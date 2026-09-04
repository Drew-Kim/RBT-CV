from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from rbtcv.condition_map import ConditionMapStore
from rbtcv.dataset import TrialVideo
from rbtcv.results_workbook import ResultsWorkbook, TAIL_CURVATURE_SHEET
from rbtcv.tail_curvature import TailCurvatureFrameRecord
from rbtcv.tail_curvature_plot import (
    TailCurvaturePlotStore,
    TrialTailCurvatureWithinTrialVariation,
)


def trial_video(cage: str, animal: str, trial: int) -> TrialVideo:
    return TrialVideo(
        dataset="Dataset",
        day="BL",
        group=f"C{cage}",
        subject=animal,
        trial=trial,
        date="2026-01-01",
        clock=f"1200{trial:02d}",
        path=Path(f"C{cage}_{animal}_T{trial}.avi"),
        relative_path=f"data/Dataset/BL/C{cage}_{animal}_T{trial}.avi",
    )


def curvature_record(
    video: TrialVideo,
    frame: int,
    curvature_degrees: float,
    back_paw_position_cm: float,
) -> TailCurvatureFrameRecord:
    return TailCurvatureFrameRecord(
        relative_video=video.relative_path,
        dataset=video.dataset,
        day=video.day,
        cage=video.cage_number,
        animal=video.rat_id,
        group=video.group,
        trial=video.trial,
        frame=frame,
        time_seconds=frame / 15.0,
        back_paw_position_cm=back_paw_position_cm,
        tail_curvature_degrees=curvature_degrees,
    )


class TailCurvatureWorkbookTests(unittest.TestCase):
    def test_position_controlled_within_trial_variation_weights_each_mouse_equally(self) -> None:
        """Day bars use trial/bin SDs, not a whole-run curvature SD."""
        series = (
            TrialTailCurvatureWithinTrialVariation(
                "BL", "1", "1", "SHAM", 1,
                (1.0, 4.0, None, None, None, None, None, None, None),
                (2, 2, 0, 0, 0, 0, 0, 0, 0),
            ),
            TrialTailCurvatureWithinTrialVariation(
                "BL", "1", "1", "SHAM", 2,
                (3.0, 6.0, None, None, None, None, None, None, None),
                (2, 2, 0, 0, 0, 0, 0, 0, 0),
            ),
            TrialTailCurvatureWithinTrialVariation(
                "BL", "1", "2", "SHAM", 1,
                (5.0, 7.0, None, None, None, None, None, None, None),
                (2, 2, 0, 0, 0, 0, 0, 0, 0),
            ),
        )

        mice, groups = TailCurvaturePlotStore._within_trial_group_summaries(series)
        self.assertEqual(mice[0].bin_means[:2], (2.0, 5.0))
        self.assertEqual(mice[1].bin_means[:2], (5.0, 7.0))
        # Mouse 1 contributes (2 + 5) / 2 = 3.5 and mouse 2 contributes
        # (5 + 7) / 2 = 6.0. Trial count therefore cannot overweight mouse 1.
        self.assertEqual(groups[0].means[:2], (3.5, 6.0))
        day_summary = TailCurvaturePlotStore._day_within_trial_variation_summaries(series)
        self.assertEqual(day_summary[0].means["BL"], 4.75)
        self.assertAlmostEqual(day_summary[0].standard_deviations["BL"], 2.5 / (2 ** 0.5))

    def test_wide_curvature_export_replaces_stale_frames_refreshes_groups_and_writes_charts(self) -> None:
        """Curvature export drives comparison and within-trial variation charts."""
        videos = tuple(
            trial_video(cage, animal, trial)
            for cage, animal in (("1", "1"), ("1", "2"), ("2", "1"), ("2", "2"))
            for trial in (1, 2)
        )
        records = tuple(
            record
            for index, video in enumerate(videos)
            for record in (
                curvature_record(video, 0, 10.12349 + index, 5.12349),
                curvature_record(video, 1, 20.12349 + index, 15.12349),
                curvature_record(video, 2, 12.12349 + index, 5.62349),
            )
        )

        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            conditions = ConditionMapStore(root)
            conditions.update_many("Dataset", (("1", "1"), ("1", "2")), "SHAM")
            conditions.update_many("Dataset", (("2", "1"), ("2", "2")), "STROKE")
            store = ResultsWorkbook(root)
            store.save_tail_curvature_measurements(videos, records)

            # Re-analysis of one source trial owns its complete row: a previous
            # later-frame curvature must not survive when that frame is absent.
            replacement = curvature_record(videos[0], 0, 99.87654, 5.0)
            replacement_same_bin = curvature_record(videos[0], 2, 97.87654, 5.5)
            store.save_tail_curvature_measurements(
                (videos[0],),
                (replacement, replacement_same_bin),
            )

            workbook = load_workbook(store.path_for_dataset("Dataset"))
            sheet = workbook[TAIL_CURVATURE_SHEET]
            headers = [cell.value for cell in sheet[1]]
            self.assertEqual(
                headers[:6],
                ["Source video", "Day", "Cage", "Animal", "Group", "Trial"],
            )
            self.assertEqual(sheet.freeze_panes, "B2")
            self.assertEqual(sheet.auto_filter.ref, "A1:I9")
            source_index = headers.index("Source video")
            group_index = headers.index("Group")
            frame_zero = headers.index("Frame 0")
            frame_one = headers.index("Frame 1")
            rows = {row[source_index]: row for row in sheet.iter_rows(min_row=2, values_only=True)}
            self.assertEqual(rows[videos[0].relative_path][group_index], "SHAM")
            self.assertEqual(rows[videos[-1].relative_path][group_index], "STROKE")
            self.assertEqual(rows[videos[0].relative_path][frame_zero], "99.877 deg; 5.000 cm")
            self.assertIsNone(rows[videos[0].relative_path][frame_one])
            self.assertEqual(rows[videos[1].relative_path][frame_zero], "11.123 deg; 5.123 cm")
            workbook.close()

            # The condition map remains authoritative for retained rows and
            # refreshes both the sheet and its charts without re-analysis.
            conditions.update_many("Dataset", (("1", "1"),), "STROKE")
            store.refresh_tail_curvature_plots("Dataset")
            workbook = load_workbook(store.path_for_dataset("Dataset"))
            refreshed_rows = {
                row[source_index]: row
                for row in workbook[TAIL_CURVATURE_SHEET].iter_rows(min_row=2, values_only=True)
            }
            self.assertEqual(refreshed_rows[videos[0].relative_path][group_index], "STROKE")
            workbook.close()

            chart_dir = root / "Dataset Results" / "Tail Curve"
            expected_charts = (
                chart_dir / "tail_curvature_group_comparison_BL.svg",
                chart_dir / "tail_curvature_within_trial_variation_BL.svg",
                chart_dir / "Day_tail_curvature_group_comparison.svg",
                chart_dir / "Day_tail_curvature_within_trial_variation.svg",
            )
            for chart in expected_charts:
                self.assertTrue(chart.exists(), chart)
            # The per-day charts retain data-driven group ribbons; day bars
            # use a neutral-gray uncertainty range with a visible whisker.
            self.assertIn(
                'fill-opacity="0.18"',
                expected_charts[0].read_text(encoding="utf-8"),
            )
            self.assertIn(
                "Each thin line is one trial",
                expected_charts[1].read_text(encoding="utf-8"),
            )
            for chart in expected_charts[2:]:
                contents = chart.read_text(encoding="utf-8")
                self.assertIn('fill="#6B7280" fill-opacity="0.24"', contents)
                self.assertIn('stroke="#374151"', contents)
            self.assertIn(
                "Tail-curvature group comparison",
                expected_charts[0].read_text(encoding="utf-8"),
            )
            self.assertIn(
                "Lower SD means more consistent posture within a run",
                expected_charts[3].read_text(encoding="utf-8"),
            )
            self.assertFalse((chart_dir / "tail_curvature_trial_consistency_BL.svg").exists())
            self.assertFalse((chart_dir / "Day_tail_curvature_trial_consistency.svg").exists())


if __name__ == "__main__":
    unittest.main()
