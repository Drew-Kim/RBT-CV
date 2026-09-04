from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from rbtcv.annotations import TrialAnnotation
from rbtcv.back_front_paw_distance import BackFrontPawDistanceFrameRecord
from rbtcv.condition_map import ConditionMapStore
from rbtcv.dataset import TrialVideo
from rbtcv.research_angle import TailAngleFrameRecord
from rbtcv.results_workbook import (
    ANGLE_CONSISTENCY_SHEET,
    BACK_FRONT_PAW_DISTANCE_SHEET,
    LEGACY_ANGLE_HEADERS,
    ResultsWorkbook,
    TICK_INTERVAL_HEADERS,
    TICK_INTERVAL_SHEET,
)
from rbtcv.tick_intervals import TickIntervalRecord
from rbtcv.tail_angle_group_plot import (
    TailAngleGroupPlotStore,
    TrialAngleSeries,
    summarize_day_by_day_group_angles,
    summarize_day_by_day_tail_angle_within_trial_variation,
    summarize_group_angles,
    summarize_tail_angle_within_trial_variation,
)


def annotation(
    *,
    cage: str,
    subject: str,
    trial: int,
    crossing_time: float,
    distance_cm: int,
    day: str = "BL",
) -> TrialAnnotation:
    return TrialAnnotation(
        relative_video=f"data/Dataset/{day}/{cage}_{subject}_T{trial}.avi",
        dataset="Dataset",
        day=day,
        group=cage,
        subject=subject,
        trial=trial,
        fps=15.0,
        start_frame=10,
        start_time=10 / 15,
        start_x=100,
        start_y=50,
        stop_frame=40,
        stop_time=40 / 15,
        stop_x=300,
        stop_y=50,
        crossing_time=crossing_time,
        outcome="reached",
        distance_cm=distance_cm,
        max_time_applied="no",
        saved_at="2026-08-15T12:00:00",
    )


class ResultsWorkbookTests(unittest.TestCase):
    def test_batch_upsert_keeps_tables_ordered_and_rebuilds_speed(self) -> None:
        with tempfile.TemporaryDirectory() as folder:
            store = ResultsWorkbook(Path(folder))
            path = store.save_many(
                (
                    annotation(cage="C2", subject="1", trial=1, crossing_time=12, distance_cm=120),
                    annotation(cage="C1", subject="2", trial=2, crossing_time=10, distance_cm=100),
                    annotation(cage="C1", subject="1", trial=1, crossing_time=8, distance_cm=120),
                )
            )["Dataset"]

            workbook = load_workbook(path)
            sheet = workbook["Forelimb"]
            self.assertEqual(
                [(sheet.cell(row=row, column=1).value, sheet.cell(row=row, column=2).value) for row in range(4, 7)],
                [(1, 1), (1, 2), (2, 1)],
            )
            self.assertEqual(sheet["A30"].value, "SPEED (cm/s)")
            self.assertEqual(sheet["C33"].value, 15.0)  # C1_1: 120 cm / 8 sec
            self.assertEqual(sheet["D34"].value, 10.0)  # C1_2: 100 cm / 10 sec

            updated = annotation(
                cage="C1",
                subject="2",
                trial=2,
                crossing_time=20,
                distance_cm=100,
            )
            store.save(updated)

            workbook = load_workbook(path)
            sheet = workbook["Forelimb"]
            self.assertEqual(sheet["D34"].value, 5.0)
            self.assertEqual(
                sum(
                    sheet.cell(row=row, column=1).value == "SPEED (cm/s)"
                    for row in range(1, sheet.max_row + 1)
                ),
                1,
            )

    def test_forelimb_days_are_chronological_left_to_right_after_analysis(self) -> None:
        with tempfile.TemporaryDirectory() as folder:
            store = ResultsWorkbook(Path(folder))
            path = store.save_many(
                (
                    annotation(cage="C1", subject="1", trial=1, crossing_time=8, distance_cm=120, day="D8"),
                    annotation(cage="C1", subject="1", trial=1, crossing_time=10, distance_cm=100, day="BL"),
                    annotation(cage="C1", subject="1", trial=2, crossing_time=20, distance_cm=110, day="D3"),
                )
            )["Dataset"]

            sheet = load_workbook(path)["Forelimb"]
            self.assertEqual([sheet.cell(row=2, column=column).value for column in (3, 6, 9)], ["BL", "D3", "D8"])
            self.assertEqual([sheet.cell(row=3, column=column).value for column in range(3, 12)], ["T1", "T2", "T3"] * 3)
            self.assertEqual(sheet.cell(row=4, column=3).value, 10)
            self.assertEqual(sheet.cell(row=4, column=7).value, 20)
            self.assertEqual(sheet.cell(row=4, column=9).value, 8)
            self.assertEqual([sheet.cell(row=2, column=column).value for column in (28, 31, 34)], ["BL", "D3", "D8"])
            self.assertEqual(sheet.cell(row=4, column=28).value, 100)
            self.assertEqual(sheet.cell(row=4, column=32).value, 110)
            self.assertEqual(sheet.cell(row=4, column=34).value, 120)

    def test_zero_distance_fall_records_zero_speed(self) -> None:
        with tempfile.TemporaryDirectory() as folder:
            store = ResultsWorkbook(Path(folder))
            path = store.save(
                annotation(
                    cage="C1",
                    subject="1",
                    trial=1,
                    crossing_time=60,
                    distance_cm=0,
                )
            )

            workbook = load_workbook(path)
            self.assertEqual(workbook["Forelimb"]["C33"].value, 0)

    def test_speed_table_moves_below_a_long_time_table(self) -> None:
        with tempfile.TemporaryDirectory() as folder:
            store = ResultsWorkbook(Path(folder))
            store.save_many(
                annotation(
                    cage=f"C{cage}",
                    subject="1",
                    trial=1,
                    crossing_time=10,
                    distance_cm=120,
                )
                for cage in range(1, 29)
            )

            workbook = load_workbook(store.path_for_dataset("Dataset"))
            sheet = workbook["Forelimb"]
            self.assertEqual(sheet["A34"].value, "SPEED (cm/s)")

    def test_tail_angle_sheet_is_wide_and_reanalysis_replaces_one_video(self) -> None:
        first_video = video(1)
        second_video = video(2)
        first_row = TailAngleFrameRecord(
            relative_video=first_video.relative_path,
            dataset="Dataset",
            day="D3",
            cage="1",
            animal="1",
            group="C1",
            trial=1,
            frame=20,
            time_seconds=20 / 15,
            back_paw_position_cm=24.5,
            signed_tail_angle_degrees=12.25,
        )
        second_row = TailAngleFrameRecord(
            relative_video=second_video.relative_path,
            dataset="Dataset",
            day="D3",
            cage="1",
            animal="1",
            group="C1",
            trial=2,
            frame=30,
            time_seconds=2.0,
            back_paw_position_cm=35.56789,
            signed_tail_angle_degrees=-8.56789,
        )
        replacement = TailAngleFrameRecord(
            relative_video=first_video.relative_path,
            dataset="Dataset",
            day="D3",
            cage="1",
            animal="1",
            group="C1",
            trial=1,
            frame=21,
            time_seconds=1.4,
            back_paw_position_cm=25.12349,
            signed_tail_angle_degrees=4.56789,
        )

        with tempfile.TemporaryDirectory() as folder:
            store = ResultsWorkbook(Path(folder))
            store.save_tail_angle_measurements((first_video, second_video), (first_row, second_row))
            store.save_tail_angle_measurements((first_video,), (replacement,))

            workbook = load_workbook(store.path_for_dataset("Dataset"))
            sheet = workbook["Frame Angles"]
            self.assertEqual(sheet.freeze_panes, "B2")
            self.assertEqual(sheet.auto_filter.ref, "A1:AK3")
            headers = [cell.value for cell in sheet[1]]
            self.assertEqual(headers[:6], ["Source video", "Day", "Cage", "Animal", "Group", "Trial"])
            self.assertNotIn("Dataset", headers)
            rows = list(sheet.iter_rows(min_row=2, values_only=True))
            self.assertEqual(len(rows), 2)
            by_source = {row[0]: row for row in rows}
            first_values = by_source[first_video.relative_path]
            second_values = by_source[second_video.relative_path]
            frame_21 = headers.index("Frame 21")
            frame_30 = headers.index("Frame 30")
            group_column = headers.index("Group")
            self.assertEqual(first_values[group_column], "Unassigned")
            self.assertEqual(first_values[frame_21], "+4.568 deg; 25.123 cm")
            self.assertEqual(second_values[frame_30], "-8.568 deg; 35.568 cm")

    def test_tail_angle_group_column_uses_saved_sham_stroke_label(self) -> None:
        trial_video = video(1)
        record = TailAngleFrameRecord(
            relative_video=trial_video.relative_path,
            dataset="Dataset",
            day="D3",
            cage="1",
            animal="1",
            group="C1",
            trial=1,
            frame=0,
            time_seconds=0.0,
            back_paw_position_cm=5.0,
            signed_tail_angle_degrees=12.0,
        )

        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            conditions = ConditionMapStore(root)
            conditions.update_many("Dataset", [("1", "1")], "SHAM")
            store = ResultsWorkbook(root)
            store.save_tail_angle_measurements((trial_video,), (record,))

            workbook = load_workbook(store.path_for_dataset("Dataset"))
            sheet = workbook["Frame Angles"]
            group_column = [cell.value for cell in sheet[1]].index("Group") + 1
            self.assertEqual(sheet.cell(row=2, column=group_column).value, "SHAM")
            workbook.close()

            conditions.update_many("Dataset", [("1", "1")], "STROKE")
            store.refresh_tail_angle_consistency("Dataset")
            workbook = load_workbook(store.path_for_dataset("Dataset"))
            self.assertEqual(workbook["Frame Angles"].cell(row=2, column=group_column).value, "STROKE")
            workbook.close()

    def test_tail_angle_sheet_migrates_existing_long_format(self) -> None:
        retained_video = video(2)
        replacement_video = video(1)
        legacy_record = TailAngleFrameRecord(
            relative_video=retained_video.relative_path,
            dataset="Dataset",
            day="D3",
            cage="1",
            animal="1",
            group="C1",
            trial=2,
            frame=30,
            time_seconds=2.0,
            back_paw_position_cm=35.5,
            signed_tail_angle_degrees=-8.5,
        )
        replacement = TailAngleFrameRecord(
            relative_video=replacement_video.relative_path,
            dataset="Dataset",
            day="D3",
            cage="1",
            animal="1",
            group="C1",
            trial=1,
            frame=21,
            time_seconds=1.4,
            back_paw_position_cm=25.0,
            signed_tail_angle_degrees=4.5,
        )

        with tempfile.TemporaryDirectory() as folder:
            store = ResultsWorkbook(Path(folder))
            path = store.path_for_dataset("Dataset")
            path.parent.mkdir(parents=True)
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Frame Angles"
            sheet.append(LEGACY_ANGLE_HEADERS)
            sheet.append(
                (
                    legacy_record.dataset,
                    legacy_record.day,
                    legacy_record.cage,
                    legacy_record.animal,
                    legacy_record.group,
                    legacy_record.trial,
                    legacy_record.frame,
                    legacy_record.time_seconds,
                    legacy_record.back_paw_position_cm,
                    legacy_record.signed_tail_angle_degrees,
                    legacy_record.relative_video,
                )
            )
            workbook.save(path)

            store.save_tail_angle_measurements((replacement_video,), (replacement,))

            migrated = load_workbook(path)["Frame Angles"]
            headers = [cell.value for cell in migrated[1]]
            rows = {row[0]: row for row in migrated.iter_rows(min_row=2, values_only=True)}
            retained = rows[retained_video.relative_path]
            self.assertEqual(retained[headers.index("Frame 30")], "-8.500 deg; 35.500 cm")

    def test_tail_angle_sheet_migrates_paired_frame_columns(self) -> None:
        retained_video = video(2)
        replacement_video = video(1)
        replacement = TailAngleFrameRecord(
            relative_video=replacement_video.relative_path,
            dataset="Dataset",
            day="D3",
            cage="1",
            animal="1",
            group="C1",
            trial=1,
            frame=21,
            time_seconds=1.4,
            back_paw_position_cm=25.0,
            signed_tail_angle_degrees=4.5,
        )

        with tempfile.TemporaryDirectory() as folder:
            store = ResultsWorkbook(Path(folder))
            path = store.path_for_dataset("Dataset")
            path.parent.mkdir(parents=True)
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Frame Angles"
            sheet.append(
                (
                    "Source video",
                    "Day",
                    "Cage",
                    "Animal",
                    "Group",
                    "Trial",
                    "Frame 30 angle (deg)",
                    "Frame 30 back paw (cm)",
                )
            )
            sheet.append((retained_video.relative_path, "D3", "1", "1", "C1", 2, -8.5, 35.5))
            workbook.save(path)

            store.save_tail_angle_measurements((replacement_video,), (replacement,))

            migrated = load_workbook(path)["Frame Angles"]
            headers = [cell.value for cell in migrated[1]]
            rows = {row[0]: row for row in migrated.iter_rows(min_row=2, values_only=True)}
            retained = rows[retained_video.relative_path]
            self.assertEqual(retained[headers.index("Frame 30")], "-8.500 deg; 35.500 cm")

    def test_tail_angle_refresh_creates_comparison_and_within_trial_variation_charts(self) -> None:
        """The legacy consistency worksheet remains inspectable, but no chart is generated from it."""
        videos = (video(1), video(2), video(3))
        records = tuple(
            TailAngleFrameRecord(
                relative_video=trial_video.relative_path,
                dataset="Dataset",
                day="D3",
                cage="1",
                animal="1",
                group="C1",
                trial=trial_video.trial,
                frame=15,
                time_seconds=1.0,
                back_paw_position_cm=5.0,
                signed_tail_angle_degrees=angle,
            )
            for trial_video, angle in zip(videos, (10.0, 14.0, 16.0))
        )

        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            ConditionMapStore(root).update_many("Dataset", [("1", "1")], "SHAM")
            store = ResultsWorkbook(root)
            store.save_tail_angle_measurements(videos, records)

            sheet = load_workbook(store.path_for_dataset("Dataset"))[ANGLE_CONSISTENCY_SHEET]
            self.assertEqual(sheet.freeze_panes, "A2")
            headers = [cell.value for cell in sheet[1]]
            values = next(sheet.iter_rows(min_row=2, values_only=True))
            self.assertEqual(values[headers.index("Condition")], "SHAM")
            self.assertEqual(values[headers.index("0-10 cm trial n")], 3)
            self.assertEqual(values[headers.index("0-10 cm SD (deg)")], 3.055)
            self.assertEqual(values[headers.index("Overall 0-90 cm trial n")], 3)
            self.assertEqual(values[headers.index("Overall 0-90 cm SD (deg)")], 3.055)
            self.assertTrue(
                (root / "Dataset Results" / "Tail Angle" / "tail_angle_group_comparison_D3.svg").exists()
            )
            self.assertTrue(
                (root / "Dataset Results" / "Tail Angle" / "Day_tail_angle_group_comparison.svg").exists()
            )
            self.assertTrue(
                (root / "Dataset Results" / "Tail Angle" / "tail_angle_within_trial_variation_D3.svg").exists()
            )
            self.assertTrue(
                (root / "Dataset Results" / "Tail Angle" / "Day_tail_angle_within_trial_variation.svg").exists()
            )
            self.assertFalse(
                (root / "Dataset Results" / "Tail Angle" / "tail_angle_trial_consistency_D3.svg").exists()
            )
            self.assertFalse(
                (root / "Dataset Results" / "Tail Angle" / "Day_tail_angle_trial_consistency.svg").exists()
            )

    def test_tick_interval_sheet_is_wide_labeled_and_creates_all_comparison_charts(self) -> None:
        """Interval times stay numeric, replace one trial, and drive all four charts."""
        videos = tuple(
            interval_video("BL", cage, animal, trial)
            for cage, animal in (("1", "1"), ("1", "2"), ("2", "1"), ("2", "2"))
            for trial in (1, 2)
        )
        elapsed_times = (1.23456, 1.60001, 2.30001, 2.90001, 4.10001, 5.20001, 6.50001, 8.40001)
        records = tuple(
            record
            for trial_video, elapsed in zip(videos, elapsed_times)
            for record in (
                interval_record(trial_video, 0, elapsed),
                interval_record(trial_video, 10, elapsed + 0.11111),
            )
        )

        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            conditions = ConditionMapStore(root)
            conditions.update_many("Dataset", (("1", "1"), ("1", "2")), "SHAM")
            conditions.update_many("Dataset", (("2", "1"), ("2", "2")), "STROKE")
            store = ResultsWorkbook(root)
            store.save_tick_interval_measurements(videos, records)

            # A recalculation replaces a source trial's wide row rather than
            # appending another row or retaining stale later intervals.
            replacement = interval_record(videos[0], 0, 9.87654)
            store.save_tick_interval_measurements((videos[0],), (replacement,))

            workbook = load_workbook(store.path_for_dataset("Dataset"))
            sheet = workbook[TICK_INTERVAL_SHEET]
            headers = [cell.value for cell in sheet[1]]
            self.assertEqual(tuple(headers), TICK_INTERVAL_HEADERS)
            self.assertEqual(sheet.freeze_panes, "B2")
            self.assertEqual(sheet.max_row, len(videos) + 1)
            self.assertEqual(sheet.auto_filter.ref, "A1:O9")

            source_index = headers.index("Source video")
            group_index = headers.index("Group")
            first_interval_index = headers.index("0-10 cm (s)")
            second_interval_index = headers.index("10-20 cm (s)")
            rows = {row[source_index]: row for row in sheet.iter_rows(min_row=2, values_only=True)}
            self.assertEqual({row[group_index] for row in rows.values()}, {"SHAM", "STROKE"})
            self.assertEqual(rows[videos[0].relative_path][group_index], "SHAM")
            self.assertEqual(rows[videos[-1].relative_path][group_index], "STROKE")
            self.assertIsInstance(rows[videos[0].relative_path][first_interval_index], float)
            self.assertEqual(rows[videos[0].relative_path][first_interval_index], 9.877)
            self.assertIsNone(rows[videos[0].relative_path][second_interval_index])
            self.assertEqual(sheet.cell(row=2, column=first_interval_index + 1).number_format, "0.000")
            workbook.close()

            chart_dir = root / "Dataset Results" / "Tick Interval"
            expected_charts = (
                chart_dir / "tick_interval_group_comparison_BL.svg",
                chart_dir / "tick_interval_within_trial_variation_BL.svg",
                chart_dir / "Day_tick_interval_group_comparison.svg",
                chart_dir / "Day_tick_interval_within_trial_variation.svg",
            )
            for chart in expected_charts:
                self.assertTrue(chart.exists(), chart)
                text = chart.read_text(encoding="utf-8")
                expected_opacity = (
                    "0.24"
                    if chart.name in {
                        "Day_tick_interval_group_comparison.svg",
                        "Day_tick_interval_within_trial_variation.svg",
                    }
                    else "0.18"
                )
                self.assertIn(f'fill-opacity="{expected_opacity}"', text)
            self.assertIn("Tick-interval elapsed-time comparison", expected_charts[0].read_text(encoding="utf-8"))
            self.assertIn("Each dot is one trial", expected_charts[1].read_text(encoding="utf-8"))
            day_group_svg = expected_charts[2].read_text(encoding="utf-8")
            self.assertIn("Colored bars are SHAM/STROKE mean elapsed times", day_group_svg)
            self.assertIn('fill="#1976D2" cursor="help"', day_group_svg)
            self.assertNotIn("<polyline", day_group_svg)
            self.assertIn("#6B7280", day_group_svg)
            self.assertIn("#374151", day_group_svg)
            self.assertIn("Lower SD means more even timing", expected_charts[3].read_text(encoding="utf-8"))
            self.assertFalse((chart_dir / "tick_interval_trial_consistency_BL.svg").exists())
            self.assertFalse((chart_dir / "Day_tick_interval_trial_consistency.svg").exists())

    def test_tail_angle_sheets_sort_days_chronologically(self) -> None:
        """All frame-level exports use BL, D3, D8 ... D30 chronological order."""
        ordered_days = ("BL", "D3", "D8", "D14", "D21", "D30")
        videos = tuple(interval_video(day, "1", "1", 1) for day in reversed(ordered_days))
        records = tuple(
            TailAngleFrameRecord(
                relative_video=trial_video.relative_path,
                dataset="Dataset",
                day=trial_video.day,
                cage="1",
                animal="1",
                group="C1",
                trial=1,
                frame=0,
                time_seconds=0.0,
                back_paw_position_cm=5.0,
                signed_tail_angle_degrees=10.0,
            )
            for trial_video in videos
        )

        with tempfile.TemporaryDirectory() as folder:
            store = ResultsWorkbook(Path(folder))
            store.save_tail_angle_measurements(videos, records)
            paw_records = tuple(
                paw_distance_record(trial_video, 0, 5.0, 5.0)
                for trial_video in videos
            )
            store.save_back_front_paw_distance_measurements(videos, paw_records)
            workbook = load_workbook(store.path_for_dataset("Dataset"), data_only=True)
            frame_days = [row[1] for row in workbook["Frame Angles"].iter_rows(min_row=2, values_only=True)]
            consistency_days = [row[0] for row in workbook[ANGLE_CONSISTENCY_SHEET].iter_rows(min_row=2, values_only=True)]
            paw_distance_days = [
                row[1]
                for row in workbook[BACK_FRONT_PAW_DISTANCE_SHEET].iter_rows(
                    min_row=2,
                    values_only=True,
                )
            ]
            workbook.close()

        self.assertEqual(frame_days, list(ordered_days))
        self.assertEqual(consistency_days, list(ordered_days))
        self.assertEqual(paw_distance_days, list(ordered_days))

    def test_back_front_paw_distance_sheet_replaces_trials_and_creates_all_charts(self) -> None:
        """Frame cells stay readable while charts use equal-mouse statistics."""
        videos = tuple(
            interval_video("BL", cage, animal, trial)
            for cage, animal in (("1", "1"), ("1", "2"), ("2", "1"), ("2", "2"))
            for trial in (1, 2)
        )
        records = tuple(
            record
            for index, trial_video in enumerate(videos)
            for record in (
                paw_distance_record(trial_video, 0, 4.12349 + index, 5.0),
                paw_distance_record(trial_video, 1, 6.12349 + index, 15.0),
            )
        )

        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            conditions = ConditionMapStore(root)
            conditions.update_many("Dataset", (("1", "1"), ("1", "2")), "SHAM")
            conditions.update_many("Dataset", (("2", "1"), ("2", "2")), "STROKE")
            store = ResultsWorkbook(root)
            store.save_back_front_paw_distance_measurements(videos, records)

            # Reanalysis of one source trial removes its old later-frame value.
            replacement = paw_distance_record(videos[0], 0, 9.87654, 5.0)
            store.save_back_front_paw_distance_measurements((videos[0],), (replacement,))

            workbook = load_workbook(store.path_for_dataset("Dataset"))
            sheet = workbook[BACK_FRONT_PAW_DISTANCE_SHEET]
            headers = [cell.value for cell in sheet[1]]
            self.assertEqual(headers[:6], ["Source video", "Day", "Cage", "Animal", "Group", "Trial"])
            self.assertEqual(sheet.freeze_panes, "B2")
            self.assertEqual(sheet.auto_filter.ref, "A1:H9")
            source_index = headers.index("Source video")
            group_index = headers.index("Group")
            frame_zero = headers.index("Frame 0")
            frame_one = headers.index("Frame 1")
            rows = {row[source_index]: row for row in sheet.iter_rows(min_row=2, values_only=True)}
            self.assertEqual(rows[videos[0].relative_path][group_index], "SHAM")
            self.assertEqual(rows[videos[-1].relative_path][group_index], "STROKE")
            self.assertEqual(rows[videos[0].relative_path][frame_zero], "9.877 cm; 5.000 cm")
            self.assertIsNone(rows[videos[0].relative_path][frame_one])
            workbook.close()

            # Relabeling reuses the condition map, including retained existing rows.
            conditions.update_many("Dataset", (("1", "1"),), "STROKE")
            store.refresh_back_front_paw_distance_plots("Dataset")
            workbook = load_workbook(store.path_for_dataset("Dataset"))
            refreshed = workbook[BACK_FRONT_PAW_DISTANCE_SHEET]
            refreshed_rows = {
                row[source_index]: row
                for row in refreshed.iter_rows(min_row=2, values_only=True)
            }
            self.assertEqual(refreshed_rows[videos[0].relative_path][group_index], "STROKE")
            workbook.close()

            chart_dir = root / "Dataset Results" / "Back_Front paw Distance"
            expected_charts = (
                chart_dir / "back_front_paw_distance_group_comparison_BL.svg",
                chart_dir / "back_front_paw_distance_within_trial_variation_BL.svg",
                chart_dir / "Day_back_front_paw_distance_group_comparison.svg",
                chart_dir / "Day_back_front_paw_distance_within_trial_variation.svg",
            )
            for chart in expected_charts:
                self.assertTrue(chart.exists(), chart)
            self.assertIn('fill-opacity="0.18"', expected_charts[0].read_text(encoding="utf-8"))
            self.assertIn(
                "Back/front paw-distance group comparison",
                expected_charts[0].read_text(encoding="utf-8"),
            )
            self.assertIn(
                "Lower SD means less within-run variation",
                expected_charts[1].read_text(encoding="utf-8"),
            )
            self.assertIn("Colored bars are SHAM/STROKE mean distances", expected_charts[2].read_text(encoding="utf-8"))
            self.assertIn("Lower SD means more consistent posture and gait", expected_charts[3].read_text(encoding="utf-8"))
            self.assertFalse(
                (chart_dir / "back_front_paw_distance_trial_consistency_BL.svg").exists()
            )
            self.assertFalse(
                (chart_dir / "Day_back_front_paw_distance_trial_consistency.svg").exists()
            )


def video(trial: int) -> TrialVideo:
    return TrialVideo(
        dataset="Dataset",
        day="D3",
        group="C1",
        subject="1",
        trial=trial,
        date="2026-08-15",
        clock=f"1200{trial:02d}",
        path=Path(f"video_{trial}.avi"),
        relative_path=f"data/Dataset/D3/C1_1_T{trial}.avi",
    )


def interval_video(day: str, cage: str, animal: str, trial: int) -> TrialVideo:
    return TrialVideo(
        dataset="Dataset",
        day=day,
        group=f"C{cage}",
        subject=animal,
        trial=trial,
        date="2026-08-15",
        clock=f"1300{trial:02d}",
        path=Path(f"interval_{day}_{cage}_{animal}_{trial}.avi"),
        relative_path=f"data/Dataset/{day}/C{cage}_{animal}_T{trial}.avi",
    )


def paw_distance_record(
    trial_video: TrialVideo,
    frame: int,
    distance_cm: float,
    back_paw_position_cm: float,
) -> BackFrontPawDistanceFrameRecord:
    return BackFrontPawDistanceFrameRecord(
        relative_video=trial_video.relative_path,
        dataset=trial_video.dataset,
        day=trial_video.day,
        cage=trial_video.cage_number,
        animal=trial_video.rat_id,
        group=trial_video.group,
        trial=trial_video.trial,
        frame=frame,
        time_seconds=frame / 15.0,
        back_paw_position_cm=back_paw_position_cm,
        back_front_paw_distance_cm=distance_cm,
    )


def interval_record(video: TrialVideo, start_cm: int, elapsed_seconds: float) -> TickIntervalRecord:
    return TickIntervalRecord(
        relative_video=video.relative_path,
        dataset=video.dataset,
        day=video.day,
        cage=video.cage_number,
        animal=video.rat_id,
        group=video.group,
        trial=video.trial,
        interval_start_cm=start_cm,
        interval_end_cm=start_cm + 10,
        start_frame=start_cm,
        end_frame=start_cm + 15,
        elapsed_seconds=elapsed_seconds,
    )


class TailAngleGroupPlotTests(unittest.TestCase):
    def test_group_summary_averages_trials_within_mouse_before_group_sd(self) -> None:
        trials = (
            TrialAngleSeries("D3", "1", "1", "STROKE", 1, (10.0,) * 9),
            TrialAngleSeries("D3", "1", "1", "STROKE", 2, (20.0,) * 9),
            TrialAngleSeries("D3", "1", "2", "STROKE", 1, (30.0,) * 9),
            TrialAngleSeries("D3", "1", "2", "STROKE", 2, (50.0,) * 9),
        )

        mice, groups = summarize_group_angles(trials)

        self.assertEqual([mouse.bin_means[0] for mouse in mice], [15.0, 40.0])
        summary = groups[0]
        self.assertEqual(summary.means[0], 27.5)
        self.assertAlmostEqual(summary.standard_deviations[0], 17.6776695297)
        self.assertEqual(summary.mouse_counts[0], 2)

        svg = TailAngleGroupPlotStore._svg("Dataset", "D3", trials, mice, groups)
        self.assertIn('fill-opacity="0.18"', svg)
        self.assertIn("Each mouse is averaged across its trials", svg)
        self.assertIn("Stroke | Cage 1 Mouse 1 | Trial T1", svg)
        self.assertIn('stroke-width="8" stroke-opacity="0" pointer-events="stroke"', svg)

    def test_day_by_day_summary_compares_mouse_mean_signed_angles(self) -> None:
        trials = (
            TrialAngleSeries("BL", "1", "1", "SHAM", 1, (10.0,) * 9),
            TrialAngleSeries("BL", "1", "1", "SHAM", 2, (20.0,) * 9),
            TrialAngleSeries("BL", "1", "2", "SHAM", 1, (30.0,) * 9),
            TrialAngleSeries("BL", "1", "2", "SHAM", 2, (50.0,) * 9),
            TrialAngleSeries("D3", "1", "1", "SHAM", 1, (30.0,) * 9),
            TrialAngleSeries("D3", "1", "1", "SHAM", 2, (40.0,) * 9),
            TrialAngleSeries("D3", "1", "2", "SHAM", 1, (50.0,) * 9),
            TrialAngleSeries("D3", "1", "2", "SHAM", 2, (70.0,) * 9),
            TrialAngleSeries("BL", "2", "1", "STROKE", 1, (-5.0,) * 9),
            TrialAngleSeries("BL", "2", "1", "STROKE", 2, (-1.0,) * 9),
            TrialAngleSeries("D3", "2", "1", "STROKE", 1, (5.0,) * 9),
            TrialAngleSeries("D3", "2", "1", "STROKE", 2, (7.0,) * 9),
        )

        summaries = {item.condition: item for item in summarize_day_by_day_group_angles(trials)}

        sham = summaries["SHAM"]
        self.assertAlmostEqual(sham.means["BL"], 27.5)
        self.assertAlmostEqual(sham.means["D3"], 47.5)
        self.assertAlmostEqual(sham.standard_deviations["BL"], 17.6776695297)
        self.assertEqual(sham.mouse_counts, {"BL": 2, "D3": 2})
        self.assertIsNone(summaries["STROKE"].standard_deviations["BL"])

        svg = TailAngleGroupPlotStore._day_by_day_svg("Dataset", summaries.values())
        self.assertIn("Day-by-day signed tail-angle comparison", svg)
        self.assertIn("Experimental day", svg)
        self.assertIn("Signed tail angle", svg)
        self.assertIn('#6B7280', svg)
        self.assertIn('#374151', svg)
        self.assertIn('+27.500°', svg)
        self.assertNotIn('<polyline', svg)

    def test_within_trial_variation_is_position_controlled_and_equal_mouse_weighted(self) -> None:
        root_two = 2.0 ** 0.5

        def frame_bins(*standard_deviations: float) -> tuple[tuple[float, ...], ...]:
            return tuple(
                (0.0, standard_deviations[index] * root_two)
                if index < len(standard_deviations)
                else ()
                for index in range(9)
            )

        trials = (
            TrialAngleSeries("BL", "1", "1", "SHAM", 1, (0.0,) * 9, frame_bins(1.0, 3.0)),
            TrialAngleSeries("BL", "1", "1", "SHAM", 2, (0.0,) * 9, frame_bins(5.0)),
            TrialAngleSeries("BL", "1", "2", "SHAM", 1, (0.0,) * 9, frame_bins(7.0)),
            TrialAngleSeries("BL", "2", "1", "STROKE", 1, (0.0,) * 9, frame_bins(2.0, 4.0)),
        )

        trial_values, mice, groups = summarize_tail_angle_within_trial_variation(trials)
        self.assertAlmostEqual(trial_values[0].mean_standard_deviation_degrees, 2.0)
        sham_mouse = next(mouse for mouse in mice if mouse.cage == "1" and mouse.animal == "1")
        # The three trial/bin SDs (1, 3, and 5) receive equal weight.
        self.assertAlmostEqual(sham_mouse.mean_standard_deviation_degrees, 3.0)
        sham_group = next(group for group in groups if group.condition == "SHAM")
        self.assertAlmostEqual(sham_group.mean_standard_deviation_degrees, 5.0)
        self.assertAlmostEqual(sham_group.standard_deviation_degrees or 0.0, 2.0 ** 1.5)

        summaries = {
            item.condition: item
            for item in summarize_day_by_day_tail_angle_within_trial_variation(trials)
        }
        self.assertAlmostEqual(summaries["SHAM"].means["BL"], 5.0)
        self.assertAlmostEqual(summaries["SHAM"].standard_deviations["BL"] or 0.0, 2.0 ** 1.5)
        self.assertEqual(summaries["SHAM"].mouse_counts, {"BL": 2})
        self.assertIsNone(summaries["STROKE"].standard_deviations["BL"])

        svg = TailAngleGroupPlotStore._day_by_day_within_trial_variation_svg(
            "Dataset", summaries.values()
        )
        self.assertIn("Day-by-day tail-angle within-trial variation", svg)
        self.assertIn("Lower SD means more stable position-controlled tail posture", svg)
        self.assertIn("Within-trial tail-angle SD", svg)
        self.assertIn('#6B7280', svg)
        self.assertIn('#374151', svg)
        self.assertIn('5.000°', svg)
        self.assertNotIn('<polyline', svg)

if __name__ == "__main__":
    unittest.main()
