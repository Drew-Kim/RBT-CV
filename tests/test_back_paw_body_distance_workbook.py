from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from rbtcv.back_paw_body_distance import BackPawBodyDistanceFrameRecord
from rbtcv.back_paw_body_distance_plot import BackPawBodyDistancePlotStore
from rbtcv.condition_map import ConditionMapStore
from rbtcv.dataset import TrialVideo
from rbtcv.results_workbook import BACK_PAW_BODY_DISTANCE_SHEET, ResultsWorkbook


def trial_video(day: str, cage: str, animal: str, trial: int) -> TrialVideo:
    return TrialVideo(
        dataset="Dataset",
        day=day,
        group=f"C{cage}",
        subject=animal,
        trial=trial,
        date="2026-01-01",
        clock=f"1200{trial:02d}",
        path=Path(f"{day}_C{cage}_{animal}_T{trial}.avi"),
        relative_path=f"data/Dataset/{day}/C{cage}_{animal}_T{trial}.avi",
    )


def distance_record(
    video: TrialVideo,
    frame: int,
    distance_cm: float,
    back_paw_position_cm: float,
) -> BackPawBodyDistanceFrameRecord:
    return BackPawBodyDistanceFrameRecord(
        relative_video=video.relative_path,
        dataset=video.dataset,
        day=video.day,
        cage=video.cage_number,
        animal=video.rat_id,
        group=video.group,
        trial=video.trial,
        frame=frame,
        time_seconds=frame / 15.0,
        back_paw_position_cm=back_paw_position_cm,
        back_paw_body_distance_cm=distance_cm,
    )


class BackPawBodyDistanceWorkbookTests(unittest.TestCase):
    def test_wide_export_replaces_stale_frames_refreshes_groups_and_writes_charts(self) -> None:
        videos = tuple(
            trial_video("BL", cage, animal, trial)
            for cage, animal in (("1", "1"), ("1", "2"), ("2", "1"), ("2", "2"))
            for trial in (1, 2)
        )
        records = tuple(
            record
            for index, video in enumerate(videos)
            for record in (
                distance_record(video, 0, 4.12349 + index, 5.12349),
                distance_record(video, 1, 6.12349 + index, 15.12349),
            )
        )

        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            conditions = ConditionMapStore(root)
            conditions.update_many("Dataset", (("1", "1"), ("1", "2")), "SHAM")
            conditions.update_many("Dataset", (("2", "1"), ("2", "2")), "STROKE")
            store = ResultsWorkbook(root)
            store.save_back_paw_body_distance_measurements(videos, records)

            # Reanalysis owns all frame cells for the supplied trial.
            replacement = distance_record(videos[0], 0, 9.87654, 5.0)
            store.save_back_paw_body_distance_measurements((videos[0],), (replacement,))

            workbook = load_workbook(store.path_for_dataset("Dataset"))
            sheet = workbook[BACK_PAW_BODY_DISTANCE_SHEET]
            headers = [cell.value for cell in sheet[1]]
            self.assertEqual(
                headers[:6],
                ["Source video", "Day", "Cage", "Animal", "Group", "Trial"],
            )
            self.assertEqual(sheet.freeze_panes, "B2")
            self.assertEqual(sheet.auto_filter.ref, "A1:H9")
            source_index = headers.index("Source video")
            group_index = headers.index("Group")
            frame_zero = headers.index("Frame 0")
            frame_one = headers.index("Frame 1")
            rows = {row[source_index]: row for row in sheet.iter_rows(min_row=2, values_only=True)}
            self.assertEqual(rows[videos[0].relative_path][group_index], "SHAM")
            self.assertEqual(rows[videos[-1].relative_path][group_index], "STROKE")
            self.assertEqual(rows[videos[0].relative_path][frame_zero], "9.877 cm; 5.000 cm")
            self.assertIsNone(rows[videos[0].relative_path][frame_one])
            workbook.close()

            # A later GUI group relabel refreshes existing rows and charts.
            conditions.update_many("Dataset", (("1", "1"),), "STROKE")
            store.refresh_back_paw_body_distance_plots("Dataset")
            workbook = load_workbook(store.path_for_dataset("Dataset"))
            refreshed_rows = {
                row[source_index]: row
                for row in workbook[BACK_PAW_BODY_DISTANCE_SHEET].iter_rows(
                    min_row=2,
                    values_only=True,
                )
            }
            self.assertEqual(refreshed_rows[videos[0].relative_path][group_index], "STROKE")
            workbook.close()

            chart_dir = root / "Dataset Results" / "Back_paw Body Distance"
            expected_charts = (
                chart_dir / "back_paw_body_distance_group_comparison_BL.svg",
                chart_dir / "back_paw_body_distance_within_trial_variation_BL.svg",
                chart_dir / "Day_back_paw_body_distance_group_comparison.svg",
                chart_dir / "Day_back_paw_body_distance_within_trial_variation.svg",
            )
            for chart in expected_charts:
                self.assertTrue(chart.exists(), chart)
            self.assertIn('fill-opacity="0.18"', expected_charts[0].read_text(encoding="utf-8"))
            text = expected_charts[2].read_text(encoding="utf-8")
            self.assertIn('fill-opacity="0.24"', text)
            self.assertIn("#6B7280", text)
            self.assertIn("#374151", text)
            self.assertIn('fill="#1976D2" cursor="help"', text)
            self.assertFalse((chart_dir / "back_paw_body_distance_trial_consistency_BL.svg").exists())
            self.assertFalse((chart_dir / "Day_back_paw_body_distance_trial_consistency.svg").exists())

    def test_wide_sheet_uses_chronological_day_order(self) -> None:
        days = ("BL", "D3", "D8", "D14", "D21", "D30")
        videos = tuple(trial_video(day, "1", "1", 1) for day in reversed(days))
        records = tuple(distance_record(video, 0, 5.0, 5.0) for video in videos)

        with tempfile.TemporaryDirectory() as folder:
            store = ResultsWorkbook(Path(folder))
            store.save_back_paw_body_distance_measurements(
                videos,
                records,
                refresh_plot=False,
            )
            workbook = load_workbook(store.path_for_dataset("Dataset"), data_only=True)
            exported_days = [
                row[1]
                for row in workbook[BACK_PAW_BODY_DISTANCE_SHEET].iter_rows(
                    min_row=2,
                    values_only=True,
                )
            ]
            workbook.close()

        self.assertEqual(exported_days, list(days))

    def test_within_trial_variation_uses_raw_frame_sd_and_retains_trial_lines(self) -> None:
        """The new chart is about frame variation inside a trial, not trial means."""
        videos = tuple(
            trial_video("BL", "1", animal, trial)
            for animal in ("1", "2")
            for trial in (1, 2)
        )
        values = {
            ("1", 1): (2.0, 4.0),
            ("1", 2): (2.0, 4.0),
            ("2", 1): (2.0, 8.0),
            ("2", 2): (2.0, 8.0),
        }
        records = tuple(
            record
            for video in videos
            for record in (
                distance_record(video, 0, values[(video.rat_id, video.trial)][0], 5.0),
                distance_record(video, 1, values[(video.rat_id, video.trial)][1], 5.0),
            )
        )

        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            conditions = ConditionMapStore(root)
            conditions.update_many("Dataset", (("1", "1"), ("1", "2")), "SHAM")
            store = ResultsWorkbook(root)
            store.save_back_paw_body_distance_measurements(videos, records)

            plot_store = BackPawBodyDistancePlotStore(root)
            variations = plot_store._read_within_trial_variation_series(
                store.path_for_dataset("Dataset"),
                conditions.load("Dataset"),
            )
            self.assertEqual(len(variations), 4)
            self.assertAlmostEqual(variations[0].standard_deviations[0], 2**0.5)
            self.assertEqual(variations[0].frame_counts[0], 2)

            mouse_series, group_summaries = plot_store._within_trial_group_summaries(variations)
            self.assertEqual(len(mouse_series), 2)
            self.assertAlmostEqual(group_summaries[0].means[0], (2**0.5 + 3 * 2**0.5) / 2)

            svg = plot_store._daily_within_trial_variation_svg(
                "Dataset",
                "BL",
                variations,
                mouse_series,
                group_summaries,
            )
            self.assertIn("Each thin line is one trial", svg)
            self.assertIn("Trial T1 | Raw frame distance SD", svg)
            self.assertIn('fill-opacity="0.18"', svg)

            chart_dir = root / "Dataset Results" / "Back_paw Body Distance"
            self.assertTrue(
                (chart_dir / "back_paw_body_distance_within_trial_variation_BL.svg").exists()
            )
            self.assertTrue(
                (chart_dir / "Day_back_paw_body_distance_within_trial_variation.svg").exists()
            )
            day_svg = (
                chart_dir / "Day_back_paw_body_distance_within_trial_variation.svg"
            ).read_text(encoding="utf-8")
            self.assertIn("within-trial variation", day_svg)
            self.assertIn('fill-opacity="0.24"', day_svg)
            self.assertIn("#6B7280", day_svg)
            self.assertIn("#374151", day_svg)


if __name__ == "__main__":
    unittest.main()
