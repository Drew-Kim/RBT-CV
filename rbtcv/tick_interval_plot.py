"""SVG comparisons for calibrated 10 cm back-paw interval times."""

from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
import html
import math
from pathlib import Path
import statistics
from typing import Iterable

from .condition_map import ConditionMapStore
from .dataset import ROOT, natural_day_key
from .tick_intervals import INTERVAL_STARTS_CM


OUTPUT_ROOT = ROOT / "outputs"
RESULTS_FILENAME = "RBT_CV_Results.xlsx"
SHEET_NAME = "Tick Interval Times"
CHART_DIRECTORY = "Tick Interval"
LEGACY_CHART_DIRECTORY = "Tail Angle"
FILE_SUFFIX = ".svg"
GROUP_PREFIX = "tick_interval_group_comparison_"
WITHIN_TRIAL_VARIATION_PREFIX = "tick_interval_within_trial_variation_"
DAY_GROUP_FILENAME = "Day_tick_interval_group_comparison.svg"
DAY_WITHIN_TRIAL_VARIATION_FILENAME = "Day_tick_interval_within_trial_variation.svg"

SHAM_COLOR = "#1976D2"
STROKE_COLOR = "#D32F2F"
UNASSIGNED_COLOR = "#616161"
ERROR_RANGE_COLOR = "#6B7280"
ERROR_WHISKER_COLOR = "#374151"


@dataclass(frozen=True)
class TrialIntervalSeries:
    day: str
    cage: str
    animal: str
    condition: str
    trial: int
    interval_seconds: tuple[float | None, ...]


@dataclass(frozen=True)
class MouseIntervalSeries:
    day: str
    cage: str
    animal: str
    condition: str
    interval_seconds: tuple[float | None, ...]


@dataclass(frozen=True)
class GroupIntervalSummary:
    condition: str
    means: tuple[float | None, ...]
    standard_deviations: tuple[float | None, ...]
    mouse_counts: tuple[int, ...]


@dataclass(frozen=True)
class TrialIntervalWithinTrialVariation:
    """One run's timing evenness across its available 10 cm intervals."""

    day: str
    cage: str
    animal: str
    condition: str
    trial: int
    standard_deviation_seconds: float
    interval_count: int


@dataclass(frozen=True)
class MouseIntervalWithinTrialVariation:
    """A mouse's mean within-trial timing SD for one experimental day."""

    day: str
    cage: str
    animal: str
    condition: str
    mean_standard_deviation_seconds: float
    trial_count: int


@dataclass(frozen=True)
class GroupIntervalWithinTrialVariation:
    """Equal-mouse group comparison of within-trial timing variation."""

    condition: str
    mean_standard_deviation_seconds: float
    standard_deviation_seconds: float | None
    mouse_count: int


@dataclass(frozen=True)
class DayGroupSummary:
    condition: str
    means: dict[str, float]
    standard_deviations: dict[str, float | None]
    mouse_counts: dict[str, int]


class TickIntervalPlotStore:
    """Create elapsed-time and within-trial-variation charts from interval data."""

    def __init__(self, output_root: Path = OUTPUT_ROOT) -> None:
        self.output_root = output_root

    def result_dir(self, dataset: str) -> Path:
        return self.output_root / f"{dataset} Results"

    def chart_dir(self, dataset: str) -> Path:
        return self.result_dir(dataset) / CHART_DIRECTORY

    def refresh_dataset(self, dataset: str) -> dict[str, Path]:
        workbook_path = self.result_dir(dataset) / RESULTS_FILENAME
        if not workbook_path.exists():
            return {}

        conditions = ConditionMapStore(self.output_root).load(dataset)
        trials = self._read_trial_series(workbook_path, conditions)
        by_day: dict[str, list[TrialIntervalSeries]] = defaultdict(list)
        for series in trials:
            by_day[series.day].append(series)

        chart_dir = self.chart_dir(dataset)
        chart_dir.mkdir(parents=True, exist_ok=True)
        paths: dict[str, Path] = {}
        for day, day_trials in sorted(by_day.items(), key=lambda item: natural_day_key(item[0])):
            mice, groups = summarize_group_interval_times(day_trials)
            group_path = chart_dir / f"{GROUP_PREFIX}{day}{FILE_SUFFIX}"
            self._write_svg(group_path, self._daily_group_svg(dataset, day, day_trials, mice, groups))
            paths[f"group:{day}"] = group_path

            variation_trials, variation_mice, variation_groups = (
                summarize_interval_within_trial_variation(day_trials)
            )
            variation_path = chart_dir / f"{WITHIN_TRIAL_VARIATION_PREFIX}{day}{FILE_SUFFIX}"
            self._write_svg(
                variation_path,
                self._daily_within_trial_variation_svg(
                    dataset,
                    day,
                    variation_trials,
                    variation_groups,
                ),
            )
            paths[f"within_trial_variation:{day}"] = variation_path

        day_group_path = chart_dir / DAY_GROUP_FILENAME
        self._write_svg(day_group_path, self._day_bar_svg(
            dataset,
            summarize_day_by_day_interval_times(trials),
            title="Day-by-day tick-interval elapsed-time comparison",
            calculation_note=(
                "For each day, each mouse is averaged across its trials and valid "
                "0-10 through 80-90 cm intervals before group calculation."
            ),
            comparison_note=(
                "Colored bars are SHAM/STROKE mean elapsed times; neutral-gray ranges and "
                "whiskers are plus/minus 1 sample SD across mice."
            ),
            empty_note="No SHAM or STROKE tick-interval data are available yet.",
            y_label="Elapsed time per 10 cm interval (s)",
        ))
        paths["day_group"] = day_group_path

        day_variation_path = chart_dir / DAY_WITHIN_TRIAL_VARIATION_FILENAME
        self._write_svg(day_variation_path, self._day_within_trial_variation_bar_svg(
            dataset,
            summarize_day_by_day_interval_within_trial_variation(trials),
        ))
        paths["day_within_trial_variation"] = day_variation_path

        live_names = {path.name for path in paths.values()}
        for prefix in (GROUP_PREFIX, WITHIN_TRIAL_VARIATION_PREFIX):
            for stale_path in chart_dir.glob(f"{prefix}*{FILE_SUFFIX}"):
                if stale_path.name not in live_names:
                    stale_path.unlink()
        # The removed trial-to-trial consistency chart may remain from an
        # earlier version; remove only that known generated filename pattern.
        for obsolete_path in chart_dir.glob(f"tick_interval_trial_consistency_*{FILE_SUFFIX}"):
            obsolete_path.unlink()
        obsolete_day_consistency = chart_dir / "Day_tick_interval_trial_consistency.svg"
        if obsolete_day_consistency.exists():
            obsolete_day_consistency.unlink()
        # Interval charts used the Tail Angle folder before they received their
        # own output location. Remove only these generated interval filenames;
        # never touch the tail-angle charts that still belong there.
        legacy_dir = self.result_dir(dataset) / LEGACY_CHART_DIRECTORY
        if legacy_dir.exists():
            legacy_names = (
                f"{GROUP_PREFIX}*{FILE_SUFFIX}",
                f"{WITHIN_TRIAL_VARIATION_PREFIX}*{FILE_SUFFIX}",
                DAY_GROUP_FILENAME,
                DAY_WITHIN_TRIAL_VARIATION_FILENAME,
                f"tick_interval_trial_consistency_*{FILE_SUFFIX}",
                "Day_tick_interval_trial_consistency.svg",
            )
            for pattern in legacy_names:
                for legacy_path in legacy_dir.glob(pattern):
                    legacy_path.unlink()
        return paths

    @staticmethod
    def _read_trial_series(
        workbook_path: Path,
        conditions: dict[tuple[str, str], str],
    ) -> list[TrialIntervalSeries]:
        from openpyxl import load_workbook

        workbook = load_workbook(workbook_path, data_only=True, read_only=True)
        try:
            if SHEET_NAME not in workbook.sheetnames:
                return []
            sheet = workbook[SHEET_NAME]
            headers = [cell.value for cell in next(sheet.iter_rows(max_row=1), ())]
            index = {str(header): position for position, header in enumerate(headers) if header is not None}
            required = ("Day", "Cage", "Animal", "Trial")
            if not all(header in index for header in required):
                return []
            output: list[TrialIntervalSeries] = []
            for values in sheet.iter_rows(min_row=2, values_only=True):
                day = _cell_text(values, index["Day"])
                cage = _cell_text(values, index["Cage"])
                animal = _cell_text(values, index["Animal"])
                trial = _cell_int(values, index["Trial"])
                if not (day and cage and animal and trial):
                    continue
                interval_seconds = tuple(
                    _cell_float(values, index.get(f"{start}-{start + 10} cm (s)"))
                    for start in INTERVAL_STARTS_CM
                )
                if not any(value is not None for value in interval_seconds):
                    continue
                output.append(
                    TrialIntervalSeries(
                        day=day,
                        cage=cage,
                        animal=animal,
                        condition=conditions.get((cage, animal), "UNASSIGNED"),
                        trial=trial,
                        interval_seconds=interval_seconds,
                    )
                )
            return output
        finally:
            workbook.close()

    @staticmethod
    def _write_svg(path: Path, text: str) -> None:
        temporary_path = path.with_suffix(f"{path.suffix}.tmp")
        try:
            temporary_path.write_text(text, encoding="utf-8")
            temporary_path.replace(path)
        finally:
            if temporary_path.exists():
                temporary_path.unlink()

    @staticmethod
    def _daily_group_svg(
        dataset: str,
        day: str,
        trial_series: Iterable[TrialIntervalSeries],
        mouse_series: Iterable[MouseIntervalSeries],
        group_summaries: Iterable[GroupIntervalSummary],
    ) -> str:
        trials = sorted(trial_series, key=_trial_sort_key)
        groups = sorted(group_summaries, key=lambda item: _condition_order(item.condition))
        values = _all_interval_values(trials)
        values.extend(_summary_edges(groups))
        lower, upper = _nonnegative_axis_bounds(values)
        width, height = 1420, 790
        left, right, top, bottom = 96, 390, 94, 150
        chart_width, chart_height = width - left - right, height - top - bottom
        x = _interval_x(left, chart_width)
        y = _value_y(top, chart_height, lower, upper)

        parts = _svg_header(
            width, height, left,
            f"Tick-interval elapsed-time comparison - {dataset} {day}",
            "Thin lines are trial interval times. Each mouse is averaged across its trials before it enters the group mean, so every mouse has equal weight.",
            "Thick lines are group means; shaded bands are plus/minus 1 sample SD across mouse means. Gaps indicate no valid interval.",
        )
        _append_interval_axes(parts, left, top, chart_width, chart_height, x, y, lower, upper, "Elapsed time (s)", height)
        _append_summary_bands(parts, groups, x, y, lower, upper)
        for series in trials:
            color = _condition_color(series.condition)
            tooltip = html.escape(f"{series.condition.title()} | Cage {series.cage} Mouse {series.animal} | Trial T{series.trial}")
            _append_hover_lines(parts, _line_segments(series.interval_seconds, x, y), color, tooltip, 1.7, 0.28)
        _append_summary_lines(parts, groups, x, y)
        _append_group_legend(parts, groups, left + chart_width + 40, top)
        parts.append("</svg>")
        return "\n".join(parts)

    @staticmethod
    def _daily_within_trial_variation_svg(
        dataset: str,
        day: str,
        trial_variations: Iterable[TrialIntervalWithinTrialVariation],
        group_summaries: Iterable[GroupIntervalWithinTrialVariation],
    ) -> str:
        """Render an equal-mouse SHAM/STROKE comparison of within-run timing SD."""
        trials = sorted(trial_variations, key=_trial_variation_sort_key)
        groups = sorted(group_summaries, key=lambda item: _condition_order(item.condition))
        values = [item.standard_deviation_seconds for item in trials]
        for group in groups:
            values.append(group.mean_standard_deviation_seconds)
            if group.standard_deviation_seconds is not None:
                values.extend((
                    group.mean_standard_deviation_seconds - group.standard_deviation_seconds,
                    group.mean_standard_deviation_seconds + group.standard_deviation_seconds,
                ))
        lower, upper = _nonnegative_axis_bounds(values)
        width, height = 1120, 740
        left, right, top, bottom = 96, 260, 94, 140
        chart_width, chart_height = width - left - right, height - top - bottom

        def y(value: float) -> float:
            return _value_y(top, chart_height, lower, upper)(value)

        parts = _svg_header(
            width,
            height,
            left,
            f"Tick-interval within-trial timing variation - {dataset} {day}",
            "Each dot is one trial's sample SD across its available 0-10 through 80-90 cm interval times. Lower SD means more even timing within that run.",
            "Large points are equal-mouse group means; shaded ranges are plus/minus 1 sample SD across mice. A mouse first contributes the mean of its trial SDs.",
        )
        if not groups:
            parts.append(
                f'<text x="{left}" y="{top + 50}" font-family="Arial" font-size="16" fill="#444">No SHAM or STROKE within-trial timing-variation data are available yet.</text>'
            )
            parts.append("</svg>")
            return "\n".join(parts)

        for tick in _time_axis_ticks(lower, upper):
            py = y(tick)
            parts.extend((
                f'<line x1="{left}" y1="{py:.1f}" x2="{left + chart_width}" y2="{py:.1f}" stroke="#e6e6e6"/>',
                f'<text x="{left - 10}" y="{py + 4:.1f}" text-anchor="end" font-family="Arial" font-size="12">{_time_tick_label(tick, upper)}</text>',
            ))

        group_x = {
            group.condition: left + chart_width * (index + 0.5) / len(groups)
            for index, group in enumerate(groups)
        }
        for group in groups:
            px = group_x[group.condition]
            parts.append(
                f'<line x1="{px:.1f}" y1="{top}" x2="{px:.1f}" y2="{top + chart_height}" stroke="#f0f0f0"/>'
            )
            parts.append(
                f'<text x="{px:.1f}" y="{top + chart_height + 26}" text-anchor="middle" font-family="Arial" font-size="14">{html.escape(group.condition.title())}</text>'
            )

        for trial in trials:
            px = group_x.get(trial.condition)
            if px is None:
                continue
            color = _condition_color(trial.condition)
            jitter = _trial_dot_jitter(trial.cage, trial.animal, trial.trial)
            tooltip = html.escape(
                f"{trial.condition.title()} | Cage {trial.cage} Mouse {trial.animal} | "
                f"Trial T{trial.trial} | SD {trial.standard_deviation_seconds:.3f} s "
                f"across {trial.interval_count} interval(s)"
            )
            parts.append(
                f'<circle cx="{px + jitter:.1f}" cy="{y(trial.standard_deviation_seconds):.1f}" r="5" fill="{color}" fill-opacity="0.38" cursor="help"><title>{tooltip}</title></circle>'
            )

        for group in groups:
            px = group_x[group.condition]
            color = _condition_color(group.condition)
            mean = group.mean_standard_deviation_seconds
            error = group.standard_deviation_seconds
            if error is not None:
                upper_y = y(min(upper, mean + error))
                lower_y = y(max(lower, mean - error))
                parts.append(
                    f'<rect x="{px - 22:.1f}" y="{upper_y:.1f}" width="44" height="{lower_y - upper_y:.1f}" fill="{color}" fill-opacity="0.18" stroke="none"/>'
                )
                parts.append(
                    f'<line x1="{px:.1f}" y1="{upper_y:.1f}" x2="{px:.1f}" y2="{lower_y:.1f}" stroke="{color}" stroke-width="2"/>'
                )
            parts.append(
                f'<circle cx="{px:.1f}" cy="{y(mean):.1f}" r="7" fill="{color}" stroke="white" stroke-width="1.5"><title>{html.escape(group.condition.title())} mean: {mean:.3f} s; n={group.mouse_count} mice</title></circle>'
            )

        legend_x = left + chart_width + 40
        parts.append(
            f'<text x="{legend_x}" y="{top + 4}" font-family="Arial" font-size="15" font-weight="bold">Groups</text>'
        )
        for index, group in enumerate(groups):
            color = _condition_color(group.condition)
            text_y = top + 30 + index * 28
            parts.extend((
                f'<circle cx="{legend_x + 12}" cy="{text_y - 5}" r="5" fill="{color}"/>',
                f'<text x="{legend_x + 30}" y="{text_y}" font-family="Arial" font-size="13">{html.escape(group.condition.title())} (n={group.mouse_count} mice)</text>',
            ))
        parts.extend((
            f'<text x="{left + chart_width / 2:.1f}" y="{height - 34}" text-anchor="middle" font-family="Arial" font-size="15">Experimental group</text>',
            f'<text x="25" y="{top + chart_height / 2:.1f}" text-anchor="middle" font-family="Arial" font-size="15" transform="rotate(-90 25 {top + chart_height / 2:.1f})">Within-trial interval-time SD (s)</text>',
            "</svg>",
        ))
        return "\n".join(parts)

    @staticmethod
    def _day_bar_svg(
        dataset: str,
        summaries: Iterable[DayGroupSummary],
        *,
        title: str,
        calculation_note: str,
        comparison_note: str,
        empty_note: str,
        y_label: str,
    ) -> str:
        """Render an equal-mouse across-day comparison as paired mean bars.

        A bar is the measured equal-mouse group mean for one experimental day.
        The translucent range and whisker are the sample SD across mouse-level
        values, drawn over the bar so both endpoints remain visible.
        """
        groups = sorted(summaries, key=lambda item: _condition_order(item.condition))
        days = sorted({day for group in groups for day in group.means}, key=natural_day_key)
        values: list[float] = []
        for group in groups:
            for day, mean in group.means.items():
                values.append(mean)
                standard_deviation = group.standard_deviations.get(day)
                if standard_deviation is not None:
                    values.extend((mean - standard_deviation, mean + standard_deviation))
        lower, upper = _nonnegative_axis_bounds(values)
        width, height = 1420, 760
        left, right, top, bottom = 96, 360, 94, 136
        chart_width, chart_height = width - left - right, height - top - bottom
        y = _value_y(top, chart_height, lower, upper)
        baseline_y = y(0.0)
        parts = _svg_header(
            width,
            height,
            left,
            f"{title} - {dataset}",
            calculation_note,
            comparison_note,
        )
        if not days or not groups:
            parts.append(
                f'<text x="{left}" y="{top + 50}" font-family="Arial" font-size="16" fill="#444">{html.escape(empty_note)}</text>'
            )
            parts.append("</svg>")
            return "\n".join(parts)

        for tick in _time_axis_ticks(lower, upper):
            py = y(tick)
            parts.extend((
                f'<line x1="{left}" y1="{py:.1f}" x2="{left + chart_width}" y2="{py:.1f}" stroke="#e6e6e6"/>',
                f'<text x="{left - 10}" y="{py + 4:.1f}" text-anchor="end" font-family="Arial" font-size="12">{_time_tick_label(tick, upper)}</text>',
            ))

        day_spacing = chart_width / len(days)
        group_count = len(groups)
        bar_width = min(64.0, day_spacing * 0.28 / max(group_count, 1))
        bar_gap = min(16.0, bar_width * 0.28)
        group_span = group_count * bar_width + max(group_count - 1, 0) * bar_gap

        for day_index, day in enumerate(days):
            center_x = left + day_spacing * (day_index + 0.5)
            parts.extend((
                f'<line x1="{center_x:.1f}" y1="{top}" x2="{center_x:.1f}" y2="{top + chart_height}" stroke="#f0f0f0"/>',
                f'<text x="{center_x:.1f}" y="{top + chart_height + 26}" text-anchor="middle" font-family="Arial" font-size="13">{html.escape(day)}</text>',
            ))
            for group_index, group in enumerate(groups):
                mean = group.means.get(day)
                if mean is None or not math.isfinite(mean):
                    continue
                color = _condition_color(group.condition)
                center_offset = -group_span / 2 + bar_width / 2 + group_index * (bar_width + bar_gap)
                bar_center = center_x + center_offset
                bar_left = bar_center - bar_width / 2
                standard_deviation = group.standard_deviations.get(day)
                error_top = mean
                error_bottom = mean
                upper_y: float | None = None
                lower_y: float | None = None
                if standard_deviation is not None and math.isfinite(standard_deviation):
                    error_top = min(upper, mean + standard_deviation)
                    error_bottom = max(lower, mean - standard_deviation)
                    upper_y = y(error_top)
                    lower_y = y(error_bottom)
                    parts.append(
                        f'<rect x="{bar_left:.1f}" y="{upper_y:.1f}" width="{bar_width:.1f}" height="{max(0.0, lower_y - upper_y):.1f}" fill="{ERROR_RANGE_COLOR}" fill-opacity="0.24" stroke="none"/>'
                    )

                mean_y = y(mean)
                bar_y = min(mean_y, baseline_y)
                bar_height = abs(baseline_y - mean_y)
                tooltip = html.escape(
                    f"{group.condition.title()} | {day}: mean {mean:.3f} s, "
                    f"SD {standard_deviation:.3f} s" if standard_deviation is not None
                    else f"{group.condition.title()} | {day}: mean {mean:.3f} s, SD unavailable (n={group.mouse_counts.get(day, 0)})"
                )
                parts.append(
                    f'<rect x="{bar_left:.1f}" y="{bar_y:.1f}" width="{bar_width:.1f}" height="{bar_height:.1f}" fill="{color}" cursor="help"><title>{tooltip}</title></rect>'
                )
                if upper_y is not None and lower_y is not None:
                    # The gray range remains behind the bar, while its whisker
                    # is deliberately drawn on top so the lower SD endpoint is
                    # visible even when it falls inside the colored bar.
                    parts.extend((
                        f'<line x1="{bar_center:.1f}" y1="{upper_y:.1f}" x2="{bar_center:.1f}" y2="{lower_y:.1f}" stroke="{ERROR_WHISKER_COLOR}" stroke-width="1.75"/>',
                        f'<line x1="{bar_center - 7:.1f}" y1="{upper_y:.1f}" x2="{bar_center + 7:.1f}" y2="{upper_y:.1f}" stroke="{ERROR_WHISKER_COLOR}" stroke-width="1.75"/>',
                        f'<line x1="{bar_center - 7:.1f}" y1="{lower_y:.1f}" x2="{bar_center + 7:.1f}" y2="{lower_y:.1f}" stroke="{ERROR_WHISKER_COLOR}" stroke-width="1.75"/>',
                    ))
                label_y = max(top + 14, y(error_top) - 8)
                parts.append(
                    f'<text x="{bar_center:.1f}" y="{label_y:.1f}" text-anchor="middle" font-family="Arial" font-size="11" font-weight="bold" fill="{ERROR_WHISKER_COLOR}">{mean:.3f} s</text>'
                )

        legend_x = left + chart_width + 42
        parts.append(f'<text x="{legend_x}" y="{top + 4}" font-family="Arial" font-size="15" font-weight="bold">Groups</text>')
        for index, group in enumerate(groups):
            color = _condition_color(group.condition)
            count = max(group.mouse_counts.values(), default=0)
            text_y = top + 32 + index * 30
            parts.extend((
                f'<rect x="{legend_x}" y="{text_y - 12}" width="18" height="12" fill="{color}"/>',
                f'<text x="{legend_x + 28}" y="{text_y}" font-family="Arial" font-size="13">{html.escape(group.condition.title())} (up to n={count} mice)</text>',
            ))
        parts.extend((
            f'<text x="{left + chart_width / 2:.1f}" y="{height - 34}" text-anchor="middle" font-family="Arial" font-size="15">Experimental day</text>',
            f'<text x="25" y="{top + chart_height / 2:.1f}" text-anchor="middle" font-family="Arial" font-size="15" transform="rotate(-90 25 {top + chart_height / 2:.1f})">{html.escape(y_label)}</text>',
            "</svg>",
        ))
        return "\n".join(parts)

    @staticmethod
    def _day_within_trial_variation_bar_svg(
        dataset: str,
        summaries: Iterable[DayGroupSummary],
    ) -> str:
        return TickIntervalPlotStore._day_bar_svg(
            dataset,
            summaries,
            title="Day-by-day tick-interval within-trial timing variation",
            calculation_note=(
                "For each day, each mouse contributes its mean trial SD across available "
                "0-10 through 80-90 cm interval times; each bar prints the equal-mouse group mean."
            ),
            comparison_note=(
                "Translucent ranges and whiskers are plus/minus 1 sample SD across mice. "
                "Lower SD means more even timing within a run."
            ),
            empty_note="No SHAM or STROKE within-trial timing-variation data are available yet.",
            y_label="Within-trial interval-time SD (s; lower = more even)",
        )

def summarize_group_interval_times(
    trial_series: Iterable[TrialIntervalSeries],
) -> tuple[list[MouseIntervalSeries], list[GroupIntervalSummary]]:
    by_mouse: dict[tuple[str, str, str, str], list[TrialIntervalSeries]] = defaultdict(list)
    for series in trial_series:
        by_mouse[(series.day, series.cage, series.animal, series.condition.upper())].append(series)
    mice: list[MouseIntervalSeries] = []
    for (day, cage, animal, condition), trials in by_mouse.items():
        mice.append(MouseIntervalSeries(
            day, cage, animal, condition,
            tuple(
                statistics.fmean(values) if (values := [trial.interval_seconds[index] for trial in trials if _valid(trial.interval_seconds[index])]) else None
                for index in range(len(INTERVAL_STARTS_CM))
            ),
        ))
    return mice, _group_summaries(mice, lambda item: item.interval_seconds)


def summarize_interval_within_trial_variation(
    trial_series: Iterable[TrialIntervalSeries],
) -> tuple[
    list[TrialIntervalWithinTrialVariation],
    list[MouseIntervalWithinTrialVariation],
    list[GroupIntervalWithinTrialVariation],
]:
    """Quantify how even interval timing is within each individual run.

    Each trial contributes one sample SD calculated across the interval times it
    validly reached.  At least two intervals are required.  The group summary
    first averages a mouse's trial SDs so an animal with three valid trials
    does not outweigh an animal with only one valid trial.
    """
    trial_variations: list[TrialIntervalWithinTrialVariation] = []
    for series in trial_series:
        valid_intervals = [
            float(value) for value in series.interval_seconds if _valid(value)
        ]
        if len(valid_intervals) < 2:
            continue
        trial_variations.append(
            TrialIntervalWithinTrialVariation(
                day=series.day,
                cage=series.cage,
                animal=series.animal,
                condition=series.condition.upper(),
                trial=series.trial,
                standard_deviation_seconds=statistics.stdev(valid_intervals),
                interval_count=len(valid_intervals),
            )
        )

    by_mouse: dict[tuple[str, str, str, str], list[TrialIntervalWithinTrialVariation]] = defaultdict(list)
    for series in trial_variations:
        by_mouse[(series.day, series.cage, series.animal, series.condition)].append(series)
    mice: list[MouseIntervalWithinTrialVariation] = []
    for (day, cage, animal, condition), mouse_trials in by_mouse.items():
        mice.append(
            MouseIntervalWithinTrialVariation(
                day=day,
                cage=cage,
                animal=animal,
                condition=condition,
                mean_standard_deviation_seconds=statistics.fmean(
                    item.standard_deviation_seconds for item in mouse_trials
                ),
                trial_count=len(mouse_trials),
            )
        )

    by_group: dict[str, list[MouseIntervalWithinTrialVariation]] = defaultdict(list)
    for mouse in mice:
        by_group[mouse.condition].append(mouse)
    groups: list[GroupIntervalWithinTrialVariation] = []
    for condition, group_mice in by_group.items():
        if condition not in {"SHAM", "STROKE"}:
            continue
        values = [mouse.mean_standard_deviation_seconds for mouse in group_mice]
        groups.append(
            GroupIntervalWithinTrialVariation(
                condition=condition,
                mean_standard_deviation_seconds=statistics.fmean(values),
                standard_deviation_seconds=(
                    statistics.stdev(values) if len(values) >= 2 else None
                ),
                mouse_count=len(values),
            )
        )
    return trial_variations, mice, groups


def summarize_day_by_day_interval_times(trial_series: Iterable[TrialIntervalSeries]) -> list[DayGroupSummary]:
    mice, _ = summarize_group_interval_times(trial_series)
    return _day_summaries(
        (mouse.day, mouse.condition, mouse.interval_seconds) for mouse in mice
    )


def summarize_day_by_day_interval_within_trial_variation(
    trial_series: Iterable[TrialIntervalSeries],
) -> list[DayGroupSummary]:
    """Compare each day's equal-mouse within-run timing variation."""
    _trials, mice, _groups = summarize_interval_within_trial_variation(trial_series)
    return _day_summaries(
        (
            mouse.day,
            mouse.condition,
            (mouse.mean_standard_deviation_seconds,),
        )
        for mouse in mice
    )


def _group_summaries(mice: Iterable, values_for) -> list[GroupIntervalSummary]:
    by_group: dict[str, list] = defaultdict(list)
    for mouse in mice:
        by_group[mouse.condition].append(mouse)
    summaries: list[GroupIntervalSummary] = []
    for condition, group_mice in by_group.items():
        means: list[float | None] = []
        errors: list[float | None] = []
        counts: list[int] = []
        for index in range(len(INTERVAL_STARTS_CM)):
            values = [values_for(mouse)[index] for mouse in group_mice if _valid(values_for(mouse)[index])]
            counts.append(len(values))
            means.append(statistics.fmean(values) if values else None)
            errors.append(statistics.stdev(values) if len(values) >= 2 else None)
        summaries.append(GroupIntervalSummary(condition, tuple(means), tuple(errors), tuple(counts)))
    return summaries


def _day_summaries(values: Iterable[tuple[str, str, tuple[float | None, ...]]]) -> list[DayGroupSummary]:
    by_group_day: dict[str, dict[str, list[float]]] = defaultdict(lambda: defaultdict(list))
    for day, condition, series in values:
        if condition not in {"SHAM", "STROKE"}:
            continue
        valid_values = [value for value in series if _valid(value)]
        if valid_values:
            by_group_day[condition][day].append(statistics.fmean(valid_values))
    return [
        DayGroupSummary(
            condition,
            {day: statistics.fmean(items) for day, items in per_day.items() if items},
            {day: statistics.stdev(items) if len(items) >= 2 else None for day, items in per_day.items() if items},
            {day: len(items) for day, items in per_day.items() if items},
        )
        for condition, per_day in by_group_day.items()
    ]


def _svg_header(width: int, height: int, left: int, title: str, note_one: str, note_two: str) -> list[str]:
    return [
        f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}">',
        '<rect width="100%" height="100%" fill="white"/>',
        f'<text x="{left}" y="32" font-family="Arial" font-size="22" font-weight="bold">{html.escape(title)}</text>',
        f'<text x="{left}" y="56" font-family="Arial" font-size="13" fill="#444">{html.escape(note_one)}</text>',
        f'<text x="{left}" y="76" font-family="Arial" font-size="13" fill="#444">{html.escape(note_two)}</text>',
    ]


def _append_interval_axes(parts: list[str], left: int, top: int, chart_width: int, chart_height: int, x, y, lower: float, upper: float, y_label: str, height: int) -> None:
    for start in INTERVAL_STARTS_CM:
        px = x(start)
        parts.extend((
            f'<line x1="{px:.1f}" y1="{top}" x2="{px:.1f}" y2="{top + chart_height}" stroke="#e6e6e6"/>',
            f'<text x="{px:.1f}" y="{top + chart_height + 25}" text-anchor="middle" font-family="Arial" font-size="12">{start}-{start + 10}</text>',
        ))
    for tick in _time_axis_ticks(lower, upper):
        py = y(tick)
        parts.extend((
            f'<line x1="{left}" y1="{py:.1f}" x2="{left + chart_width}" y2="{py:.1f}" stroke="#e6e6e6"/>',
            f'<text x="{left - 10}" y="{py + 4:.1f}" text-anchor="end" font-family="Arial" font-size="12">{_time_tick_label(tick, upper)}</text>',
        ))
    parts.extend((
        f'<text x="{left + chart_width / 2:.1f}" y="{height - 34}" text-anchor="middle" font-family="Arial" font-size="15">Back-paw beam interval (cm)</text>',
        f'<text x="25" y="{top + chart_height / 2:.1f}" text-anchor="middle" font-family="Arial" font-size="15" transform="rotate(-90 25 {top + chart_height / 2:.1f})">{html.escape(y_label)}</text>',
    ))


def _append_summary_bands(parts: list[str], groups: Iterable[GroupIntervalSummary], x, y, lower: float, upper: float) -> None:
    for summary in groups:
        color = _condition_color(summary.condition)
        for segment in _band_segments(summary, x, y, lower, upper):
            parts.append(f'<polygon points="{" ".join(segment)}" fill="{color}" fill-opacity="0.18" stroke="none"/>')


def _append_hover_lines(parts: list[str], segments: Iterable[list[str]], color: str, tooltip: str, width: float, opacity: float) -> None:
    for points in segments:
        joined = " ".join(points)
        parts.append(
            f'<polyline points="{joined}" fill="none" stroke="{color}" stroke-width="{width}" stroke-opacity="{opacity}" pointer-events="none"/>'
            f'<polyline points="{joined}" fill="none" stroke="{color}" stroke-width="8" stroke-opacity="0" pointer-events="stroke" cursor="help"><title>{tooltip}</title></polyline>'
        )


def _append_summary_lines(parts: list[str], groups: Iterable[GroupIntervalSummary], x, y) -> None:
    for summary in groups:
        color = _condition_color(summary.condition)
        for segment in _line_segments(summary.means, x, y):
            parts.append(f'<polyline points="{" ".join(segment)}" fill="none" stroke="{color}" stroke-width="4" stroke-linecap="round"/>')


def _append_group_legend(parts: list[str], groups: Iterable[GroupIntervalSummary], legend_x: float, top: int) -> None:
    parts.append(f'<text x="{legend_x}" y="{top + 4}" font-family="Arial" font-size="15" font-weight="bold">Groups</text>')
    for index, summary in enumerate(groups):
        color = _condition_color(summary.condition)
        text_y = top + 30 + index * 28
        count = max(summary.mouse_counts, default=0)
        parts.extend((
            f'<line x1="{legend_x}" y1="{text_y - 5}" x2="{legend_x + 24}" y2="{text_y - 5}" stroke="{color}" stroke-width="4"/>',
            f'<text x="{legend_x + 33}" y="{text_y}" font-family="Arial" font-size="13">{html.escape(summary.condition.title())} (up to n={count} mice)</text>',
        ))


def _line_segments(values: Iterable[float | None], x, y) -> list[list[str]]:
    segments: list[list[str]] = []
    segment: list[str] = []
    for start, value in zip(INTERVAL_STARTS_CM, values):
        if not _valid(value):
            if segment:
                segments.append(segment)
                segment = []
            continue
        segment.append(f"{x(start):.1f},{y(float(value)):.1f}")
    if segment:
        segments.append(segment)
    return segments


def _band_segments(summary: GroupIntervalSummary, x, y, lower: float, upper: float) -> list[list[str]]:
    segments: list[list[tuple[float, float, float]]] = []
    segment: list[tuple[float, float, float]] = []
    for start, mean, error in zip(INTERVAL_STARTS_CM, summary.means, summary.standard_deviations):
        if not _valid(mean) or not _valid(error):
            if segment:
                segments.append(segment)
                segment = []
            continue
        segment.append((x(start), float(mean), float(error)))
    if segment:
        segments.append(segment)
    return [
        [
            *(f"{px:.1f},{y(min(upper, mean + error)):.1f}" for px, mean, error in segment),
            *(f"{px:.1f},{y(max(lower, mean - error)):.1f}" for px, mean, error in reversed(segment)),
        ]
        for segment in segments
    ]


def _interval_x(left: int, chart_width: int):
    return lambda start: left + chart_width * ((start + 5) / 90)


def _value_y(top: int, chart_height: int, lower: float, upper: float):
    return lambda value: top + chart_height * (upper - value) / (upper - lower)


def _nonnegative_axis_bounds(values: Iterable[float]) -> tuple[float, float]:
    valid = [float(value) for value in values if _valid(value)]
    maximum = max(valid, default=0.0)
    if maximum <= 0:
        return 0.0, 1.0
    target = maximum * 1.1
    step = _nice_time_step(target / 6)
    upper = max(step * 2, math.ceil(target / step) * step)
    return 0.0, upper


def _nice_time_step(target: float) -> float:
    """Return a readable seconds increment near one sixth of the plotted range."""
    if target <= 0 or not math.isfinite(target):
        return 1.0
    magnitude = 10 ** math.floor(math.log10(target))
    normalized = target / magnitude
    for candidate in (1.0, 2.0, 5.0, 10.0):
        if normalized <= candidate:
            return candidate * magnitude
    return 10.0 * magnitude


def _time_axis_ticks(lower: float, upper: float) -> list[float]:
    step = _nice_time_step(upper / 6)
    count = max(1, int(round((upper - lower) / step)))
    return [lower + (index * step) for index in range(count + 1)]


def _time_tick_label(value: float, upper: float) -> str:
    step = _nice_time_step(upper / 6)
    if step >= 1:
        return f"{value:.0f}"
    if step >= 0.1:
        return f"{value:.1f}"
    if step >= 0.01:
        return f"{value:.2f}"
    return f"{value:.3f}"


def _all_interval_values(series: Iterable[TrialIntervalSeries]) -> list[float]:
    return [value for item in series for value in item.interval_seconds if _valid(value)]


def _summary_edges(groups: Iterable[GroupIntervalSummary]) -> list[float]:
    values: list[float] = []
    for group in groups:
        for mean, error in zip(group.means, group.standard_deviations):
            if _valid(mean):
                values.append(float(mean))
                if _valid(error):
                    values.append(float(mean) + float(error))
    return values


def _valid(value: float | None) -> bool:
    return value is not None and math.isfinite(value)


def _cell_text(values: tuple[object, ...], index: int | None) -> str:
    if index is None or index >= len(values) or values[index] is None:
        return ""
    value = values[index]
    return str(int(value)) if isinstance(value, float) and value.is_integer() else str(value).strip()


def _cell_int(values: tuple[object, ...], index: int | None) -> int | None:
    if index is None or index >= len(values) or values[index] is None:
        return None
    try:
        return int(float(values[index]))
    except (TypeError, ValueError):
        return None


def _cell_float(values: tuple[object, ...], index: int | None) -> float | None:
    if index is None or index >= len(values) or values[index] is None:
        return None
    try:
        value = float(values[index])
    except (TypeError, ValueError):
        return None
    return value if math.isfinite(value) else None


def _condition_color(condition: str) -> str:
    return {
        "SHAM": SHAM_COLOR,
        "STROKE": STROKE_COLOR,
    }.get(condition.upper(), UNASSIGNED_COLOR)


def _condition_order(condition: str) -> tuple[int, str]:
    upper = condition.upper()
    return ({"SHAM": 0, "STROKE": 1}.get(upper, 2), upper)


def _numeric_key(value: str) -> tuple[int, int | str]:
    return (0, int(value)) if value.isdigit() else (1, value.casefold())


def _trial_sort_key(series: TrialIntervalSeries):
    return (_condition_order(series.condition), _numeric_key(series.cage), _numeric_key(series.animal), series.trial)


def _trial_variation_sort_key(series: TrialIntervalWithinTrialVariation):
    return (
        _condition_order(series.condition),
        _numeric_key(series.cage),
        _numeric_key(series.animal),
        series.trial,
    )


def _trial_dot_jitter(cage: str, animal: str, trial: int) -> float:
    """Return a stable small horizontal offset so same-group trial dots show."""
    identifiers = f"{cage}|{animal}|{trial}"
    checksum = sum((index + 1) * ord(character) for index, character in enumerate(identifiers))
    return float((checksum % 13) - 6) * 4.0
