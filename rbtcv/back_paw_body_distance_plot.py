"""SVG comparisons for calibrated back-paw to body-center distance.

The human-readable workbook stores one frame per wide cell as the measured
distance and matching calibrated back-paw position.  This module turns those
records into the same equal-mouse 10-cm-bin group and consistency charts used
by the other continuous research measurements.
"""

from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
import html
import math
from pathlib import Path
import re
import statistics

from .back_front_paw_distance_plot import (
    BackFrontPawDistancePlotStore,
    GroupPawDistanceSummary,
    MousePawDistanceSeries,
    TrialPawDistanceSeries,
    _all_bin_values,
    _append_bin_axes,
    _append_group_legend,
    _append_hover_lines,
    _append_summary_bands,
    _append_summary_lines,
    _bin_start,
    _bin_x,
    _cell_float,
    _cell_int,
    _cell_text,
    _condition_color,
    _condition_order,
    _line_segments,
    _summary_edges,
    _trial_sort_key,
    _value_y,
    _zoomed_nonnegative_axis_bounds,
    summarize_day_by_day_back_front_paw_distances,
    summarize_group_back_front_paw_distances,
)
from .condition_map import ConditionMapStore
from .dataset import ROOT, natural_day_key


OUTPUT_ROOT = ROOT / "outputs"
RESULTS_FILENAME = "RBT_CV_Results.xlsx"
SHEET_NAME = "Back Paw-Body Distance"
CHART_DIRECTORY = "Back_paw Body Distance"
FILE_SUFFIX = ".svg"
GROUP_PREFIX = "back_paw_body_distance_group_comparison_"
WITHIN_TRIAL_VARIATION_PREFIX = "back_paw_body_distance_within_trial_variation_"
DAY_GROUP_FILENAME = "Day_back_paw_body_distance_group_comparison.svg"
DAY_WITHIN_TRIAL_VARIATION_FILENAME = "Day_back_paw_body_distance_within_trial_variation.svg"
FRAME_CELL_RE = re.compile(r"([+-]?\d+(?:\.\d+)?) cm; ([+-]?\d+(?:\.\d+)?) cm")
ERROR_RANGE_COLOR = "#6B7280"
ERROR_WHISKER_COLOR = "#374151"

# These aliases preserve the generic, metric-independent equal-mouse
# statistics data shape while making the public intent of this module clear.
TrialBackPawBodyDistanceSeries = TrialPawDistanceSeries
MouseBackPawBodyDistanceSeries = MousePawDistanceSeries
GroupBackPawBodyDistanceSummary = GroupPawDistanceSummary


@dataclass(frozen=True)
class TrialBackPawBodyWithinTrialVariation:
    """One trial's sample SD across its raw frame distances in each bin."""

    day: str
    cage: str
    animal: str
    condition: str
    trial: int
    standard_deviations: tuple[float | None, ...]
    frame_counts: tuple[int, ...]


class BackPawBodyDistancePlotStore(BackFrontPawDistancePlotStore):
    """Create back-paw/body-distance comparison and variation charts."""

    def __init__(self, output_root: Path = OUTPUT_ROOT) -> None:
        super().__init__(output_root)

    def chart_dir(self, dataset: str) -> Path:
        return self.result_dir(dataset) / CHART_DIRECTORY

    def refresh_dataset(self, dataset: str) -> dict[str, Path]:
        workbook_path = self.result_dir(dataset) / RESULTS_FILENAME
        if not workbook_path.exists():
            return {}

        conditions = ConditionMapStore(self.output_root).load(dataset)
        trials = self._read_trial_series(workbook_path, conditions)
        within_trial_variation = self._read_within_trial_variation_series(
            workbook_path,
            conditions,
        )
        by_day: dict[str, list[TrialBackPawBodyDistanceSeries]] = defaultdict(list)
        for series in trials:
            by_day[series.day].append(series)
        variation_by_day: dict[str, list[TrialBackPawBodyWithinTrialVariation]] = defaultdict(list)
        for series in within_trial_variation:
            variation_by_day[series.day].append(series)

        chart_dir = self.chart_dir(dataset)
        chart_dir.mkdir(parents=True, exist_ok=True)
        paths: dict[str, Path] = {}
        for day, day_trials in sorted(by_day.items(), key=lambda item: natural_day_key(item[0])):
            mice, groups = summarize_group_back_front_paw_distances(day_trials)
            group_path = chart_dir / f"{GROUP_PREFIX}{day}{FILE_SUFFIX}"
            self._write_svg(
                group_path,
                self._daily_group_svg(dataset, day, day_trials, mice, groups),
            )
            paths[f"group:{day}"] = group_path

            within_trial_path = chart_dir / f"{WITHIN_TRIAL_VARIATION_PREFIX}{day}{FILE_SUFFIX}"
            variation_trials = variation_by_day.get(day, [])
            variation_mice, variation_groups = self._within_trial_group_summaries(
                variation_trials
            )
            self._write_svg(
                within_trial_path,
                self._daily_within_trial_variation_svg(
                    dataset,
                    day,
                    variation_trials,
                    variation_mice,
                    variation_groups,
                ),
            )
            paths[f"within_trial_variation:{day}"] = within_trial_path

        day_group_path = chart_dir / DAY_GROUP_FILENAME
        self._write_svg(
            day_group_path,
            self._day_bar_svg(
                dataset,
                summarize_day_by_day_back_front_paw_distances(trials),
                title="Day-by-day back-paw/body-distance comparison",
                calculation_note=(
                    "For each day, each mouse is averaged across its trials and valid "
                    "10 cm bins from 0-90 cm before group calculation."
                ),
                comparison_note=(
                    "Colored bars are SHAM/STROKE mean distances; neutral-gray ranges and "
                    "whiskers are plus/minus 1 sample SD across mice."
                ),
                empty_note="No SHAM or STROKE back-paw/body-distance data are available yet.",
                y_label="Back-paw to body-center distance (cm)",
            ),
        )
        paths["day_group"] = day_group_path

        day_within_trial_path = chart_dir / DAY_WITHIN_TRIAL_VARIATION_FILENAME
        self._write_svg(
            day_within_trial_path,
            self._day_bar_svg(
                dataset,
                self._day_within_trial_variation_summaries(within_trial_variation),
                title="Day-by-day back-paw/body-distance within-trial variation",
                calculation_note=(
                    "For each trial and 10 cm bin, calculate sample SD across raw valid "
                    "frame distances. Each mouse contributes its mean trial/bin SD."
                ),
                comparison_note=(
                    "Colored bars are SHAM/STROKE mean within-trial variation; neutral-gray "
                    "ranges and whiskers are plus/minus 1 sample SD across mice."
                ),
                empty_note=(
                    "No SHAM or STROKE back-paw/body-distance within-trial variation "
                    "data are available yet."
                ),
                y_label="Within-trial back-paw/body-distance SD (cm)",
            ),
        )
        paths["day_within_trial_variation"] = day_within_trial_path

        live_names = {path.name for path in paths.values()}
        for prefix in (GROUP_PREFIX, WITHIN_TRIAL_VARIATION_PREFIX):
            for stale_path in chart_dir.glob(f"{prefix}*{FILE_SUFFIX}"):
                if stale_path.name not in live_names:
                    stale_path.unlink()
        for obsolete_path in chart_dir.glob(f"back_paw_body_distance_trial_consistency_*{FILE_SUFFIX}"):
            obsolete_path.unlink()
        obsolete_day_consistency = chart_dir / "Day_back_paw_body_distance_trial_consistency.svg"
        if obsolete_day_consistency.exists():
            obsolete_day_consistency.unlink()
        return paths

    @staticmethod
    def _read_trial_series(
        workbook_path: Path,
        conditions: dict[tuple[str, str], str],
    ) -> list[TrialBackPawBodyDistanceSeries]:
        from openpyxl import load_workbook

        workbook = load_workbook(workbook_path, data_only=True, read_only=True)
        try:
            if SHEET_NAME not in workbook.sheetnames:
                return []
            sheet = workbook[SHEET_NAME]
            headers = [cell.value for cell in next(sheet.iter_rows(max_row=1), ())]
            index = {
                str(header): position
                for position, header in enumerate(headers)
                if header is not None
            }
            required = ("Day", "Cage", "Animal", "Trial")
            if not all(header in index for header in required):
                return []
            compact, distance_columns, position_columns = _frame_columns(index)

            output: list[TrialBackPawBodyDistanceSeries] = []
            for values in sheet.iter_rows(min_row=2, values_only=True):
                day = _cell_text(values, index["Day"])
                cage = _cell_text(values, index["Cage"])
                animal = _cell_text(values, index["Animal"])
                trial = _cell_int(values, index["Trial"])
                if not (day and cage and animal and trial):
                    continue

                by_bin: dict[int, list[float]] = defaultdict(list)
                for frame in set(compact) | set(distance_columns) | set(position_columns):
                    distance, position = _frame_measurement(
                        values,
                        frame,
                        compact,
                        distance_columns,
                        position_columns,
                    )
                    if distance is None or position is None:
                        continue
                    bin_start = _bin_start(position)
                    if bin_start is not None:
                        by_bin[bin_start].append(distance)
                if not by_bin:
                    continue
                output.append(
                    TrialBackPawBodyDistanceSeries(
                        day=day,
                        cage=cage,
                        animal=animal,
                        condition=conditions.get((cage, animal), "UNASSIGNED"),
                        trial=trial,
                        bin_means=tuple(
                            statistics.fmean(by_bin[start]) if by_bin[start] else None
                            for start in range(0, 90, 10)
                        ),
                    )
                )
            return output
        finally:
            workbook.close()

    @staticmethod
    def _read_within_trial_variation_series(
        workbook_path: Path,
        conditions: dict[tuple[str, str], str],
    ) -> list[TrialBackPawBodyWithinTrialVariation]:
        """Read raw frame distances and calculate a sample SD for each trial/bin.

        A bin needs at least two valid frame distances.  This intentionally
        does *not* average across trials: a returned line remains one trial's
        within-run variation profile.
        """
        from openpyxl import load_workbook

        workbook = load_workbook(workbook_path, data_only=True, read_only=True)
        try:
            if SHEET_NAME not in workbook.sheetnames:
                return []
            sheet = workbook[SHEET_NAME]
            headers = [cell.value for cell in next(sheet.iter_rows(max_row=1), ())]
            index = {
                str(header): position
                for position, header in enumerate(headers)
                if header is not None
            }
            required = ("Day", "Cage", "Animal", "Trial")
            if not all(header in index for header in required):
                return []
            compact, distance_columns, position_columns = _frame_columns(index)

            output: list[TrialBackPawBodyWithinTrialVariation] = []
            for values in sheet.iter_rows(min_row=2, values_only=True):
                day = _cell_text(values, index["Day"])
                cage = _cell_text(values, index["Cage"])
                animal = _cell_text(values, index["Animal"])
                trial = _cell_int(values, index["Trial"])
                if not (day and cage and animal and trial):
                    continue

                by_bin: dict[int, list[float]] = defaultdict(list)
                for frame in set(compact) | set(distance_columns) | set(position_columns):
                    distance, position = _frame_measurement(
                        values,
                        frame,
                        compact,
                        distance_columns,
                        position_columns,
                    )
                    if distance is None or position is None:
                        continue
                    bin_start = _bin_start(position)
                    if bin_start is not None:
                        by_bin[bin_start].append(distance)

                standard_deviations = tuple(
                    statistics.stdev(by_bin[start]) if len(by_bin[start]) >= 2 else None
                    for start in range(0, 90, 10)
                )
                if not any(value is not None for value in standard_deviations):
                    continue
                output.append(
                    TrialBackPawBodyWithinTrialVariation(
                        day=day,
                        cage=cage,
                        animal=animal,
                        condition=conditions.get((cage, animal), "UNASSIGNED"),
                        trial=trial,
                        standard_deviations=standard_deviations,
                        frame_counts=tuple(len(by_bin[start]) for start in range(0, 90, 10)),
                    )
                )
            return output
        finally:
            workbook.close()

    @staticmethod
    def _within_trial_summary_inputs(
        variation_trials: list[TrialBackPawBodyWithinTrialVariation],
    ) -> list[TrialPawDistanceSeries]:
        """Adapt per-trial SD profiles to the shared equal-mouse summaries."""
        return [
            TrialPawDistanceSeries(
                day=series.day,
                cage=series.cage,
                animal=series.animal,
                condition=series.condition,
                trial=series.trial,
                bin_means=series.standard_deviations,
            )
            for series in variation_trials
        ]

    @classmethod
    def _within_trial_group_summaries(
        cls,
        variation_trials: list[TrialBackPawBodyWithinTrialVariation],
    ) -> tuple[list[MouseBackPawBodyDistanceSeries], list[GroupBackPawBodyDistanceSummary]]:
        """Average each mouse's trial SDs before computing group statistics."""
        return summarize_group_back_front_paw_distances(
            cls._within_trial_summary_inputs(variation_trials)
        )

    @classmethod
    def _day_within_trial_variation_summaries(
        cls,
        variation_trials: list[TrialBackPawBodyWithinTrialVariation],
    ):
        """Return day summaries with equal mouse, not equal trial, weighting."""
        return summarize_day_by_day_back_front_paw_distances(
            cls._within_trial_summary_inputs(variation_trials)
        )

    @staticmethod
    def _daily_group_svg(
        dataset: str,
        day: str,
        trial_series: list[TrialBackPawBodyDistanceSeries],
        mouse_series: list[MouseBackPawBodyDistanceSeries],
        group_summaries: list[GroupBackPawBodyDistanceSummary],
    ) -> str:
        del mouse_series  # Group legend summarizes mouse counts.
        trials = sorted(trial_series, key=_trial_sort_key)
        groups = sorted(group_summaries, key=lambda item: _condition_order(item.condition))
        values = _all_bin_values(trials)
        values.extend(_summary_edges(groups))
        lower, upper = _zoomed_nonnegative_axis_bounds(values)
        width, height = 1420, 790
        left, right, top, bottom = 96, 390, 94, 150
        chart_width, chart_height = width - left - right, height - top - bottom
        x = _bin_x(left, chart_width)
        y = _value_y(top, chart_height, lower, upper)

        parts = [
            f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}">',
            '<rect width="100%" height="100%" fill="white"/>',
            f'<text x="{left}" y="32" font-family="Arial" font-size="22" font-weight="bold">Back-paw/body-distance group comparison - {html.escape(dataset)} {html.escape(day)}</text>',
            f'<text x="{left}" y="56" font-family="Arial" font-size="13" fill="#444">Thin lines are trial means within each 10 cm bin. Each mouse is averaged across its trials before it enters the group mean, so every mouse has equal weight.</text>',
            f'<text x="{left}" y="76" font-family="Arial" font-size="13" fill="#444">Thick lines are group means; shaded bands are plus/minus 1 sample SD across mouse means. Gaps indicate no valid observations.</text>',
        ]
        _append_bin_axes(
            parts,
            left,
            top,
            chart_width,
            chart_height,
            x,
            y,
            lower,
            upper,
            "Back-paw to body-center distance (cm)",
            height,
        )
        _append_summary_bands(parts, groups, x, y, lower, upper)
        for series in trials:
            color = _condition_color(series.condition)
            tooltip = html.escape(
                f"{series.condition.title()} | Cage {series.cage} Mouse {series.animal} | Trial T{series.trial}"
            )
            _append_hover_lines(
                parts,
                _line_segments(series.bin_means, x, y),
                color,
                tooltip,
                1.7,
                0.28,
            )
        _append_summary_lines(parts, groups, x, y)
        _append_group_legend(parts, groups, left + chart_width + 40, top)
        parts.append("</svg>")
        return "\n".join(parts)

    @staticmethod
    def _day_bar_svg(
        dataset: str,
        summaries,
        *,
        title: str,
        calculation_note: str,
        comparison_note: str,
        empty_note: str,
        y_label: str,
    ) -> str:
        """Render equal-mouse day summaries as data-driven paired bars."""
        groups = sorted(summaries, key=lambda item: _condition_order(item.condition))
        days = sorted(
            {day for group in groups for day in group.means},
            key=natural_day_key,
        )
        values: list[float] = []
        for group in groups:
            for day, mean in group.means.items():
                values.append(mean)
                error = group.standard_deviations.get(day)
                if error is not None:
                    values.extend((mean - error, mean + error))
        lower, upper = _bar_axis_bounds(values)
        width, height = 1420, 760
        left, right, top, bottom = 96, 360, 94, 136
        chart_width, chart_height = width - left - right, height - top - bottom
        y = _value_y(top, chart_height, lower, upper)
        baseline_y = y(0.0)
        parts = [
            f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}">',
            '<rect width="100%" height="100%" fill="white"/>',
            f'<text x="{left}" y="32" font-family="Arial" font-size="22" font-weight="bold">{html.escape(title)} - {html.escape(dataset)}</text>',
            f'<text x="{left}" y="56" font-family="Arial" font-size="13" fill="#444">{html.escape(calculation_note)}</text>',
            f'<text x="{left}" y="76" font-family="Arial" font-size="13" fill="#444">{html.escape(comparison_note)}</text>',
        ]
        if not days or not groups:
            parts.append(
                f'<text x="{left}" y="{top + 50}" font-family="Arial" font-size="16" fill="#444">{html.escape(empty_note)}</text>'
            )
            parts.append("</svg>")
            return "\n".join(parts)

        for tick in _bar_axis_ticks(lower, upper):
            py = y(tick)
            parts.extend((
                f'<line x1="{left}" y1="{py:.1f}" x2="{left + chart_width}" y2="{py:.1f}" stroke="#e6e6e6"/>',
                f'<text x="{left - 10}" y="{py + 4:.1f}" text-anchor="end" font-family="Arial" font-size="12">{_bar_tick_label(tick, upper)}</text>',
            ))

        day_spacing = chart_width / len(days)
        group_count = len(groups)
        bar_width = min(64.0, day_spacing * 0.28 / max(group_count, 1))
        bar_gap = min(16.0, bar_width * 0.28)
        group_span = group_count * bar_width + max(group_count - 1, 0) * bar_gap
        for day_index, day in enumerate(days):
            center_x = left + day_spacing * (day_index + 0.5)
            parts.extend((
                f'<line x1="{center_x:.1f}" y1="{top}" x2="{center_x:.1f}" y2="{top + chart_height}" stroke="#f0f0f0"/>',
                f'<text x="{center_x:.1f}" y="{top + chart_height + 26}" text-anchor="middle" font-family="Arial" font-size="13">{html.escape(day)}</text>',
            ))
            for group_index, group in enumerate(groups):
                mean = group.means.get(day)
                if mean is None or not math.isfinite(mean):
                    continue
                color = _condition_color(group.condition)
                center_offset = -group_span / 2 + bar_width / 2 + group_index * (bar_width + bar_gap)
                bar_center = center_x + center_offset
                bar_left = bar_center - bar_width / 2
                error = group.standard_deviations.get(day)
                error_top = mean
                error_bottom = mean
                upper_y: float | None = None
                lower_y: float | None = None
                if error is not None and math.isfinite(error):
                    error_top = min(upper, mean + error)
                    error_bottom = max(lower, mean - error)
                    upper_y = y(error_top)
                    lower_y = y(error_bottom)
                    parts.append(
                        f'<rect x="{bar_left:.1f}" y="{upper_y:.1f}" width="{bar_width:.1f}" height="{max(0.0, lower_y - upper_y):.1f}" fill="{ERROR_RANGE_COLOR}" fill-opacity="0.24" stroke="none"/>'
                    )

                mean_y = y(mean)
                parts.append(
                    f'<rect x="{bar_left:.1f}" y="{min(mean_y, baseline_y):.1f}" width="{bar_width:.1f}" height="{abs(baseline_y - mean_y):.1f}" fill="{color}" cursor="help"><title>{html.escape(_day_bar_tooltip(group.condition, day, mean, error, group.mouse_counts.get(day, 0)))}</title></rect>'
                )
                if upper_y is not None and lower_y is not None:
                    parts.extend((
                        f'<line x1="{bar_center:.1f}" y1="{upper_y:.1f}" x2="{bar_center:.1f}" y2="{lower_y:.1f}" stroke="{ERROR_WHISKER_COLOR}" stroke-width="1.75"/>',
                        f'<line x1="{bar_center - 7:.1f}" y1="{upper_y:.1f}" x2="{bar_center + 7:.1f}" y2="{upper_y:.1f}" stroke="{ERROR_WHISKER_COLOR}" stroke-width="1.75"/>',
                        f'<line x1="{bar_center - 7:.1f}" y1="{lower_y:.1f}" x2="{bar_center + 7:.1f}" y2="{lower_y:.1f}" stroke="{ERROR_WHISKER_COLOR}" stroke-width="1.75"/>',
                    ))
                parts.append(
                    f'<text x="{bar_center:.1f}" y="{max(top + 14, y(error_top) - 8):.1f}" text-anchor="middle" font-family="Arial" font-size="11" font-weight="bold" fill="{ERROR_WHISKER_COLOR}">{mean:.3f} cm</text>'
                )

        legend_x = left + chart_width + 42
        parts.append(f'<text x="{legend_x}" y="{top + 4}" font-family="Arial" font-size="15" font-weight="bold">Groups</text>')
        for index, group in enumerate(groups):
            color = _condition_color(group.condition)
            count = max(group.mouse_counts.values(), default=0)
            text_y = top + 32 + index * 30
            parts.extend((
                f'<rect x="{legend_x}" y="{text_y - 12}" width="18" height="12" fill="{color}"/>',
                f'<text x="{legend_x + 28}" y="{text_y}" font-family="Arial" font-size="13">{html.escape(group.condition.title())} (up to n={count} mice)</text>',
            ))
        parts.extend((
            f'<text x="{left + chart_width / 2:.1f}" y="{height - 34}" text-anchor="middle" font-family="Arial" font-size="15">Experimental day</text>',
            f'<text x="25" y="{top + chart_height / 2:.1f}" text-anchor="middle" font-family="Arial" font-size="15" transform="rotate(-90 25 {top + chart_height / 2:.1f})">{html.escape(y_label)}</text>',
            "</svg>",
        ))
        return "\n".join(parts)

    @staticmethod
    def _daily_within_trial_variation_svg(
        dataset: str,
        day: str,
        trial_variation: list[TrialBackPawBodyWithinTrialVariation],
        mouse_series: list[MouseBackPawBodyDistanceSeries],
        group_summaries: list[GroupBackPawBodyDistanceSummary],
    ) -> str:
        """Render raw within-trial frame variation without hiding trial lines."""
        del mouse_series  # Mouse averages are represented by group lines/bands.
        trials = sorted(trial_variation, key=_trial_sort_key)
        groups = sorted(group_summaries, key=lambda item: _condition_order(item.condition))
        values = [
            value
            for trial in trials
            for value in trial.standard_deviations
            if value is not None
        ]
        values.extend(_summary_edges(groups))
        lower, upper = _zoomed_nonnegative_axis_bounds(values)
        width, height = 1420, 790
        left, right, top, bottom = 96, 390, 94, 150
        chart_width, chart_height = width - left - right, height - top - bottom
        x = _bin_x(left, chart_width)
        y = _value_y(top, chart_height, lower, upper)

        parts = [
            f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}">',
            '<rect width="100%" height="100%" fill="white"/>',
            f'<text x="{left}" y="32" font-family="Arial" font-size="22" font-weight="bold">Back-paw/body-distance within-trial variation - {html.escape(dataset)} {html.escape(day)}</text>',
            f'<text x="{left}" y="56" font-family="Arial" font-size="13" fill="#444">Each thin line is one trial. A point is that trial\'s sample SD across raw valid frame distances in the 10 cm bin; higher values mean more distance variation during that run.</text>',
            f'<text x="{left}" y="76" font-family="Arial" font-size="13" fill="#444">For group values, each mouse first contributes the mean of its own trial SDs. Thick lines are group means; shaded bands are plus/minus 1 sample SD across mice.</text>',
        ]
        _append_bin_axes(
            parts,
            left,
            top,
            chart_width,
            chart_height,
            x,
            y,
            lower,
            upper,
            "Within-trial back-paw/body-distance SD (cm)",
            height,
        )
        _append_summary_bands(parts, groups, x, y, lower, upper)
        for trial in trials:
            color = _condition_color(trial.condition)
            tooltip = html.escape(
                f"{trial.condition.title()} | Cage {trial.cage} Mouse {trial.animal} | "
                f"Trial T{trial.trial} | Raw frame distance SD"
            )
            _append_hover_lines(
                parts,
                _line_segments(trial.standard_deviations, x, y),
                color,
                tooltip,
                1.7,
                0.28,
            )
        _append_summary_lines(parts, groups, x, y)
        _append_group_legend(parts, groups, left + chart_width + 40, top)
        parts.append("</svg>")
        return "\n".join(parts)


def _frame_columns(index: dict[str, int]) -> tuple[dict[int, int], dict[int, int], dict[int, int]]:
    compact: dict[int, int] = {}
    distance_columns: dict[int, int] = {}
    position_columns: dict[int, int] = {}
    for header, column in index.items():
        if not header.startswith("Frame "):
            continue
        try:
            if header.endswith(" distance (cm)"):
                distance_columns[
                    int(header.removeprefix("Frame ").removesuffix(" distance (cm)"))
                ] = column
            elif header.endswith(" back paw (cm)"):
                position_columns[
                    int(header.removeprefix("Frame ").removesuffix(" back paw (cm)"))
                ] = column
            else:
                compact[int(header.removeprefix("Frame "))] = column
        except ValueError:
            continue
    return compact, distance_columns, position_columns


def _frame_measurement(
    values: tuple[object, ...],
    frame: int,
    compact: dict[int, int],
    distance_columns: dict[int, int],
    position_columns: dict[int, int],
) -> tuple[float | None, float | None]:
    if frame in compact:
        column = compact[frame]
        if column >= len(values) or values[column] is None:
            return None, None
        match = FRAME_CELL_RE.fullmatch(str(values[column]).strip())
        if match is None:
            return None, None
        return float(match.group(1)), float(match.group(2))
    return (
        _cell_float(values, distance_columns.get(frame)),
        _cell_float(values, position_columns.get(frame)),
    )


def _numeric_key(value: str) -> tuple[int, int | str]:
    return (0, int(value)) if value.isdigit() else (1, value.casefold())


def _bar_axis_bounds(values: list[float]) -> tuple[float, float]:
    """Return a zero-baseline, readable axis appropriate for mean bars."""
    maximum = max((float(value) for value in values if math.isfinite(value)), default=0.0)
    if maximum <= 0:
        return 0.0, 1.0
    target = maximum * 1.12
    magnitude = 10 ** math.floor(math.log10(target))
    normalized = target / magnitude
    for candidate in (1.0, 2.0, 5.0, 10.0):
        if normalized <= candidate:
            return 0.0, candidate * magnitude
    return 0.0, 10.0 * magnitude


def _bar_axis_ticks(lower: float, upper: float) -> list[float]:
    span = upper - lower
    if span <= 0:
        return [lower]
    step = span / 6.0
    return [lower + index * step for index in range(7)]


def _bar_tick_label(value: float, upper: float) -> str:
    if upper >= 10:
        return f"{value:.0f}"
    if upper >= 1:
        return f"{value:.1f}"
    return f"{value:.2f}"


def _day_bar_tooltip(
    condition: str,
    day: str,
    mean: float,
    standard_deviation: float | None,
    mouse_count: int,
) -> str:
    if standard_deviation is None:
        return f"{condition.title()} | {day}: mean {mean:.3f} cm; SD unavailable (n={mouse_count})"
    return f"{condition.title()} | {day}: mean {mean:.3f} cm; SD {standard_deviation:.3f} cm (n={mouse_count})"
