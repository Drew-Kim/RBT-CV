"""Group comparison plots for signed tail angle, weighted equally by mouse."""

from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
import html
import math
from pathlib import Path
import re
import statistics
from typing import Iterable

from .condition_map import ConditionMapStore
from .dataset import ROOT, natural_day_key


OUTPUT_ROOT = ROOT / "outputs"
RESULTS_FILENAME = "RBT_CV_Results.xlsx"
ANGLE_SHEET = "Frame Angles"
FILE_PREFIX = "tail_angle_group_comparison_"
FILE_SUFFIX = ".svg"
CHART_DIRECTORY = "Tail Angle"
DAY_COMPARISON_FILENAME = "Day_tail_angle_group_comparison.svg"
WITHIN_TRIAL_VARIATION_PREFIX = "tail_angle_within_trial_variation_"
DAY_WITHIN_TRIAL_VARIATION_FILENAME = "Day_tail_angle_within_trial_variation.svg"
OBSOLETE_CONSISTENCY_PREFIX = "tail_angle_trial_consistency_"
OBSOLETE_DAY_CONSISTENCY_FILENAME = "Day_tail_angle_trial_consistency.svg"
LEGACY_DAY_FILENAMES = (
    "tail_angle_group_comparison_by_day.svg",
    "tail_angle_trial_consistency_by_day.svg",
)
BIN_STARTS_CM = tuple(range(0, 90, 10))

SHAM_COLOR = "#1976D2"
STROKE_COLOR = "#D32F2F"
UNASSIGNED_COLOR = "#616161"
ERROR_RANGE_COLOR = "#6B7280"
ERROR_WHISKER_COLOR = "#374151"
FRAME_CELL_RE = re.compile(r"([+-]?\d+(?:\.\d+)?) deg; ([+-]?\d+(?:\.\d+)?) cm")


@dataclass(frozen=True)
class TrialAngleSeries:
    day: str
    cage: str
    animal: str
    condition: str
    trial: int
    bin_means: tuple[float | None, ...]
    # Frame-level values remain separated by beam position so within-run
    # variation is not inflated by the normal posture change along the beam.
    frame_angles_by_bin: tuple[tuple[float, ...], ...] = ()


@dataclass(frozen=True)
class MouseAngleSeries:
    day: str
    cage: str
    animal: str
    condition: str
    bin_means: tuple[float | None, ...]


@dataclass(frozen=True)
class GroupAngleSummary:
    condition: str
    means: tuple[float | None, ...]
    standard_deviations: tuple[float | None, ...]
    mouse_counts: tuple[int, ...]


@dataclass(frozen=True)
class DayGroupAngleSummary:
    """One group-level 0-90 cm tail-angle value for each experimental day."""

    condition: str
    means: dict[str, float]
    standard_deviations: dict[str, float | None]
    mouse_counts: dict[str, int]


@dataclass(frozen=True)
class TrialAngleWithinTrialVariation:
    """One run's position-controlled tail-angle variation."""

    day: str
    cage: str
    animal: str
    condition: str
    trial: int
    mean_standard_deviation_degrees: float
    bin_standard_deviations: tuple[float, ...]

    @property
    def bin_count(self) -> int:
        return len(self.bin_standard_deviations)


@dataclass(frozen=True)
class MouseAngleWithinTrialVariation:
    """Equal-weight within-run variation summary for one mouse/day."""

    day: str
    cage: str
    animal: str
    condition: str
    mean_standard_deviation_degrees: float
    trial_count: int
    bin_observation_count: int


@dataclass(frozen=True)
class GroupAngleWithinTrialVariation:
    """Equal-mouse SHAM/STROKE summary of within-run angle variation."""

    condition: str
    mean_standard_deviation_degrees: float
    standard_deviation_degrees: float | None
    mouse_count: int


class TailAngleGroupPlotStore:
    """Create daily signed-tail-angle comparisons from exported frame data."""

    def __init__(self, output_root: Path = OUTPUT_ROOT) -> None:
        self.output_root = output_root

    def result_dir(self, dataset: str) -> Path:
        return self.output_root / f"{dataset} Results"

    def chart_dir(self, dataset: str) -> Path:
        return self.result_dir(dataset) / CHART_DIRECTORY

    def refresh_dataset(self, dataset: str) -> dict[str, Path]:
        workbook_path = self.result_dir(dataset) / RESULTS_FILENAME
        if not workbook_path.exists():
            return {}

        conditions = ConditionMapStore(self.output_root).load(dataset)
        trial_series = self._read_trial_series(workbook_path, conditions)
        by_day: dict[str, list[TrialAngleSeries]] = defaultdict(list)
        for series in trial_series:
            by_day[series.day].append(series)

        result_dir = self.result_dir(dataset)
        chart_dir = self.chart_dir(dataset)
        chart_dir.mkdir(parents=True, exist_ok=True)
        paths: dict[str, Path] = {}
        for day, day_trials in sorted(by_day.items(), key=lambda item: natural_day_key(item[0])):
            mouse_series, group_summaries = summarize_group_angles(day_trials)
            path = chart_dir / f"{FILE_PREFIX}{day}{FILE_SUFFIX}"
            self._write_svg(path, self._svg(dataset, day, day_trials, mouse_series, group_summaries))
            paths[day] = path

            variation_trials, _variation_mice, variation_groups = (
                summarize_tail_angle_within_trial_variation(day_trials)
            )
            variation_path = chart_dir / f"{WITHIN_TRIAL_VARIATION_PREFIX}{day}{FILE_SUFFIX}"
            self._write_svg(
                variation_path,
                self._daily_within_trial_variation_svg(
                    dataset,
                    day,
                    variation_trials,
                    variation_groups,
                ),
            )
            paths[f"within_trial_variation:{day}"] = variation_path

        by_day_path = chart_dir / DAY_COMPARISON_FILENAME
        self._write_svg(
            by_day_path,
            self._day_by_day_svg(
                dataset,
                summarize_day_by_day_group_angles(trial_series),
            ),
        )
        paths["by_day"] = by_day_path

        by_day_variation_path = chart_dir / DAY_WITHIN_TRIAL_VARIATION_FILENAME
        self._write_svg(
            by_day_variation_path,
            self._day_by_day_within_trial_variation_svg(
                dataset,
                summarize_day_by_day_tail_angle_within_trial_variation(trial_series),
            ),
        )
        paths["by_day_within_trial_variation"] = by_day_variation_path

        live_names = {path.name for path in paths.values()}
        for prefix in (FILE_PREFIX, WITHIN_TRIAL_VARIATION_PREFIX):
            for stale_path in chart_dir.glob(f"{prefix}*{FILE_SUFFIX}"):
                if stale_path.name not in live_names:
                    stale_path.unlink()
        for stale_path in chart_dir.glob(f"{OBSOLETE_CONSISTENCY_PREFIX}*{FILE_SUFFIX}"):
            stale_path.unlink()
        for stale_name in (
            OBSOLETE_DAY_CONSISTENCY_FILENAME,
            *LEGACY_DAY_FILENAMES,
        ):
            stale_path = chart_dir / stale_name
            if stale_path.exists() and stale_path.name not in live_names:
                stale_path.unlink()
        # These are generated charts from the pre-folder layout, not source
        # data. Remove them after their replacements have been written.
        for stale_path in result_dir.glob(f"{FILE_PREFIX}*{FILE_SUFFIX}"):
            stale_path.unlink()
        for stale_name in (
            DAY_COMPARISON_FILENAME,
            DAY_WITHIN_TRIAL_VARIATION_FILENAME,
            OBSOLETE_DAY_CONSISTENCY_FILENAME,
            *LEGACY_DAY_FILENAMES,
        ):
            stale_path = result_dir / stale_name
            if stale_path.exists():
                stale_path.unlink()
        for stale_path in result_dir.glob(f"{WITHIN_TRIAL_VARIATION_PREFIX}*{FILE_SUFFIX}"):
            stale_path.unlink()
        for stale_path in result_dir.glob(f"{OBSOLETE_CONSISTENCY_PREFIX}*{FILE_SUFFIX}"):
            stale_path.unlink()
        return paths

    @staticmethod
    def _day_by_day_svg(
        dataset: str,
        group_summaries: Iterable[DayGroupAngleSummary],
    ) -> str:
        return TailAngleGroupPlotStore._day_bar_svg(
            dataset,
            group_summaries,
            title="Day-by-day signed tail-angle comparison",
            calculation_note=(
                "For each day, each mouse is averaged across its trials and valid "
                "10 cm bins from 0-90 cm before group calculation."
            ),
            comparison_note=(
                "Colored bars are SHAM/STROKE mean signed tail angles; neutral-gray ranges and "
                "whiskers are plus/minus 1 sample SD across mice."
            ),
            empty_note="No SHAM or STROKE signed tail-angle data are available yet.",
            y_label="Signed tail angle (degrees)",
            value_suffix="\N{DEGREE SIGN}",
            signed_values=True,
        )

    @staticmethod
    def _day_by_day_within_trial_variation_svg(
        dataset: str,
        group_summaries: Iterable[DayGroupAngleSummary],
    ) -> str:
        return TailAngleGroupPlotStore._day_bar_svg(
            dataset,
            group_summaries,
            title="Day-by-day tail-angle within-trial variation",
            calculation_note=(
                "For each trial, calculate frame-level SD within each valid 10 cm bin. "
                "Each mouse contributes the mean of its valid trial/bin SDs; each bar prints the group mean."
            ),
            comparison_note=(
                "Neutral-gray ranges and whiskers are plus/minus 1 sample SD across mice. "
                "Lower SD means more stable position-controlled tail posture within runs."
            ),
            empty_note="No SHAM or STROKE within-trial tail-angle variation data are available yet.",
            y_label="Within-trial tail-angle SD (degrees; lower = more stable)",
            value_suffix="\N{DEGREE SIGN}",
            signed_values=False,
        )

    @staticmethod
    def _day_by_day_summary_svg(
        dataset: str,
        group_summaries: Iterable[DayGroupAngleSummary],
        *,
        title: str,
        calculation_note: str,
        comparison_note: str,
        empty_note: str,
        y_label: str,
        show_zero_reference: bool,
    ) -> str:
        groups = sorted(group_summaries, key=lambda item: _condition_order(item.condition))
        days = sorted(
            {day for summary in groups for day in summary.means},
            key=natural_day_key,
        )
        values: list[float] = []
        for summary in groups:
            for day, mean in summary.means.items():
                values.append(mean)
                error = summary.standard_deviations.get(day)
                if error is not None:
                    values.extend((mean - error, mean + error))
        lower, upper = _axis_bounds(values)

        width, height = 1320, 720
        left, right, top, bottom = 96, 330, 94, 126
        chart_width = width - left - right
        chart_height = height - top - bottom

        def x(day: str) -> float:
            if len(days) <= 1:
                return left + (chart_width / 2)
            return left + chart_width * days.index(day) / (len(days) - 1)

        def y(value: float) -> float:
            return top + chart_height * (upper - value) / (upper - lower)

        # Both groups share the experimental-day coordinate. Their vertical
        # positions and bands carry the comparison; no horizontal offset is used.
        def group_x(day: str, _condition: str) -> float:
            return x(day)

        parts = [
            f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}">',
            '<rect width="100%" height="100%" fill="white"/>',
            f'<text x="{left}" y="32" font-family="Arial" font-size="22" font-weight="bold">{html.escape(title)} - {html.escape(dataset)}</text>',
            f'<text x="{left}" y="56" font-family="Arial" font-size="13" fill="#444">{html.escape(calculation_note)}</text>',
            f'<text x="{left}" y="76" font-family="Arial" font-size="13" fill="#444">{html.escape(comparison_note)}</text>',
        ]
        if not days:
            parts.extend(
                (
                    f'<text x="{left}" y="{top + 50}" font-family="Arial" font-size="16" fill="#444">{html.escape(empty_note)}</text>',
                    "</svg>",
                )
            )
            return "\n".join(parts)

        for day in days:
            px = x(day)
            parts.extend(
                (
                    f'<line x1="{px:.1f}" y1="{top}" x2="{px:.1f}" y2="{top + chart_height}" stroke="#e6e6e6"/>',
                    f'<text x="{px:.1f}" y="{top + chart_height + 25}" text-anchor="middle" font-family="Arial" font-size="12">{html.escape(day)}</text>',
                )
            )
        for tick in range(int(lower), int(upper) + 1, 5):
            py = y(tick)
            parts.extend(
                (
                    f'<line x1="{left}" y1="{py:.1f}" x2="{left + chart_width}" y2="{py:.1f}" stroke="#e6e6e6"/>',
                    f'<text x="{left - 10}" y="{py + 4:.1f}" text-anchor="end" font-family="Arial" font-size="12">{tick}</text>',
                )
            )
        if show_zero_reference and lower < 0 < upper:
            zero = y(0)
            parts.append(
                f'<line x1="{left}" y1="{zero:.1f}" x2="{left + chart_width}" y2="{zero:.1f}" stroke="#555" stroke-width="1.5" stroke-dasharray="6 4"/>'
            )

        for summary in groups:
            color = _condition_color(summary.condition)
            summary_x = lambda day, condition=summary.condition: group_x(day, condition)
            for segment in _day_band_segments(summary, days, summary_x, y, lower, upper):
                parts.append(
                    f'<polygon points="{" ".join(segment)}" fill="{color}" fill-opacity="0.18" stroke="none"/>'
                )

        for summary in groups:
            color = _condition_color(summary.condition)
            summary_x = lambda day, condition=summary.condition: group_x(day, condition)
            for segment in _day_line_segments(summary, days, summary_x, y):
                parts.append(
                    f'<polyline points="{" ".join(segment)}" fill="none" stroke="{color}" stroke-width="4" stroke-linecap="round"/>'
                )
            for day, mean in summary.means.items():
                parts.append(f'<circle cx="{summary_x(day):.1f}" cy="{y(mean):.1f}" r="4.5" fill="{color}"/>')

        legend_x = left + chart_width + 40
        parts.append(
            f'<text x="{legend_x}" y="{top + 4}" font-family="Arial" font-size="15" font-weight="bold">Groups</text>'
        )
        for index, summary in enumerate(groups):
            color = _condition_color(summary.condition)
            text_y = top + 30 + index * 28
            largest_n = max(summary.mouse_counts.values(), default=0)
            parts.extend(
                (
                    f'<line x1="{legend_x}" y1="{text_y - 5}" x2="{legend_x + 24}" y2="{text_y - 5}" stroke="{color}" stroke-width="4"/>',
                    f'<text x="{legend_x + 33}" y="{text_y}" font-family="Arial" font-size="13">{html.escape(summary.condition.title())} (up to n={largest_n} mice)</text>',
                )
            )

        parts.extend(
            (
                f'<text x="{left + chart_width / 2:.1f}" y="{height - 34}" text-anchor="middle" font-family="Arial" font-size="15">Experimental day</text>',
                f'<text x="25" y="{top + chart_height / 2:.1f}" text-anchor="middle" font-family="Arial" font-size="15" transform="rotate(-90 25 {top + chart_height / 2:.1f})">{html.escape(y_label)}</text>',
                "</svg>",
            )
        )
        return "\n".join(parts)

    @staticmethod
    def _day_bar_svg(
        dataset: str,
        group_summaries: Iterable[DayGroupAngleSummary],
        *,
        title: str,
        calculation_note: str,
        comparison_note: str,
        empty_note: str,
        y_label: str,
        value_suffix: str,
        signed_values: bool,
    ) -> str:
        """Render data-driven daily means as paired bars with visible SD limits."""
        groups = sorted(group_summaries, key=lambda item: _condition_order(item.condition))
        days = sorted(
            {day for summary in groups for day in summary.means}, key=natural_day_key
        )
        values: list[float] = []
        for summary in groups:
            for day, mean in summary.means.items():
                values.append(mean)
                standard_deviation = summary.standard_deviations.get(day)
                if standard_deviation is not None:
                    values.extend((mean - standard_deviation, mean + standard_deviation))
        lower, upper = (
            _axis_bounds(values) if signed_values else _nonnegative_axis_bounds(values)
        )
        width, height = 1420, 760
        left, right, top, bottom = 96, 360, 94, 136
        chart_width, chart_height = width - left - right, height - top - bottom

        def y(value: float) -> float:
            return top + chart_height * (upper - value) / (upper - lower)

        baseline_y = y(0.0)
        parts = [
            f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}">',
            '<rect width="100%" height="100%" fill="white"/>',
            f'<text x="{left}" y="32" font-family="Arial" font-size="22" font-weight="bold">{html.escape(title)} - {html.escape(dataset)}</text>',
            f'<text x="{left}" y="56" font-family="Arial" font-size="13" fill="#444">{html.escape(calculation_note)}</text>',
            f'<text x="{left}" y="76" font-family="Arial" font-size="13" fill="#444">{html.escape(comparison_note)}</text>',
        ]
        if not days or not groups:
            parts.extend(
                (
                    f'<text x="{left}" y="{top + 50}" font-family="Arial" font-size="16" fill="#444">{html.escape(empty_note)}</text>',
                    "</svg>",
                )
            )
            return "\n".join(parts)

        for tick in _angle_axis_ticks(lower, upper):
            py = y(tick)
            parts.extend(
                (
                    f'<line x1="{left}" y1="{py:.1f}" x2="{left + chart_width}" y2="{py:.1f}" stroke="#e6e6e6"/>',
                    f'<text x="{left - 10}" y="{py + 4:.1f}" text-anchor="end" font-family="Arial" font-size="12">{_angle_tick_label(tick)}</text>',
                )
            )
        if signed_values and lower < 0 < upper:
            parts.append(
                f'<line x1="{left}" y1="{baseline_y:.1f}" x2="{left + chart_width}" y2="{baseline_y:.1f}" stroke="#555" stroke-width="1.5" stroke-dasharray="6 4"/>'
            )

        day_spacing = chart_width / len(days)
        group_count = len(groups)
        bar_width = min(64.0, day_spacing * 0.28 / max(group_count, 1))
        bar_gap = min(16.0, bar_width * 0.28)
        group_span = group_count * bar_width + max(group_count - 1, 0) * bar_gap
        for day_index, day in enumerate(days):
            day_x = left + day_spacing * (day_index + 0.5)
            parts.extend(
                (
                    f'<line x1="{day_x:.1f}" y1="{top}" x2="{day_x:.1f}" y2="{top + chart_height}" stroke="#f0f0f0"/>',
                    f'<text x="{day_x:.1f}" y="{top + chart_height + 26}" text-anchor="middle" font-family="Arial" font-size="13">{html.escape(day)}</text>',
                )
            )
            for group_index, group in enumerate(groups):
                mean = group.means.get(day)
                if mean is None or not math.isfinite(mean):
                    continue
                color = _condition_color(group.condition)
                center_offset = -group_span / 2 + bar_width / 2 + group_index * (bar_width + bar_gap)
                bar_center = day_x + center_offset
                bar_left = bar_center - bar_width / 2
                standard_deviation = group.standard_deviations.get(day)
                error_top = mean
                error_bottom = mean
                upper_y: float | None = None
                lower_y: float | None = None
                if standard_deviation is not None and math.isfinite(standard_deviation):
                    error_top = min(upper, mean + standard_deviation)
                    error_bottom = max(lower, mean - standard_deviation)
                    upper_y, lower_y = y(error_top), y(error_bottom)
                    parts.append(
                        f'<rect x="{bar_left:.1f}" y="{upper_y:.1f}" width="{bar_width:.1f}" height="{max(0.0, lower_y - upper_y):.1f}" fill="{ERROR_RANGE_COLOR}" fill-opacity="0.24" stroke="none"/>'
                    )

                mean_y = y(mean)
                bar_y = min(mean_y, baseline_y)
                bar_height = abs(baseline_y - mean_y)
                if standard_deviation is None:
                    tooltip_text = (
                        f"{group.condition.title()} | {day}: mean {mean:.3f}{value_suffix}, "
                        f"SD unavailable (n={group.mouse_counts.get(day, 0)})"
                    )
                else:
                    tooltip_text = (
                        f"{group.condition.title()} | {day}: mean {mean:.3f}{value_suffix}, "
                        f"SD {standard_deviation:.3f}{value_suffix}"
                    )
                parts.append(
                    f'<rect x="{bar_left:.1f}" y="{bar_y:.1f}" width="{bar_width:.1f}" height="{bar_height:.1f}" fill="{color}" cursor="help"><title>{html.escape(tooltip_text)}</title></rect>'
                )
                if upper_y is not None and lower_y is not None:
                    # Draw the whisker after the bar so its lower SD endpoint
                    # remains visible when it falls inside the colored bar.
                    parts.extend(
                        (
                            f'<line x1="{bar_center:.1f}" y1="{upper_y:.1f}" x2="{bar_center:.1f}" y2="{lower_y:.1f}" stroke="{ERROR_WHISKER_COLOR}" stroke-width="1.75"/>',
                            f'<line x1="{bar_center - 7:.1f}" y1="{upper_y:.1f}" x2="{bar_center + 7:.1f}" y2="{upper_y:.1f}" stroke="{ERROR_WHISKER_COLOR}" stroke-width="1.75"/>',
                            f'<line x1="{bar_center - 7:.1f}" y1="{lower_y:.1f}" x2="{bar_center + 7:.1f}" y2="{lower_y:.1f}" stroke="{ERROR_WHISKER_COLOR}" stroke-width="1.75"/>',
                        )
                    )
                label = f"{mean:+.3f}{value_suffix}" if signed_values else f"{mean:.3f}{value_suffix}"
                parts.append(
                    f'<text x="{bar_center:.1f}" y="{max(top + 14, y(error_top) - 8):.1f}" text-anchor="middle" font-family="Arial" font-size="11" font-weight="bold" fill="{ERROR_WHISKER_COLOR}">{html.escape(label)}</text>'
                )

        legend_x = left + chart_width + 42
        parts.append(
            f'<text x="{legend_x}" y="{top + 4}" font-family="Arial" font-size="15" font-weight="bold">Groups</text>'
        )
        for index, group in enumerate(groups):
            color = _condition_color(group.condition)
            text_y = top + 32 + index * 30
            count = max(group.mouse_counts.values(), default=0)
            parts.extend(
                (
                    f'<rect x="{legend_x}" y="{text_y - 12}" width="18" height="12" fill="{color}"/>',
                    f'<text x="{legend_x + 28}" y="{text_y}" font-family="Arial" font-size="13">{html.escape(group.condition.title())} (up to n={count} mice)</text>',
                )
            )
        parts.extend(
            (
                f'<text x="{left + chart_width / 2:.1f}" y="{height - 34}" text-anchor="middle" font-family="Arial" font-size="15">Experimental day</text>',
                f'<text x="25" y="{top + chart_height / 2:.1f}" text-anchor="middle" font-family="Arial" font-size="15" transform="rotate(-90 25 {top + chart_height / 2:.1f})">{html.escape(y_label)}</text>',
                "</svg>",
            )
        )
        return "\n".join(parts)

    @staticmethod
    def _daily_within_trial_variation_svg(
        dataset: str,
        day: str,
        trial_variations: Iterable[TrialAngleWithinTrialVariation],
        group_summaries: Iterable[GroupAngleWithinTrialVariation],
    ) -> str:
        """Show individual-run position-controlled angle SDs for one day."""
        trials = sorted(trial_variations, key=_trial_variation_sort_key)
        groups = sorted(group_summaries, key=lambda item: _condition_order(item.condition))
        values = [item.mean_standard_deviation_degrees for item in trials]
        for group in groups:
            values.append(group.mean_standard_deviation_degrees)
            if group.standard_deviation_degrees is not None:
                values.extend(
                    (
                        group.mean_standard_deviation_degrees - group.standard_deviation_degrees,
                        group.mean_standard_deviation_degrees + group.standard_deviation_degrees,
                    )
                )
        lower, upper = _nonnegative_axis_bounds(values)
        width, height = 1120, 740
        left, right, top, bottom = 96, 260, 94, 140
        chart_width, chart_height = width - left - right, height - top - bottom

        def y(value: float) -> float:
            return top + chart_height * (upper - value) / (upper - lower)

        parts = [
            f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}">',
            '<rect width="100%" height="100%" fill="white"/>',
            f'<text x="{left}" y="32" font-family="Arial" font-size="22" font-weight="bold">Tail-angle within-trial variation - {html.escape(dataset)} {html.escape(day)}</text>',
            f'<text x="{left}" y="56" font-family="Arial" font-size="13" fill="#444">Each dot is one trial\'s mean frame-level tail-angle SD across its valid 10 cm beam bins. Lower SD means more stable position-controlled tail posture.</text>',
            f'<text x="{left}" y="76" font-family="Arial" font-size="13" fill="#444">Large points are equal-mouse group means; neutral-gray ranges and whiskers are plus/minus 1 sample SD across mice.</text>',
        ]
        if not groups:
            parts.extend(
                (
                    f'<text x="{left}" y="{top + 50}" font-family="Arial" font-size="16" fill="#444">No SHAM or STROKE within-trial tail-angle variation data are available yet.</text>',
                    "</svg>",
                )
            )
            return "\n".join(parts)

        for tick in _angle_axis_ticks(lower, upper):
            py = y(tick)
            parts.extend(
                (
                    f'<line x1="{left}" y1="{py:.1f}" x2="{left + chart_width}" y2="{py:.1f}" stroke="#e6e6e6"/>',
                    f'<text x="{left - 10}" y="{py + 4:.1f}" text-anchor="end" font-family="Arial" font-size="12">{_angle_tick_label(tick)}</text>',
                )
            )
        group_x = {
            group.condition: left + chart_width * (index + 0.5) / len(groups)
            for index, group in enumerate(groups)
        }
        for group in groups:
            px = group_x[group.condition]
            parts.extend(
                (
                    f'<line x1="{px:.1f}" y1="{top}" x2="{px:.1f}" y2="{top + chart_height}" stroke="#f0f0f0"/>',
                    f'<text x="{px:.1f}" y="{top + chart_height + 26}" text-anchor="middle" font-family="Arial" font-size="14">{html.escape(group.condition.title())}</text>',
                )
            )
        for trial in trials:
            px = group_x.get(trial.condition)
            if px is None:
                continue
            jitter = _trial_dot_jitter(trial.cage, trial.animal, trial.trial)
            tooltip = html.escape(
                f"{trial.condition.title()} | Cage {trial.cage} Mouse {trial.animal} | Trial T{trial.trial} | "
                f"mean bin SD {trial.mean_standard_deviation_degrees:.3f} deg across {trial.bin_count} bin(s)"
            )
            parts.append(
                f'<circle cx="{px + jitter:.1f}" cy="{y(trial.mean_standard_deviation_degrees):.1f}" r="5" fill="{_condition_color(trial.condition)}" fill-opacity="0.38" cursor="help"><title>{tooltip}</title></circle>'
            )
        for group in groups:
            px = group_x[group.condition]
            mean = group.mean_standard_deviation_degrees
            error = group.standard_deviation_degrees
            upper_y: float | None = None
            lower_y: float | None = None
            if error is not None:
                upper_y = y(min(upper, mean + error))
                lower_y = y(max(lower, mean - error))
                parts.append(
                    f'<rect x="{px - 22:.1f}" y="{upper_y:.1f}" width="44" height="{max(0.0, lower_y - upper_y):.1f}" fill="{ERROR_RANGE_COLOR}" fill-opacity="0.24" stroke="none"/>'
                )
            if upper_y is not None and lower_y is not None:
                parts.extend(
                    (
                        f'<line x1="{px:.1f}" y1="{upper_y:.1f}" x2="{px:.1f}" y2="{lower_y:.1f}" stroke="{ERROR_WHISKER_COLOR}" stroke-width="1.75"/>',
                        f'<line x1="{px - 7:.1f}" y1="{upper_y:.1f}" x2="{px + 7:.1f}" y2="{upper_y:.1f}" stroke="{ERROR_WHISKER_COLOR}" stroke-width="1.75"/>',
                        f'<line x1="{px - 7:.1f}" y1="{lower_y:.1f}" x2="{px + 7:.1f}" y2="{lower_y:.1f}" stroke="{ERROR_WHISKER_COLOR}" stroke-width="1.75"/>',
                    )
                )
            parts.append(
                f'<circle cx="{px:.1f}" cy="{y(mean):.1f}" r="7" fill="{_condition_color(group.condition)}" stroke="white" stroke-width="1.5"><title>{html.escape(group.condition.title())} mean: {mean:.3f} deg; n={group.mouse_count} mice</title></circle>'
            )

        legend_x = left + chart_width + 40
        parts.append(
            f'<text x="{legend_x}" y="{top + 4}" font-family="Arial" font-size="15" font-weight="bold">Groups</text>'
        )
        for index, group in enumerate(groups):
            text_y = top + 30 + index * 28
            parts.extend(
                (
                    f'<circle cx="{legend_x + 12}" cy="{text_y - 5}" r="5" fill="{_condition_color(group.condition)}"/>',
                    f'<text x="{legend_x + 30}" y="{text_y}" font-family="Arial" font-size="13">{html.escape(group.condition.title())} (n={group.mouse_count} mice)</text>',
                )
            )
        parts.extend(
            (
                f'<text x="{left + chart_width / 2:.1f}" y="{height - 34}" text-anchor="middle" font-family="Arial" font-size="15">Experimental group</text>',
                f'<text x="25" y="{top + chart_height / 2:.1f}" text-anchor="middle" font-family="Arial" font-size="15" transform="rotate(-90 25 {top + chart_height / 2:.1f})">Within-trial tail-angle SD (degrees)</text>',
                "</svg>",
            )
        )
        return "\n".join(parts)

    @staticmethod
    def _read_trial_series(
        workbook_path: Path,
        conditions: dict[tuple[str, str], str],
    ) -> list[TrialAngleSeries]:
        from openpyxl import load_workbook

        workbook = load_workbook(workbook_path, data_only=True, read_only=True)
        try:
            if ANGLE_SHEET not in workbook.sheetnames:
                return []
            sheet = workbook[ANGLE_SHEET]
            headers = [cell.value for cell in next(sheet.iter_rows(max_row=1), ())]
            index = {str(header): column for column, header in enumerate(headers) if header is not None}
            required = ("Day", "Cage", "Animal", "Trial")
            if not all(header in index for header in required):
                return []
            compact, separate_angles, separate_positions = _frame_columns(index)

            output: list[TrialAngleSeries] = []
            for values in sheet.iter_rows(min_row=2, values_only=True):
                day = _cell_text(values, index["Day"])
                cage = _cell_text(values, index["Cage"])
                animal = _cell_text(values, index["Animal"])
                trial = _cell_int(values, index["Trial"])
                if not (day and cage and animal and trial):
                    continue

                by_bin: dict[int, list[float]] = defaultdict(list)
                for frame in set(compact) | set(separate_angles) | set(separate_positions):
                    angle, position = _frame_measurement(
                        values,
                        frame,
                        compact,
                        separate_angles,
                        separate_positions,
                    )
                    if angle is None or position is None:
                        continue
                    bin_start = _bin_start(position)
                    if bin_start is not None:
                        by_bin[bin_start].append(angle)
                if not by_bin:
                    continue
                output.append(
                    TrialAngleSeries(
                        day=day,
                        cage=cage,
                        animal=animal,
                        condition=conditions.get((cage, animal), "UNASSIGNED"),
                        trial=trial,
                        bin_means=tuple(
                            statistics.fmean(by_bin[start]) if by_bin[start] else None
                            for start in BIN_STARTS_CM
                        ),
                        frame_angles_by_bin=tuple(
                            tuple(by_bin[start]) for start in BIN_STARTS_CM
                        ),
                    )
                )
            return output
        finally:
            workbook.close()

    @staticmethod
    def _write_svg(path: Path, text: str) -> None:
        temporary_path = path.with_suffix(f"{path.suffix}.tmp")
        try:
            temporary_path.write_text(text, encoding="utf-8")
            temporary_path.replace(path)
        finally:
            if temporary_path.exists():
                temporary_path.unlink()

    @staticmethod
    def _svg(
        dataset: str,
        day: str,
        trial_series: Iterable[TrialAngleSeries],
        mouse_series: Iterable[MouseAngleSeries],
        group_summaries: Iterable[GroupAngleSummary],
    ) -> str:
        trials = sorted(
            trial_series,
            key=lambda item: (_condition_order(item.condition), _numeric_key(item.cage), _numeric_key(item.animal), item.trial),
        )
        mice = sorted(
            mouse_series,
            key=lambda item: (_condition_order(item.condition), _numeric_key(item.cage), _numeric_key(item.animal)),
        )
        groups = sorted(group_summaries, key=lambda item: _condition_order(item.condition))
        values = [
            value
            for series in trials
            for value in series.bin_means
            if value is not None and math.isfinite(value)
        ]
        for summary in groups:
            for mean, error in zip(summary.means, summary.standard_deviations):
                if mean is not None:
                    values.append(mean)
                    if error is not None:
                        values.extend((mean - error, mean + error))
        lower, upper = _axis_bounds(values)

        width, height = 1420, 790
        left, right, top, bottom = 96, 390, 94, 150
        chart_width = width - left - right
        chart_height = height - top - bottom

        def x(position_cm: float) -> float:
            return left + chart_width * position_cm / 90

        def y(value: float) -> float:
            return top + chart_height * (upper - value) / (upper - lower)

        parts = [
            f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}">',
            '<rect width="100%" height="100%" fill="white"/>',
            f'<text x="{left}" y="32" font-family="Arial" font-size="22" font-weight="bold">Signed tail-angle group comparison - {html.escape(dataset)} {html.escape(day)}</text>',
            f'<text x="{left}" y="56" font-family="Arial" font-size="13" fill="#444">Thin lines are trial means within each 10 cm bin. Each mouse is averaged across its trials before it enters the group mean, so every mouse has equal weight.</text>',
            f'<text x="{left}" y="76" font-family="Arial" font-size="13" fill="#444">Thick lines are group means; shaded bands are plus/minus 1 sample SD across mouse means. Gaps indicate no valid observations.</text>',
        ]

        for tick in range(0, 91, 10):
            px = x(tick)
            parts.extend(
                (
                    f'<line x1="{px:.1f}" y1="{top}" x2="{px:.1f}" y2="{top + chart_height}" stroke="#e6e6e6"/>',
                    f'<text x="{px:.1f}" y="{top + chart_height + 25}" text-anchor="middle" font-family="Arial" font-size="12">{tick}</text>',
                )
            )
        for tick in range(int(lower), int(upper) + 1, 5):
            py = y(tick)
            parts.extend(
                (
                    f'<line x1="{left}" y1="{py:.1f}" x2="{left + chart_width}" y2="{py:.1f}" stroke="#e6e6e6"/>',
                    f'<text x="{left - 10}" y="{py + 4:.1f}" text-anchor="end" font-family="Arial" font-size="12">{tick}</text>',
                )
            )
        if lower < 0 < upper:
            zero = y(0)
            parts.append(
                f'<line x1="{left}" y1="{zero:.1f}" x2="{left + chart_width}" y2="{zero:.1f}" stroke="#555" stroke-width="1.5" stroke-dasharray="6 4"/>'
            )

        for summary in groups:
            color = _condition_color(summary.condition)
            for segment in _band_segments(summary, x, y, lower, upper):
                parts.append(
                    f'<polygon points="{" ".join(segment)}" fill="{color}" fill-opacity="0.18" stroke="none"/>'
                )

        for series in trials:
            color = _condition_color(series.condition)
            tooltip = html.escape(
                f"{series.condition.title()} | Cage {series.cage} Mouse {series.animal} | Trial T{series.trial}"
            )
            for segment in _line_segments(series.bin_means, x, y):
                parts.append(
                    f'<polyline points="{" ".join(segment)}" fill="none" stroke="{color}" stroke-width="1.7" stroke-opacity="0.28" pointer-events="none"/>'
                    f'<polyline points="{" ".join(segment)}" fill="none" stroke="{color}" stroke-width="8" stroke-opacity="0" pointer-events="stroke" cursor="help"><title>{tooltip}</title></polyline>'
                )

        for summary in groups:
            color = _condition_color(summary.condition)
            for segment in _line_segments(summary.means, x, y):
                parts.append(
                    f'<polyline points="{" ".join(segment)}" fill="none" stroke="{color}" stroke-width="4" stroke-linecap="round"/>'
                )

        legend_x = left + chart_width + 40
        parts.append(
            f'<text x="{legend_x}" y="{top + 4}" font-family="Arial" font-size="15" font-weight="bold">Groups</text>'
        )
        for index, summary in enumerate(groups):
            color = _condition_color(summary.condition)
            text_y = top + 30 + index * 28
            largest_n = max(summary.mouse_counts, default=0)
            parts.extend(
                (
                    f'<line x1="{legend_x}" y1="{text_y - 5}" x2="{legend_x + 24}" y2="{text_y - 5}" stroke="{color}" stroke-width="4"/>',
                    f'<text x="{legend_x + 33}" y="{text_y}" font-family="Arial" font-size="13">{html.escape(summary.condition.title())} (up to n={largest_n} mice)</text>',
                )
            )

        parts.extend(
            (
                f'<text x="{left + chart_width / 2:.1f}" y="{height - 34}" text-anchor="middle" font-family="Arial" font-size="15">Back-paw position on calibrated beam (cm)</text>',
                f'<text x="25" y="{top + chart_height / 2:.1f}" text-anchor="middle" font-family="Arial" font-size="15" transform="rotate(-90 25 {top + chart_height / 2:.1f})">Signed tail angle (degrees)</text>',
                "</svg>",
            )
        )
        return "\n".join(parts)


def summarize_group_angles(
    trial_series: Iterable[TrialAngleSeries],
) -> tuple[list[MouseAngleSeries], list[GroupAngleSummary]]:
    """Average T1-T3 within mouse before calculating group statistics."""
    by_mouse: dict[tuple[str, str, str, str], list[TrialAngleSeries]] = defaultdict(list)
    for series in trial_series:
        by_mouse[(series.day, series.cage, series.animal, series.condition)].append(series)

    mice: list[MouseAngleSeries] = []
    for (day, cage, animal, condition), trials in by_mouse.items():
        mice.append(
            MouseAngleSeries(
                day=day,
                cage=cage,
                animal=animal,
                condition=condition,
                bin_means=tuple(
                    statistics.fmean(values)
                    if (values := [trial.bin_means[index] for trial in trials if trial.bin_means[index] is not None])
                    else None
                    for index in range(len(BIN_STARTS_CM))
                ),
            )
        )

    by_group: dict[str, list[MouseAngleSeries]] = defaultdict(list)
    for mouse in mice:
        by_group[mouse.condition].append(mouse)
    summaries: list[GroupAngleSummary] = []
    for condition, group_mice in by_group.items():
        means: list[float | None] = []
        errors: list[float | None] = []
        counts: list[int] = []
        for index in range(len(BIN_STARTS_CM)):
            values = [mouse.bin_means[index] for mouse in group_mice if mouse.bin_means[index] is not None]
            counts.append(len(values))
            means.append(statistics.fmean(values) if values else None)
            errors.append(statistics.stdev(values) if len(values) >= 2 else None)
        summaries.append(
            GroupAngleSummary(
                condition=condition,
                means=tuple(means),
                standard_deviations=tuple(errors),
                mouse_counts=tuple(counts),
            )
        )
    return mice, summaries


def summarize_day_by_day_group_angles(
    trial_series: Iterable[TrialAngleSeries],
) -> list[DayGroupAngleSummary]:
    """Compare daily signed tail angles, weighting every mouse equally."""
    mouse_series, _ = summarize_group_angles(trial_series)
    values_by_group_day: dict[str, dict[str, list[float]]] = defaultdict(lambda: defaultdict(list))
    for mouse in mouse_series:
        condition = mouse.condition.upper()
        if condition not in {"SHAM", "STROKE"}:
            continue
        valid_bin_means = [
            value for value in mouse.bin_means if value is not None and math.isfinite(value)
        ]
        if valid_bin_means:
            values_by_group_day[condition][mouse.day].append(statistics.fmean(valid_bin_means))

    summaries: list[DayGroupAngleSummary] = []
    for condition, by_day in values_by_group_day.items():
        means = {day: statistics.fmean(values) for day, values in by_day.items() if values}
        standard_deviations = {
            day: statistics.stdev(values) if len(values) >= 2 else None
            for day, values in by_day.items()
            if values
        }
        summaries.append(
            DayGroupAngleSummary(
                condition=condition,
                means=means,
                standard_deviations=standard_deviations,
                mouse_counts={day: len(values) for day, values in by_day.items() if values},
            )
        )
    return summaries


def summarize_tail_angle_within_trial_variation(
    trial_series: Iterable[TrialAngleSeries],
) -> tuple[
    list[TrialAngleWithinTrialVariation],
    list[MouseAngleWithinTrialVariation],
    list[GroupAngleWithinTrialVariation],
]:
    """Measure frame-to-frame posture variation without mixing beam positions.

    For each trial, a sample SD is calculated from the frame-level signed angle
    values in each valid 10 cm bin.  The trial's displayed value is its mean
    bin SD.  A mouse then averages *all* valid trial/bin SDs, so it contributes
    equally to its SHAM/STROKE group even if it has more valid trials than
    another mouse.
    """
    trial_variations: list[TrialAngleWithinTrialVariation] = []
    for series in trial_series:
        bin_standard_deviations = tuple(
            statistics.stdev(values)
            for values in series.frame_angles_by_bin
            if len(values) >= 2
        )
        if not bin_standard_deviations:
            continue
        trial_variations.append(
            TrialAngleWithinTrialVariation(
                day=series.day,
                cage=series.cage,
                animal=series.animal,
                condition=series.condition.upper(),
                trial=series.trial,
                mean_standard_deviation_degrees=statistics.fmean(bin_standard_deviations),
                bin_standard_deviations=bin_standard_deviations,
            )
        )

    by_mouse: dict[
        tuple[str, str, str, str], list[TrialAngleWithinTrialVariation]
    ] = defaultdict(list)
    for trial in trial_variations:
        by_mouse[(trial.day, trial.cage, trial.animal, trial.condition)].append(trial)
    mice: list[MouseAngleWithinTrialVariation] = []
    for (day, cage, animal, condition), mouse_trials in by_mouse.items():
        bin_values = [
            value
            for trial in mouse_trials
            for value in trial.bin_standard_deviations
            if math.isfinite(value)
        ]
        if not bin_values:
            continue
        mice.append(
            MouseAngleWithinTrialVariation(
                day=day,
                cage=cage,
                animal=animal,
                condition=condition,
                mean_standard_deviation_degrees=statistics.fmean(bin_values),
                trial_count=len(mouse_trials),
                bin_observation_count=len(bin_values),
            )
        )

    by_group: dict[str, list[MouseAngleWithinTrialVariation]] = defaultdict(list)
    for mouse in mice:
        if mouse.condition in {"SHAM", "STROKE"}:
            by_group[mouse.condition].append(mouse)
    groups: list[GroupAngleWithinTrialVariation] = []
    for condition, group_mice in by_group.items():
        values = [mouse.mean_standard_deviation_degrees for mouse in group_mice]
        groups.append(
            GroupAngleWithinTrialVariation(
                condition=condition,
                mean_standard_deviation_degrees=statistics.fmean(values),
                standard_deviation_degrees=(
                    statistics.stdev(values) if len(values) >= 2 else None
                ),
                mouse_count=len(values),
            )
        )
    return trial_variations, mice, groups


def summarize_day_by_day_tail_angle_within_trial_variation(
    trial_series: Iterable[TrialAngleSeries],
) -> list[DayGroupAngleSummary]:
    """Compare equal-mouse, position-controlled within-run variation by day."""
    _trials, mice, _groups = summarize_tail_angle_within_trial_variation(trial_series)
    values_by_group_day: dict[str, dict[str, list[float]]] = defaultdict(
        lambda: defaultdict(list)
    )
    for mouse in mice:
        if mouse.condition in {"SHAM", "STROKE"}:
            values_by_group_day[mouse.condition][mouse.day].append(
                mouse.mean_standard_deviation_degrees
            )

    summaries: list[DayGroupAngleSummary] = []
    for condition, by_day in values_by_group_day.items():
        summaries.append(
            DayGroupAngleSummary(
                condition=condition,
                means={day: statistics.fmean(values) for day, values in by_day.items() if values},
                standard_deviations={
                    day: statistics.stdev(values) if len(values) >= 2 else None
                    for day, values in by_day.items()
                    if values
                },
                mouse_counts={day: len(values) for day, values in by_day.items() if values},
            )
        )
    return summaries


def _frame_columns(index: dict[str, int]) -> tuple[dict[int, int], dict[int, int], dict[int, int]]:
    compact: dict[int, int] = {}
    separate_angles: dict[int, int] = {}
    separate_positions: dict[int, int] = {}
    for header, column in index.items():
        if not header.startswith("Frame "):
            continue
        try:
            if header.endswith(" angle (deg)"):
                separate_angles[int(header.removeprefix("Frame ").removesuffix(" angle (deg)"))] = column
            elif header.endswith(" back paw (cm)"):
                separate_positions[int(header.removeprefix("Frame ").removesuffix(" back paw (cm)"))] = column
            else:
                compact[int(header.removeprefix("Frame "))] = column
        except ValueError:
            continue
    return compact, separate_angles, separate_positions


def _frame_measurement(
    values: tuple[object, ...],
    frame: int,
    compact: dict[int, int],
    separate_angles: dict[int, int],
    separate_positions: dict[int, int],
) -> tuple[float | None, float | None]:
    if frame in compact:
        column = compact[frame]
        if column >= len(values) or values[column] is None:
            return None, None
        match = FRAME_CELL_RE.fullmatch(str(values[column]).strip())
        if match is None:
            return None, None
        return float(match.group(1)), float(match.group(2))
    return _cell_float(values, separate_angles.get(frame)), _cell_float(values, separate_positions.get(frame))


def _bin_start(position_cm: float) -> int | None:
    if not 0.0 <= position_cm <= 90.0:
        return None
    return min(int(position_cm // 10) * 10, 80)


def _axis_bounds(values: list[float]) -> tuple[float, float]:
    lower_value = min(values + [0.0])
    upper_value = max(values + [0.0])
    lower = math.floor(lower_value / 5.0) * 5.0
    upper = math.ceil(upper_value / 5.0) * 5.0
    if upper - lower < 10.0:
        lower -= 5.0
        upper += 5.0
    return lower, upper


def _nonnegative_axis_bounds(values: Iterable[float]) -> tuple[float, float]:
    valid = [float(value) for value in values if math.isfinite(value)]
    maximum = max(valid, default=0.0)
    if maximum <= 0.0:
        return 0.0, 1.0
    step = _nice_axis_step((maximum * 1.1) / 6.0)
    upper = max(step * 2.0, math.ceil((maximum * 1.1) / step) * step)
    return 0.0, upper


def _nice_axis_step(target: float) -> float:
    if target <= 0.0 or not math.isfinite(target):
        return 1.0
    magnitude = 10 ** math.floor(math.log10(target))
    normalized = target / magnitude
    for candidate in (1.0, 2.0, 5.0, 10.0):
        if normalized <= candidate:
            return candidate * magnitude
    return 10.0 * magnitude


def _angle_axis_ticks(lower: float, upper: float) -> list[float]:
    step = _nice_axis_step((upper - lower) / 6.0)
    first = math.floor(lower / step) * step
    ticks: list[float] = []
    value = first
    while value <= upper + (step * 0.001):
        if value >= lower - (step * 0.001):
            ticks.append(value)
        value += step
    return ticks


def _angle_tick_label(value: float) -> str:
    if math.isclose(value, round(value), abs_tol=1e-9):
        return str(int(round(value)))
    if abs(value) >= 1.0:
        return f"{value:.1f}"
    return f"{value:.2f}"


def _line_segments(
    values: Iterable[float | None],
    x,
    y,
) -> list[list[str]]:
    segments: list[list[str]] = []
    segment: list[str] = []
    for start, value in zip(BIN_STARTS_CM, values):
        if value is None or not math.isfinite(value):
            if segment:
                segments.append(segment)
                segment = []
            continue
        segment.append(f"{x(start + 5):.1f},{y(value):.1f}")
    if segment:
        segments.append(segment)
    return segments


def _band_segments(summary: GroupAngleSummary, x, y, lower: float, upper: float) -> list[list[str]]:
    segments: list[list[tuple[float, float, float]]] = []
    segment: list[tuple[float, float, float]] = []
    for start, mean, error in zip(BIN_STARTS_CM, summary.means, summary.standard_deviations):
        if mean is None or error is None:
            if segment:
                segments.append(segment)
                segment = []
            continue
        segment.append((x(start + 5), mean, error))
    if segment:
        segments.append(segment)
    return [
        [
            *(f"{px:.1f},{y(min(upper, mean + error)):.1f}" for px, mean, error in segment),
            *(f"{px:.1f},{y(max(lower, mean - error)):.1f}" for px, mean, error in reversed(segment)),
        ]
        for segment in segments
    ]


def _day_line_segments(
    summary: DayGroupAngleSummary,
    days: Iterable[str],
    x,
    y,
) -> list[list[str]]:
    segments: list[list[str]] = []
    segment: list[str] = []
    for day in days:
        mean = summary.means.get(day)
        if mean is None or not math.isfinite(mean):
            if segment:
                segments.append(segment)
                segment = []
            continue
        segment.append(f"{x(day):.1f},{y(mean):.1f}")
    if segment:
        segments.append(segment)
    return segments


def _day_band_segments(
    summary: DayGroupAngleSummary,
    days: Iterable[str],
    x,
    y,
    lower: float,
    upper: float,
) -> list[list[str]]:
    segments: list[list[tuple[float, float, float]]] = []
    segment: list[tuple[float, float, float]] = []
    for day in days:
        mean = summary.means.get(day)
        error = summary.standard_deviations.get(day)
        if mean is None or error is None or not math.isfinite(mean) or not math.isfinite(error):
            if segment:
                segments.append(segment)
                segment = []
            continue
        segment.append((x(day), mean, error))
    if segment:
        segments.append(segment)
    polygons: list[list[str]] = []
    for segment in segments:
        if len(segment) == 1:
            # A one-day chart has no horizontal span for a conventional ribbon.
            # Render the SD as a narrow shaded column around that day's point.
            px, mean, error = segment[0]
            half_width = 10.0
            upper_y = y(min(upper, mean + error))
            lower_y = y(max(lower, mean - error))
            polygons.append(
                [
                    f"{px - half_width:.1f},{upper_y:.1f}",
                    f"{px + half_width:.1f},{upper_y:.1f}",
                    f"{px + half_width:.1f},{lower_y:.1f}",
                    f"{px - half_width:.1f},{lower_y:.1f}",
                ]
            )
            continue
        polygons.append(
            [
                *(f"{px:.1f},{y(min(upper, mean + error)):.1f}" for px, mean, error in segment),
                *(f"{px:.1f},{y(max(lower, mean - error)):.1f}" for px, mean, error in reversed(segment)),
            ]
        )
    return polygons


def _cell_text(values: tuple[object, ...], index: int | None) -> str:
    if index is None or index >= len(values) or values[index] is None:
        return ""
    value = values[index]
    return str(int(value)) if isinstance(value, float) and value.is_integer() else str(value).strip()


def _cell_float(values: tuple[object, ...], index: int | None) -> float | None:
    if index is None or index >= len(values) or values[index] is None:
        return None
    try:
        value = float(values[index])
    except (TypeError, ValueError):
        return None
    return value if math.isfinite(value) else None


def _cell_int(values: tuple[object, ...], index: int | None) -> int:
    value = _cell_float(values, index)
    return int(value) if value is not None else 0


def _condition_color(condition: str) -> str:
    return {"SHAM": SHAM_COLOR, "STROKE": STROKE_COLOR}.get(condition.upper(), UNASSIGNED_COLOR)


def _condition_order(condition: str) -> int:
    return {"SHAM": 0, "STROKE": 1}.get(condition.upper(), 2)


def _numeric_key(value: str) -> tuple[int, int | str]:
    return (0, int(value)) if value.isdigit() else (1, value.casefold())


def _trial_variation_sort_key(series: TrialAngleWithinTrialVariation):
    return (
        _condition_order(series.condition),
        _numeric_key(series.cage),
        _numeric_key(series.animal),
        series.trial,
    )


def _trial_dot_jitter(cage: str, animal: str, trial: int) -> float:
    """Return a stable small offset so individual trial points stay visible."""
    identifiers = f"{cage}|{animal}|{trial}"
    checksum = sum(
        (index + 1) * ord(character) for index, character in enumerate(identifiers)
    )
    return float((checksum % 13) - 6) * 4.0
