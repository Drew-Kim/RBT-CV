"""SVG group comparisons for frame-wise tail-curvature measurements.

Tail curvature uses the same binning and equal-mouse statistical treatment as
the other continuous research metrics.  The shared plotting primitives live in
``back_front_paw_distance_plot`` because both metrics are nonnegative,
frame-wise values paired with a calibrated back-paw position.
"""

from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
import html
import math
from pathlib import Path
import re
import statistics
from typing import Iterable

from .back_front_paw_distance_plot import (
    BackFrontPawDistancePlotStore,
    DayGroupPawDistanceSummary,
    GroupPawDistanceSummary,
    MousePawDistanceSeries,
    TrialPawDistanceSeries,
    _all_bin_values,
    _append_bin_axes,
    _append_group_legend,
    _append_hover_lines,
    _append_summary_bands,
    _append_summary_lines,
    _bin_start,
    _bin_x,
    _axis_ticks,
    _cell_float,
    _cell_int,
    _cell_text,
    _condition_color,
    _condition_order,
    _summary_edges,
    _svg_header,
    _tick_label,
    _trial_sort_key,
    _valid,
    _value_y,
    _zoomed_nonnegative_axis_bounds,
    summarize_day_by_day_back_front_paw_distances,
    summarize_group_back_front_paw_distances,
)
from .condition_map import ConditionMapStore
from .dataset import ROOT, natural_day_key


OUTPUT_ROOT = ROOT / "outputs"
RESULTS_FILENAME = "RBT_CV_Results.xlsx"
SHEET_NAME = "Tail Curvature"
CHART_DIRECTORY = "Tail Curve"
FILE_SUFFIX = ".svg"
GROUP_PREFIX = "tail_curvature_group_comparison_"
WITHIN_TRIAL_VARIATION_PREFIX = "tail_curvature_within_trial_variation_"
DAY_GROUP_FILENAME = "Day_tail_curvature_group_comparison.svg"
DAY_WITHIN_TRIAL_VARIATION_FILENAME = "Day_tail_curvature_within_trial_variation.svg"
FRAME_CELL_RE = re.compile(r"([+-]?\d+(?:\.\d+)?) deg; ([+-]?\d+(?:\.\d+)?) cm")

# Keep uncertainty visually separate from the SHAM/STROKE bar colors so both
# ends of each sample-SD range remain legible over a colored bar.
ERROR_RANGE_COLOR = "#6B7280"
ERROR_WHISKER_COLOR = "#374151"

# These aliases communicate the metric while reusing the deliberately generic
# binned-series data shape and equal-mouse summary calculations.
TrialTailCurvatureSeries = TrialPawDistanceSeries
MouseTailCurvatureSeries = MousePawDistanceSeries
GroupTailCurvatureSummary = GroupPawDistanceSummary


@dataclass(frozen=True)
class TrialTailCurvatureWithinTrialVariation:
    """Frame-level curvature variation within one trial and 10 cm bin.

    Each value is a sample SD calculated only from valid frames inside that
    beam-position bin. This avoids treating normal posture changes along the
    beam as instability.
    """

    day: str
    cage: str
    animal: str
    condition: str
    trial: int
    standard_deviations: tuple[float | None, ...]
    frame_counts: tuple[int, ...]


class TailCurvaturePlotStore(BackFrontPawDistancePlotStore):
    """Create tail-curvature comparison and within-trial-variation charts."""

    def __init__(self, output_root: Path = OUTPUT_ROOT) -> None:
        super().__init__(output_root)

    def chart_dir(self, dataset: str) -> Path:
        return self.result_dir(dataset) / CHART_DIRECTORY

    def refresh_dataset(self, dataset: str) -> dict[str, Path]:
        workbook_path = self.result_dir(dataset) / RESULTS_FILENAME
        if not workbook_path.exists():
            return {}

        conditions = ConditionMapStore(self.output_root).load(dataset)
        trials = self._read_trial_series(workbook_path, conditions)
        within_trial_variation = self._read_within_trial_variation_series(
            workbook_path,
            conditions,
        )
        by_day: dict[str, list[TrialTailCurvatureSeries]] = defaultdict(list)
        for series in trials:
            by_day[series.day].append(series)
        variation_by_day: dict[str, list[TrialTailCurvatureWithinTrialVariation]] = defaultdict(list)
        for series in within_trial_variation:
            variation_by_day[series.day].append(series)

        chart_dir = self.chart_dir(dataset)
        chart_dir.mkdir(parents=True, exist_ok=True)
        paths: dict[str, Path] = {}
        for day, day_trials in sorted(by_day.items(), key=lambda item: natural_day_key(item[0])):
            mice, groups = summarize_group_back_front_paw_distances(day_trials)
            group_path = chart_dir / f"{GROUP_PREFIX}{day}{FILE_SUFFIX}"
            self._write_svg(
                group_path,
                self._daily_group_svg(dataset, day, day_trials, mice, groups),
            )
            paths[f"group:{day}"] = group_path

            variation_trials = variation_by_day.get(day, [])
            variation_mice, variation_groups = self._within_trial_group_summaries(
                variation_trials
            )
            variation_path = chart_dir / f"{WITHIN_TRIAL_VARIATION_PREFIX}{day}{FILE_SUFFIX}"
            self._write_svg(
                variation_path,
                self._daily_within_trial_variation_svg(
                    dataset,
                    day,
                    variation_trials,
                    variation_mice,
                    variation_groups,
                ),
            )
            paths[f"within_trial_variation:{day}"] = variation_path

        day_group_path = chart_dir / DAY_GROUP_FILENAME
        self._write_svg(
            day_group_path,
            self._day_bar_svg(
                dataset,
                summarize_day_by_day_back_front_paw_distances(trials),
                title="Day-by-day tail-curvature comparison",
                calculation_note=(
                    "For each day, each mouse is averaged across its trials and valid "
                    "10 cm bins from 0-90 cm before group calculation."
                ),
                comparison_note=(
                    "Colored bars are SHAM/STROKE mean curvature; neutral-gray ranges and "
                    "whiskers are plus/minus 1 sample SD across mice."
                ),
                empty_note="No SHAM or STROKE tail-curvature data are available yet.",
                y_label="Tail curvature (degrees)",
                value_suffix=" deg",
            ),
        )
        paths["day_group"] = day_group_path

        day_variation_path = chart_dir / DAY_WITHIN_TRIAL_VARIATION_FILENAME
        self._write_svg(
            day_variation_path,
            self._day_bar_svg(
                dataset,
                self._day_within_trial_variation_summaries(within_trial_variation),
                title="Day-by-day tail-curvature within-trial variation",
                calculation_note=(
                    "For each trial and 10 cm bin, calculate sample SD across raw valid "
                    "frame curvatures. Each mouse contributes the mean of its trial/bin SDs."
                ),
                comparison_note=(
                    "Neutral-gray ranges and whiskers are plus/minus 1 sample SD across "
                    "mice. Lower SD means more consistent posture within a run."
                ),
                empty_note="No SHAM or STROKE tail-curvature variation data are available yet.",
                y_label="Within-trial tail-curvature SD (degrees; lower = more consistent)",
                value_suffix=" deg",
            ),
        )
        paths["day_within_trial_variation"] = day_variation_path

        live_names = {path.name for path in paths.values()}
        for prefix in (GROUP_PREFIX, WITHIN_TRIAL_VARIATION_PREFIX):
            for stale_path in chart_dir.glob(f"{prefix}*{FILE_SUFFIX}"):
                if stale_path.name not in live_names:
                    stale_path.unlink()
        # Delete only known obsolete generated trial-to-trial consistency
        # files. The posture-variation charts above are the replacement.
        for obsolete_path in chart_dir.glob(f"tail_curvature_trial_consistency_*{FILE_SUFFIX}"):
            obsolete_path.unlink()
        obsolete_day_consistency = chart_dir / "Day_tail_curvature_trial_consistency.svg"
        if obsolete_day_consistency.exists():
            obsolete_day_consistency.unlink()
        return paths

    @staticmethod
    def _read_trial_series(
        workbook_path: Path,
        conditions: dict[tuple[str, str], str],
    ) -> list[TrialTailCurvatureSeries]:
        from openpyxl import load_workbook

        workbook = load_workbook(workbook_path, data_only=True, read_only=True)
        try:
            if SHEET_NAME not in workbook.sheetnames:
                return []
            sheet = workbook[SHEET_NAME]
            headers = [cell.value for cell in next(sheet.iter_rows(max_row=1), ())]
            index = {
                str(header): position
                for position, header in enumerate(headers)
                if header is not None
            }
            required = ("Day", "Cage", "Animal", "Trial")
            if not all(header in index for header in required):
                return []
            compact, curvature_columns, position_columns = _frame_columns(index)

            output: list[TrialTailCurvatureSeries] = []
            for values in sheet.iter_rows(min_row=2, values_only=True):
                day = _cell_text(values, index["Day"])
                cage = _cell_text(values, index["Cage"])
                animal = _cell_text(values, index["Animal"])
                trial = _cell_int(values, index["Trial"])
                if not (day and cage and animal and trial):
                    continue

                by_bin: dict[int, list[float]] = defaultdict(list)
                for frame in set(compact) | set(curvature_columns) | set(position_columns):
                    curvature, position = _frame_measurement(
                        values,
                        frame,
                        compact,
                        curvature_columns,
                        position_columns,
                    )
                    if curvature is None or position is None:
                        continue
                    bin_start = _bin_start(position)
                    if bin_start is not None:
                        by_bin[bin_start].append(curvature)
                if not by_bin:
                    continue
                output.append(
                    TrialTailCurvatureSeries(
                        day=day,
                        cage=cage,
                        animal=animal,
                        condition=conditions.get((cage, animal), "UNASSIGNED"),
                        trial=trial,
                        bin_means=tuple(
                            statistics.fmean(by_bin[start]) if by_bin[start] else None
                            for start in range(0, 90, 10)
                        ),
                    )
                )
            return output
        finally:
            workbook.close()

    @staticmethod
    def _read_within_trial_variation_series(
        workbook_path: Path,
        conditions: dict[tuple[str, str], str],
    ) -> list[TrialTailCurvatureWithinTrialVariation]:
        """Read raw frames and calculate position-controlled curvature SDs.

        A value is available only when a 10 cm bin contains at least two valid
        frames. No values are imputed across missing frames, bins, or falls.
        """
        from openpyxl import load_workbook

        workbook = load_workbook(workbook_path, data_only=True, read_only=True)
        try:
            if SHEET_NAME not in workbook.sheetnames:
                return []
            sheet = workbook[SHEET_NAME]
            headers = [cell.value for cell in next(sheet.iter_rows(max_row=1), ())]
            index = {
                str(header): position
                for position, header in enumerate(headers)
                if header is not None
            }
            required = ("Day", "Cage", "Animal", "Trial")
            if not all(header in index for header in required):
                return []
            compact, curvature_columns, position_columns = _frame_columns(index)

            output: list[TrialTailCurvatureWithinTrialVariation] = []
            for values in sheet.iter_rows(min_row=2, values_only=True):
                day = _cell_text(values, index["Day"])
                cage = _cell_text(values, index["Cage"])
                animal = _cell_text(values, index["Animal"])
                trial = _cell_int(values, index["Trial"])
                if not (day and cage and animal and trial):
                    continue

                by_bin: dict[int, list[float]] = defaultdict(list)
                for frame in set(compact) | set(curvature_columns) | set(position_columns):
                    curvature, position = _frame_measurement(
                        values,
                        frame,
                        compact,
                        curvature_columns,
                        position_columns,
                    )
                    if curvature is None or position is None:
                        continue
                    bin_start = _bin_start(position)
                    if bin_start is not None:
                        by_bin[bin_start].append(curvature)

                standard_deviations = tuple(
                    statistics.stdev(by_bin[start]) if len(by_bin[start]) >= 2 else None
                    for start in range(0, 90, 10)
                )
                if not any(value is not None for value in standard_deviations):
                    continue
                output.append(
                    TrialTailCurvatureWithinTrialVariation(
                        day=day,
                        cage=cage,
                        animal=animal,
                        condition=conditions.get((cage, animal), "UNASSIGNED"),
                        trial=trial,
                        standard_deviations=standard_deviations,
                        frame_counts=tuple(len(by_bin[start]) for start in range(0, 90, 10)),
                    )
                )
            return output
        finally:
            workbook.close()

    @staticmethod
    def _within_trial_summary_inputs(
        variation_trials: Iterable[TrialTailCurvatureWithinTrialVariation],
    ) -> list[TrialTailCurvatureSeries]:
        """Adapt per-trial, per-bin SDs to the shared equal-mouse summaries."""
        return [
            TrialTailCurvatureSeries(
                day=series.day,
                cage=series.cage,
                animal=series.animal,
                condition=series.condition,
                trial=series.trial,
                bin_means=series.standard_deviations,
            )
            for series in variation_trials
        ]

    @classmethod
    def _within_trial_group_summaries(
        cls,
        variation_trials: Iterable[TrialTailCurvatureWithinTrialVariation],
    ) -> tuple[list[MouseTailCurvatureSeries], list[GroupTailCurvatureSummary]]:
        """Give each mouse equal weight after averaging its trial/bin SDs."""
        return summarize_group_back_front_paw_distances(
            cls._within_trial_summary_inputs(variation_trials)
        )

    @classmethod
    def _day_within_trial_variation_summaries(
        cls,
        variation_trials: Iterable[TrialTailCurvatureWithinTrialVariation],
    ) -> list[DayGroupPawDistanceSummary]:
        """Summarize one position-controlled within-trial SD value per mouse/day."""
        return summarize_day_by_day_back_front_paw_distances(
            cls._within_trial_summary_inputs(variation_trials)
        )

    @staticmethod
    def _daily_group_svg(
        dataset: str,
        day: str,
        trial_series: list[TrialTailCurvatureSeries],
        mouse_series: list[MouseTailCurvatureSeries],
        group_summaries: list[GroupTailCurvatureSummary],
    ) -> str:
        del mouse_series  # The legend summarizes group-level mouse counts.
        trials = sorted(trial_series, key=_trial_sort_key)
        groups = sorted(group_summaries, key=lambda item: _condition_order(item.condition))
        values = _all_bin_values(trials)
        values.extend(_summary_edges(groups))
        lower, upper = _zoomed_nonnegative_axis_bounds(values)
        width, height = 1420, 790
        left, right, top, bottom = 96, 390, 94, 150
        chart_width, chart_height = width - left - right, height - top - bottom
        x = _bin_x(left, chart_width)
        y = _value_y(top, chart_height, lower, upper)

        parts = [
            f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}">',
            '<rect width="100%" height="100%" fill="white"/>',
            f'<text x="{left}" y="32" font-family="Arial" font-size="22" font-weight="bold">Tail-curvature group comparison - {html.escape(dataset)} {html.escape(day)}</text>',
            f'<text x="{left}" y="56" font-family="Arial" font-size="13" fill="#444">Thin lines are trial means within each 10 cm bin. Each mouse is averaged across its trials before it enters the group mean, so every mouse has equal weight.</text>',
            f'<text x="{left}" y="76" font-family="Arial" font-size="13" fill="#444">Thick lines are group means; shaded bands are plus/minus 1 sample SD across mouse means. Gaps indicate no valid observations.</text>',
        ]
        _append_bin_axes(
            parts,
            left,
            top,
            chart_width,
            chart_height,
            x,
            y,
            lower,
            upper,
            "Tail curvature (degrees)",
            height,
        )
        _append_summary_bands(parts, groups, x, y, lower, upper)
        for series in trials:
            color = _condition_color(series.condition)
            tooltip = html.escape(
                f"{series.condition.title()} | Cage {series.cage} Mouse {series.animal} | Trial T{series.trial}"
            )
            _append_hover_lines(
                parts,
                _line_segments(series.bin_means, x, y),
                color,
                tooltip,
                1.7,
                0.28,
            )
        _append_summary_lines(parts, groups, x, y)
        _append_group_legend(parts, groups, left + chart_width + 40, top)
        parts.append("</svg>")
        return "\n".join(parts)

    @staticmethod
    def _daily_within_trial_variation_svg(
        dataset: str,
        day: str,
        trial_variation: Iterable[TrialTailCurvatureWithinTrialVariation],
        mouse_series: Iterable[MouseTailCurvatureSeries],
        group_summaries: list[GroupTailCurvatureSummary],
    ) -> str:
        """Show frame-level curvature variability without pooling trials."""
        del mouse_series  # Group lines/bands already use equal-mouse values.
        trials = sorted(trial_variation, key=_trial_sort_key)
        groups = sorted(group_summaries, key=lambda item: _condition_order(item.condition))
        values = [
            value
            for trial in trials
            for value in trial.standard_deviations
            if _valid(value)
        ]
        values.extend(_summary_edges(groups))
        lower, upper = _zoomed_nonnegative_axis_bounds(values)
        width, height = 1420, 790
        left, right, top, bottom = 96, 390, 94, 150
        chart_width, chart_height = width - left - right, height - top - bottom
        x = _bin_x(left, chart_width)
        y = _value_y(top, chart_height, lower, upper)

        parts = [
            f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}">',
            '<rect width="100%" height="100%" fill="white"/>',
            f'<text x="{left}" y="32" font-family="Arial" font-size="22" font-weight="bold">Tail-curvature within-trial variation - {html.escape(dataset)} {html.escape(day)}</text>',
            f'<text x="{left}" y="56" font-family="Arial" font-size="13" fill="#444">Each thin line is one trial. A point is that trial\'s sample SD across raw valid frame curvatures in the 10 cm bin; higher values mean more posture variation during that run.</text>',
            f'<text x="{left}" y="76" font-family="Arial" font-size="13" fill="#444">For group values, each mouse first contributes the mean of its own trial SDs. Thick lines are group means; shaded bands are plus/minus 1 sample SD across mice.</text>',
        ]
        _append_bin_axes(
            parts,
            left,
            top,
            chart_width,
            chart_height,
            x,
            y,
            lower,
            upper,
            "Within-trial tail-curvature SD (degrees)",
            height,
        )
        _append_summary_bands(parts, groups, x, y, lower, upper)
        for trial in trials:
            color = _condition_color(trial.condition)
            tooltip = html.escape(
                f"{trial.condition.title()} | Cage {trial.cage} Mouse {trial.animal} | "
                f"Trial T{trial.trial} | Raw frame tail-curvature SD"
            )
            _append_hover_lines(
                parts,
                _line_segments(trial.standard_deviations, x, y),
                color,
                tooltip,
                1.7,
                0.28,
            )
        _append_summary_lines(parts, groups, x, y)
        _append_group_legend(parts, groups, left + chart_width + 40, top)
        parts.append("</svg>")
        return "\n".join(parts)

    @staticmethod
    def _day_bar_svg(
        dataset: str,
        summaries: Iterable[DayGroupPawDistanceSummary],
        *,
        title: str,
        calculation_note: str,
        comparison_note: str,
        empty_note: str,
        y_label: str,
        value_suffix: str,
    ) -> str:
        """Render paired SHAM/STROKE day bars with transparent SD ranges.

        Bars are always calculated from the data. The neutral-gray uncertainty
        layer and whisker keep ±1 sample SD readable on top of either group
        color without suggesting that the groups were separated by design.
        """
        groups = sorted(summaries, key=lambda item: _condition_order(item.condition))
        days = sorted({day for group in groups for day in group.means}, key=natural_day_key)
        values: list[float] = []
        for group in groups:
            for day, mean in group.means.items():
                if not _valid(mean):
                    continue
                values.append(float(mean))
                standard_deviation = group.standard_deviations.get(day)
                if _valid(standard_deviation):
                    values.extend((
                        float(mean) - float(standard_deviation),
                        float(mean) + float(standard_deviation),
                    ))
        lower, upper = _bar_nonnegative_axis_bounds(values)
        width, height = 1420, 760
        left, right, top, bottom = 96, 360, 94, 136
        chart_width, chart_height = width - left - right, height - top - bottom
        y = _value_y(top, chart_height, lower, upper)
        baseline_y = y(0.0)
        parts = _svg_header(
            width,
            height,
            left,
            f"{title} - {dataset}",
            calculation_note,
            comparison_note,
        )
        if not days or not groups:
            parts.append(
                f'<text x="{left}" y="{top + 50}" font-family="Arial" font-size="16" fill="#444">{html.escape(empty_note)}</text>'
            )
            parts.append("</svg>")
            return "\n".join(parts)

        for tick in _axis_ticks(lower, upper):
            py = y(tick)
            parts.extend((
                f'<line x1="{left}" y1="{py:.1f}" x2="{left + chart_width}" y2="{py:.1f}" stroke="#e6e6e6"/>',
                f'<text x="{left - 10}" y="{py + 4:.1f}" text-anchor="end" font-family="Arial" font-size="12">{_tick_label(tick)}</text>',
            ))

        day_spacing = chart_width / len(days)
        group_count = len(groups)
        bar_width = min(64.0, day_spacing * 0.28 / max(group_count, 1))
        bar_gap = min(16.0, bar_width * 0.28)
        group_span = group_count * bar_width + max(group_count - 1, 0) * bar_gap

        for day_index, day in enumerate(days):
            day_center = left + day_spacing * (day_index + 0.5)
            parts.extend((
                f'<line x1="{day_center:.1f}" y1="{top}" x2="{day_center:.1f}" y2="{top + chart_height}" stroke="#f0f0f0"/>',
                f'<text x="{day_center:.1f}" y="{top + chart_height + 26}" text-anchor="middle" font-family="Arial" font-size="13">{html.escape(day)}</text>',
            ))
            for group_index, group in enumerate(groups):
                mean = group.means.get(day)
                if not _valid(mean):
                    continue
                mean = float(mean)
                center_offset = -group_span / 2 + bar_width / 2 + group_index * (bar_width + bar_gap)
                bar_center = day_center + center_offset
                bar_left = bar_center - bar_width / 2
                standard_deviation = group.standard_deviations.get(day)
                error_top = mean
                upper_y: float | None = None
                lower_y: float | None = None
                if _valid(standard_deviation):
                    error = float(standard_deviation)
                    error_top = min(upper, mean + error)
                    error_bottom = max(lower, mean - error)
                    upper_y = y(error_top)
                    lower_y = y(error_bottom)
                    parts.append(
                        f'<rect x="{bar_left:.1f}" y="{upper_y:.1f}" width="{bar_width:.1f}" height="{max(0.0, lower_y - upper_y):.1f}" fill="{ERROR_RANGE_COLOR}" fill-opacity="0.24" stroke="none"/>'
                    )

                color = _condition_color(group.condition)
                mean_y = y(mean)
                bar_y = min(mean_y, baseline_y)
                bar_height = abs(baseline_y - mean_y)
                if _valid(standard_deviation):
                    tooltip_text = (
                        f"{group.condition.title()} | {day}: mean {mean:.3f}{value_suffix}, "
                        f"SD {float(standard_deviation):.3f}{value_suffix}; "
                        f"n={group.mouse_counts.get(day, 0)} mice"
                    )
                else:
                    tooltip_text = (
                        f"{group.condition.title()} | {day}: mean {mean:.3f}{value_suffix}, "
                        f"SD unavailable; n={group.mouse_counts.get(day, 0)} mice"
                    )
                parts.append(
                    f'<rect x="{bar_left:.1f}" y="{bar_y:.1f}" width="{bar_width:.1f}" height="{bar_height:.1f}" fill="{color}" cursor="help"><title>{html.escape(tooltip_text)}</title></rect>'
                )
                if upper_y is not None and lower_y is not None:
                    # Draw this *after* the colored bar. In particular, the
                    # lower cap remains visible when it lies inside the bar.
                    parts.extend((
                        f'<line x1="{bar_center:.1f}" y1="{upper_y:.1f}" x2="{bar_center:.1f}" y2="{lower_y:.1f}" stroke="{ERROR_WHISKER_COLOR}" stroke-width="1.75"/>',
                        f'<line x1="{bar_center - 7:.1f}" y1="{upper_y:.1f}" x2="{bar_center + 7:.1f}" y2="{upper_y:.1f}" stroke="{ERROR_WHISKER_COLOR}" stroke-width="1.75"/>',
                        f'<line x1="{bar_center - 7:.1f}" y1="{lower_y:.1f}" x2="{bar_center + 7:.1f}" y2="{lower_y:.1f}" stroke="{ERROR_WHISKER_COLOR}" stroke-width="1.75"/>',
                    ))
                label_y = max(top + 14, y(error_top) - 8)
                parts.append(
                    f'<text x="{bar_center:.1f}" y="{label_y:.1f}" text-anchor="middle" font-family="Arial" font-size="11" font-weight="bold" fill="{ERROR_WHISKER_COLOR}">{mean:.3f}{html.escape(value_suffix)}</text>'
                )

        legend_x = left + chart_width + 42
        parts.append(
            f'<text x="{legend_x}" y="{top + 4}" font-family="Arial" font-size="15" font-weight="bold">Groups</text>'
        )
        for index, group in enumerate(groups):
            color = _condition_color(group.condition)
            count = max(group.mouse_counts.values(), default=0)
            text_y = top + 32 + index * 30
            parts.extend((
                f'<rect x="{legend_x}" y="{text_y - 12}" width="18" height="12" fill="{color}"/>',
                f'<text x="{legend_x + 28}" y="{text_y}" font-family="Arial" font-size="13">{html.escape(group.condition.title())} (up to n={count} mice)</text>',
            ))
        parts.extend((
            f'<text x="{left + chart_width / 2:.1f}" y="{height - 34}" text-anchor="middle" font-family="Arial" font-size="15">Experimental day</text>',
            f'<text x="25" y="{top + chart_height / 2:.1f}" text-anchor="middle" font-family="Arial" font-size="15" transform="rotate(-90 25 {top + chart_height / 2:.1f})">{html.escape(y_label)}</text>',
            "</svg>",
        ))
        return "\n".join(parts)


def _bar_nonnegative_axis_bounds(values: Iterable[float]) -> tuple[float, float]:
    """Use a zero baseline while preserving a small headroom above SD bars."""
    finite_values = [float(value) for value in values if math.isfinite(float(value))]
    maximum = max(finite_values, default=0.0)
    if maximum <= 0.0:
        return 0.0, 1.0
    padded_maximum = maximum * 1.08
    target_step = padded_maximum / 6.0
    magnitude = 10 ** math.floor(math.log10(target_step))
    normalized = target_step / magnitude
    step = next(
        candidate * magnitude
        for candidate in (1.0, 2.0, 5.0, 10.0)
        if normalized <= candidate
    )
    return 0.0, max(step * 2.0, math.ceil(padded_maximum / step) * step)


def _frame_columns(index: dict[str, int]) -> tuple[dict[int, int], dict[int, int], dict[int, int]]:
    compact: dict[int, int] = {}
    curvature_columns: dict[int, int] = {}
    position_columns: dict[int, int] = {}
    for header, column in index.items():
        if not header.startswith("Frame "):
            continue
        try:
            if header.endswith(" curvature (deg)"):
                curvature_columns[
                    int(header.removeprefix("Frame ").removesuffix(" curvature (deg)"))
                ] = column
            elif header.endswith(" back paw (cm)"):
                position_columns[
                    int(header.removeprefix("Frame ").removesuffix(" back paw (cm)"))
                ] = column
            else:
                compact[int(header.removeprefix("Frame "))] = column
        except ValueError:
            continue
    return compact, curvature_columns, position_columns


def _frame_measurement(
    values: tuple[object, ...],
    frame: int,
    compact: dict[int, int],
    curvature_columns: dict[int, int],
    position_columns: dict[int, int],
) -> tuple[float | None, float | None]:
    if frame in compact:
        column = compact[frame]
        if column >= len(values) or values[column] is None:
            return None, None
        match = FRAME_CELL_RE.fullmatch(str(values[column]).strip())
        if match is None:
            return None, None
        return float(match.group(1)), float(match.group(2))
    return (
        _cell_float(values, curvature_columns.get(frame)),
        _cell_float(values, position_columns.get(frame)),
    )


def _line_segments(values, x, y) -> list[list[str]]:
    """Keep metric-local curve code readable while using the shared bin shape."""
    from .back_front_paw_distance_plot import _line_segments as shared_line_segments

    return shared_line_segments(values, x, y)


def _numeric_key(value: str) -> tuple[int, int | str]:
    return (0, int(value)) if value.isdigit() else (1, value.casefold())
