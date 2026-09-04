"""SVG summaries of within-mouse tail-angle consistency across trials."""

from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
import html
import math
from pathlib import Path
import statistics
from typing import Iterable

from .dataset import ROOT


OUTPUT_ROOT = ROOT / "outputs"
RESULTS_FILENAME = "RBT_CV_Results.xlsx"
SHEET_NAME = "Tail Angle Trial Consistency"
FILE_PREFIX = "tail_angle_trial_consistency_"
FILE_SUFFIX = ".svg"
CHART_DIRECTORY = "Tail Angle"
BIN_STARTS_CM = tuple(range(0, 90, 10))

SHAM_COLOR = "#1976D2"
STROKE_COLOR = "#D32F2F"
UNASSIGNED_COLOR = "#616161"


@dataclass(frozen=True)
class ConsistencySeries:
    """One mouse's per-position sample SD values for one experimental day."""

    day: str
    cage: str
    animal: str
    condition: str
    standard_deviations: tuple[float | None, ...]
    trial_counts: tuple[int, ...]
    overall_standard_deviation: float | None
    overall_trial_count: int

    @property
    def label(self) -> str:
        return f"Cage {self.cage} Mouse {self.animal}"


class TailAngleConsistencyPlotStore:
    """Render one readable within-mouse trial-consistency chart for each day."""

    def __init__(self, output_root: Path = OUTPUT_ROOT) -> None:
        self.output_root = output_root

    def result_dir(self, dataset: str) -> Path:
        return self.output_root / f"{dataset} Results"

    def chart_dir(self, dataset: str) -> Path:
        return self.result_dir(dataset) / CHART_DIRECTORY

    def refresh_dataset(self, dataset: str) -> dict[str, Path]:
        """Refresh charts from the numeric Excel consistency sheet.

        The workbook is the single data source so the plotted SD values exactly
        match the values the user can inspect in Excel.
        """
        workbook_path = self.result_dir(dataset) / RESULTS_FILENAME
        if not workbook_path.exists():
            return {}

        series = self._read_series(workbook_path)
        by_day: dict[str, list[ConsistencySeries]] = defaultdict(list)
        for item in series:
            by_day[item.day].append(item)

        result_dir = self.result_dir(dataset)
        chart_dir = self.chart_dir(dataset)
        chart_dir.mkdir(parents=True, exist_ok=True)
        paths: dict[str, Path] = {}
        for day, day_series in by_day.items():
            path = chart_dir / f"{FILE_PREFIX}{day}{FILE_SUFFIX}"
            self._write_svg(path, self._svg(dataset, day, day_series))
            paths[day] = path

        # A later reanalysis may invalidate every frame for one day. In that
        # case remove its obsolete chart rather than leaving stale results.
        live_names = {path.name for path in paths.values()}
        for stale_path in chart_dir.glob(f"{FILE_PREFIX}*{FILE_SUFFIX}"):
            if stale_path.name not in live_names:
                stale_path.unlink()
        # Clean up only generated charts from the older root-level layout.
        for stale_path in result_dir.glob(f"{FILE_PREFIX}*{FILE_SUFFIX}"):
            stale_path.unlink()
        return paths

    @staticmethod
    def _read_series(workbook_path: Path) -> list[ConsistencySeries]:
        from openpyxl import load_workbook

        workbook = load_workbook(workbook_path, data_only=True, read_only=True)
        try:
            if SHEET_NAME not in workbook.sheetnames:
                return []
            sheet = workbook[SHEET_NAME]
            headers = [cell.value for cell in next(sheet.iter_rows(max_row=1), ())]
            index = {str(header): position for position, header in enumerate(headers) if header is not None}
            required = ("Day", "Cage", "Animal", "Condition")
            if not all(header in index for header in required):
                return []

            series: list[ConsistencySeries] = []
            for values in sheet.iter_rows(min_row=2, values_only=True):
                day = _cell_text(values, index["Day"])
                cage = _cell_text(values, index["Cage"])
                animal = _cell_text(values, index["Animal"])
                if not (day and cage and animal):
                    continue
                standard_deviations = tuple(
                    _cell_float(values, index.get(f"{start}-{start + 10} cm SD (deg)"))
                    for start in BIN_STARTS_CM
                )
                trial_counts = tuple(
                    _cell_int(values, index.get(f"{start}-{start + 10} cm trial n"))
                    for start in BIN_STARTS_CM
                )
                series.append(
                    ConsistencySeries(
                        day=day,
                        cage=cage,
                        animal=animal,
                        condition=_cell_text(values, index["Condition"]).upper() or "UNASSIGNED",
                        standard_deviations=standard_deviations,
                        trial_counts=trial_counts,
                        overall_standard_deviation=_cell_float(
                            values, index.get("Overall 0-90 cm SD (deg)")
                        ),
                        overall_trial_count=_cell_int(
                            values, index.get("Overall 0-90 cm trial n")
                        ),
                    )
                )
            return series
        finally:
            workbook.close()

    @staticmethod
    def _write_svg(path: Path, text: str) -> None:
        temporary_path = path.with_suffix(f"{path.suffix}.tmp")
        try:
            temporary_path.write_text(text, encoding="utf-8")
            temporary_path.replace(path)
        finally:
            if temporary_path.exists():
                temporary_path.unlink()

    @staticmethod
    def _svg(dataset: str, day: str, series: Iterable[ConsistencySeries]) -> str:
        rows = sorted(
            series,
            key=lambda item: (_condition_order(item.condition), _numeric_key(item.cage), _numeric_key(item.animal)),
        )
        width = 1380
        height = max(720, 250 + 22 * len(rows))
        left, right, top, bottom = 96, 360, 90, 94
        chart_width = width - left - right
        chart_height = height - top - bottom
        values = [
            value
            for row in rows
            for value in row.standard_deviations
            if value is not None and math.isfinite(value)
        ]
        group_values: dict[str, dict[int, list[float]]] = defaultdict(lambda: defaultdict(list))
        for row in rows:
            for start, value in zip(BIN_STARTS_CM, row.standard_deviations):
                if value is not None and math.isfinite(value):
                    group_values[row.condition][start].append(value)

        # The band is the sample SD of the individual mouse SDs in each group.
        # Include its upper edge while choosing the axis range so it never clips.
        group_statistics: dict[str, dict[int, tuple[float, float | None]]] = {}
        band_edges = list(values)
        for condition, per_bin in group_values.items():
            group_statistics[condition] = {}
            for start, values_at_bin in per_bin.items():
                mean = statistics.fmean(values_at_bin)
                error = statistics.stdev(values_at_bin) if len(values_at_bin) >= 2 else None
                group_statistics[condition][start] = (mean, error)
                band_edges.append(mean + error if error is not None else mean)
        upper = max(35.0, math.ceil(max(band_edges, default=0.0) / 5.0) * 5.0)

        def x(position_cm: float) -> float:
            return left + chart_width * position_cm / 90

        def y(value: float) -> float:
            return top + chart_height * (upper - value) / upper

        parts = [
            f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}">',
            '<rect width="100%" height="100%" fill="white"/>',
            f'<text x="{left}" y="32" font-family="Arial" font-size="22" font-weight="bold">Tail-angle trial consistency - {html.escape(dataset)} {html.escape(day)}</text>',
            f'<text x="{left}" y="56" font-family="Arial" font-size="13" fill="#444">Thin lines show each mouse\'s trial-to-trial tail-angle SD in each 10 cm bin. Thick lines show the group average; shaded bands show variation across mice.</text>',
        ]

        for tick in range(0, 91, 10):
            px = x(tick)
            parts.extend(
                (
                    f'<line x1="{px:.1f}" y1="{top}" x2="{px:.1f}" y2="{top + chart_height}" stroke="#e6e6e6"/>',
                    f'<text x="{px:.1f}" y="{top + chart_height + 25}" text-anchor="middle" font-family="Arial" font-size="12">{tick}</text>',
                )
            )
        for value in range(0, int(upper) + 1, 5):
            py = y(value)
            parts.extend(
                (
                    f'<line x1="{left}" y1="{py:.1f}" x2="{left + chart_width}" y2="{py:.1f}" stroke="#e6e6e6"/>',
                    f'<text x="{left - 10}" y="{py + 4:.1f}" text-anchor="end" font-family="Arial" font-size="12">{value}</text>',
                )
            )

        if not values:
            parts.append(
                f'<text x="{left + chart_width / 2:.1f}" y="{top + chart_height / 2:.1f}" text-anchor="middle" font-family="Arial" font-size="17" fill="#555">No SD values yet: each mouse needs at least two valid trials in a position bin.</text>'
            )

        for row in rows:
            color = _condition_color(row.condition)
            tooltip = html.escape(
                f"{row.condition.title()} | Cage {row.cage} Mouse {row.animal} | Trial-to-trial consistency SD"
            )
            segments: list[list[str]] = []
            segment: list[str] = []
            for start, value in zip(BIN_STARTS_CM, row.standard_deviations):
                if value is None or not math.isfinite(value):
                    if segment:
                        segments.append(segment)
                        segment = []
                    continue
                segment.append(f"{x(start + 5):.1f},{y(value):.1f}")
            if segment:
                segments.append(segment)
            for points in segments:
                parts.append(
                    f'<polyline points="{" ".join(points)}" fill="none" stroke="{color}" stroke-width="2" stroke-opacity="0.38" pointer-events="none"/>'
                    f'<polyline points="{" ".join(points)}" fill="none" stroke="{color}" stroke-width="8" stroke-opacity="0" pointer-events="stroke" cursor="help"><title>{tooltip}</title></polyline>'
                )

        for condition, per_bin in sorted(group_statistics.items(), key=lambda item: _condition_order(item[0])):
            color = _condition_color(condition)
            band_segments: list[list[tuple[float, float, float]]] = []
            band_segment: list[tuple[float, float, float]] = []
            for start in BIN_STARTS_CM:
                statistics_at_bin = per_bin.get(start)
                if statistics_at_bin is None or statistics_at_bin[1] is None:
                    if band_segment:
                        band_segments.append(band_segment)
                        band_segment = []
                    continue
                mean, error = statistics_at_bin
                band_segment.append((x(start + 5), mean, error))
            if band_segment:
                band_segments.append(band_segment)
            for segment in band_segments:
                high_points = [
                    f"{px:.1f},{y(min(upper, mean + error)):.1f}"
                    for px, mean, error in segment
                ]
                low_points = [
                    f"{px:.1f},{y(max(0.0, mean - error)):.1f}"
                    for px, mean, error in reversed(segment)
                ]
                parts.append(
                    f'<polygon points="{" ".join((*high_points, *low_points))}" fill="{color}" fill-opacity="0.18" stroke="none"/>'
                )

        for condition, per_bin in sorted(group_statistics.items(), key=lambda item: _condition_order(item[0])):
            color = _condition_color(condition)
            segments: list[list[str]] = []
            segment = []
            for start in BIN_STARTS_CM:
                statistics_at_bin = per_bin.get(start)
                if statistics_at_bin is None:
                    if segment:
                        segments.append(segment)
                        segment = []
                    continue
                mean, _error = statistics_at_bin
                segment.append(f"{x(start + 5):.1f},{y(mean):.1f}")
            if segment:
                segments.append(segment)
            for points in segments:
                parts.append(
                    f'<polyline points="{" ".join(points)}" fill="none" stroke="{color}" stroke-width="4" stroke-linecap="round"/>'
                )

        legend_x = left + chart_width + 42
        parts.extend(
            (
                f'<text x="{legend_x}" y="{top + 4}" font-family="Arial" font-size="15" font-weight="bold">Mouse summary</text>',
                f'<line x1="{legend_x}" y1="{top + 23}" x2="{legend_x + 28}" y2="{top + 23}" stroke="{SHAM_COLOR}" stroke-width="4"/>',
                f'<text x="{legend_x + 36}" y="{top + 28}" font-family="Arial" font-size="13">SHAM</text>',
                f'<line x1="{legend_x + 95}" y1="{top + 23}" x2="{legend_x + 123}" y2="{top + 23}" stroke="{STROKE_COLOR}" stroke-width="4"/>',
                f'<text x="{legend_x + 131}" y="{top + 28}" font-family="Arial" font-size="13">STROKE</text>',
            )
        )
        for index, row in enumerate(rows):
            overall = "—" if row.overall_standard_deviation is None else f"{row.overall_standard_deviation:.3f}°"
            text_y = top + 57 + index * 22
            parts.extend(
                (
                    f'<line x1="{legend_x}" y1="{text_y - 5}" x2="{legend_x + 18}" y2="{text_y - 5}" stroke="{_condition_color(row.condition)}" stroke-width="3" stroke-opacity="0.7"/>',
                    f'<text x="{legend_x + 26}" y="{text_y}" font-family="Arial" font-size="12">{html.escape(row.label)}: overall {overall} (n={row.overall_trial_count})</text>',
                )
            )

        parts.extend(
            (
                f'<text x="{left + chart_width / 2:.1f}" y="{height - 32}" text-anchor="middle" font-family="Arial" font-size="15">Back-paw position on calibrated beam (cm)</text>',
                f'<text x="25" y="{top + chart_height / 2:.1f}" text-anchor="middle" font-family="Arial" font-size="15" transform="rotate(-90 25 {top + chart_height / 2:.1f})">Tail-angle sample SD across trials (degrees)</text>',
                "</svg>",
            )
        )
        return "\n".join(parts)


def _cell_text(values: tuple[object, ...], index: int | None) -> str:
    if index is None or index >= len(values) or values[index] is None:
        return ""
    value = values[index]
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _cell_float(values: tuple[object, ...], index: int | None) -> float | None:
    if index is None or index >= len(values) or values[index] is None:
        return None
    try:
        value = float(values[index])
    except (TypeError, ValueError):
        return None
    return value if math.isfinite(value) else None


def _cell_int(values: tuple[object, ...], index: int | None) -> int:
    value = _cell_float(values, index)
    return int(value) if value is not None else 0


def _condition_color(condition: str) -> str:
    return {"SHAM": SHAM_COLOR, "STROKE": STROKE_COLOR}.get(condition.upper(), UNASSIGNED_COLOR)


def _condition_order(condition: str) -> int:
    return {"SHAM": 0, "STROKE": 1}.get(condition.upper(), 2)


def _numeric_key(value: str) -> tuple[int, int | str]:
    return (0, int(value)) if value.isdigit() else (1, value.casefold())
