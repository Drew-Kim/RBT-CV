from __future__ import annotations

from collections import defaultdict
from collections.abc import Iterable
from pathlib import Path
import re
import statistics
from typing import TYPE_CHECKING

from .condition_map import ConditionMapStore
from .dataset import ROOT, natural_day_key
from .tail_angle_consistency_plot import TailAngleConsistencyPlotStore
from .tail_angle_group_plot import TailAngleGroupPlotStore

if TYPE_CHECKING:
    from .annotations import TrialAnnotation
    from .dataset import TrialVideo
    from .research_angle import TailAngleFrameRecord


RESULTS_ROOT = ROOT / "outputs"
RESULTS_FILENAME = "RBT_CV_Results.xlsx"
TIME_SHEET = "Forelimb"
AUDIT_SHEET = "RBT-CV Results"
DISTANCE_TABLE_COLUMN = 26
SPEED_TABLE_ROW = 30
ANGLE_SHEET = "Frame Angles"
ANGLE_CONSISTENCY_SHEET = "Tail Angle Trial Consistency"
ANGLE_METADATA_HEADERS = (
    "Source video",
    "Day",
    "Cage",
    "Animal",
    "Group",
    "Trial",
)
LEGACY_ANGLE_HEADERS = (
    "Dataset",
    "Day",
    "Cage",
    "Animal",
    "Group",
    "Trial",
    "Frame",
    "Time (s)",
    "Back paw position (cm)",
    "Signed tail angle (deg)",
    "Source video",
)
ANGLE_BIN_STARTS_CM = tuple(range(0, 90, 10))
ANGLE_CONSISTENCY_HEADERS = (
    "Day",
    "Cage",
    "Animal",
    "Condition",
    *(
        header
        for start in ANGLE_BIN_STARTS_CM
        for header in (f"{start}-{start + 10} cm SD (deg)", f"{start}-{start + 10} cm trial n")
    ),
    "Overall 0-90 cm SD (deg)",
    "Overall 0-90 cm trial n",
)


class ResultsWorkbookError(RuntimeError):
    """Raised when the results workbook cannot be read or safely written."""


class ResultsWorkbook:
    """Maintain ordered Time, Distance, Speed, and audit tables for each dataset."""

    def __init__(self, results_root: Path = RESULTS_ROOT) -> None:
        self.results_root = results_root

    def path_for_dataset(self, dataset: str) -> Path:
        return self.results_root / f"{dataset} Results" / RESULTS_FILENAME

    def save(self, annotation: TrialAnnotation) -> Path:
        return self.save_many((annotation,))[annotation.dataset]

    def save_many(self, annotations: Iterable[TrialAnnotation]) -> dict[str, Path]:
        """Upsert a result batch with one workbook write per dataset.

        This is used by animal/day analysis so recalculation of the Speed table
        and the Excel save happen once instead of once for every video.
        """
        by_dataset: dict[str, list[TrialAnnotation]] = defaultdict(list)
        for annotation in annotations:
            by_dataset[annotation.dataset].append(annotation)

        saved_paths: dict[str, Path] = {}
        for dataset, dataset_annotations in by_dataset.items():
            path = self.path_for_dataset(dataset)
            workbook = self._load_or_create(path)
            sheet = self._time_sheet(workbook)

            self._clear_speed_tables(sheet)
            self._ensure_result_tables(sheet)
            for annotation in dataset_annotations:
                self._write_measurement(sheet, 1, 1, annotation, annotation.crossing_time)
                self._write_measurement(
                    sheet,
                    1,
                    DISTANCE_TABLE_COLUMN,
                    annotation,
                    annotation.distance_cm,
                )
                self._upsert_audit_row(self._audit_sheet(workbook), annotation)
            self._order_forelimb_day_columns(sheet)
            self._rebuild_speed_table(sheet)

            self._save_workbook(
                workbook,
                path,
                "Close it in Excel and run the analysis again.",
            )
            saved_paths[dataset] = path
        return saved_paths

    def refresh_forelimb_layout(self, dataset: str) -> Path | None:
        """Put existing Forelimb day columns in chronological left-to-right order."""
        path = self.path_for_dataset(dataset)
        if not path.exists():
            return None
        workbook = self._load_or_create(path)
        sheet = self._time_sheet(workbook)
        self._clear_speed_tables(sheet)
        self._ensure_result_tables(sheet)
        self._order_forelimb_day_columns(sheet)
        self._rebuild_speed_table(sheet)
        self._save_workbook(
            workbook,
            path,
            "Close it in Excel and try refreshing the Forelimb layout again.",
        )
        return path

    def save_tail_angle_measurements(
        self,
        videos: Iterable[TrialVideo],
        records: Iterable[TailAngleFrameRecord],
        *,
        refresh_plot: bool = True,
    ) -> dict[str, Path]:
        """Upsert valid frame-angle rows for completed DLC analysis batches.

        Each dataset uses its normal results workbook and a single filterable
        ``Frame Angles`` sheet. A reanalysis first removes rows from the same
        source video, including stale rows when no valid angle remains.
        """
        videos_by_dataset: dict[str, list[TrialVideo]] = defaultdict(list)
        for video in videos:
            videos_by_dataset[video.dataset].append(video)

        records_by_dataset: dict[str, list[TailAngleFrameRecord]] = defaultdict(list)
        for record in records:
            records_by_dataset[record.dataset].append(record)

        saved_paths: dict[str, Path] = {}
        for dataset, dataset_videos in videos_by_dataset.items():
            path = self.path_for_dataset(dataset)
            workbook = self._load_or_create(path)
            sheet = self._angle_sheet(workbook)
            conditions = ConditionMapStore(self.results_root).load(dataset)
            self._upsert_tail_angle_rows(
                sheet,
                records_by_dataset[dataset],
                dataset_videos,
                conditions,
            )
            self._write_tail_angle_consistency_sheet(workbook, sheet, dataset)
            self._save_workbook(
                workbook,
                path,
                "Close it in Excel and run the analysis again.",
            )
            if refresh_plot:
                TailAngleConsistencyPlotStore(self.results_root).refresh_dataset(dataset)
                TailAngleGroupPlotStore(self.results_root).refresh_dataset(dataset)
            saved_paths[dataset] = path
        return saved_paths

    def refresh_tail_angle_consistency(self, dataset: str) -> Path | None:
        """Rebuild the consistency sheet and chart from already exported angles.

        This makes a newly assigned SHAM/STROKE label visible immediately and
        also supports creating the first chart from an existing Frame Angles
        worksheet without re-running DeepLabCut.
        """
        path = self.path_for_dataset(dataset)
        if not path.exists():
            return None
        workbook = self._load_or_create(path)
        if ANGLE_SHEET not in workbook.sheetnames:
            return None
        angle_sheet = workbook[ANGLE_SHEET]
        self._upsert_tail_angle_rows(
            angle_sheet,
            [],
            [],
            ConditionMapStore(self.results_root).load(dataset),
        )
        self._write_tail_angle_consistency_sheet(workbook, angle_sheet, dataset)
        self._save_workbook(
            workbook,
            path,
            "Close it in Excel and try refreshing the tail-angle results again.",
        )
        TailAngleConsistencyPlotStore(self.results_root).refresh_dataset(dataset)
        TailAngleGroupPlotStore(self.results_root).refresh_dataset(dataset)
        return path

    def _load_or_create(self, path: Path):
        try:
            from openpyxl import Workbook, load_workbook
            from openpyxl.utils.exceptions import InvalidFileException
            from zipfile import BadZipFile
        except ImportError as exc:
            raise ResultsWorkbookError(
                "Excel support is missing. Install the project requirements."
            ) from exc

        if path.exists():
            try:
                return load_workbook(path)
            except (OSError, InvalidFileException, BadZipFile) as exc:
                raise ResultsWorkbookError(
                    f"Could not open {path.name}. Close it in Excel and try Save Trial Result again."
                ) from exc

        workbook = Workbook()
        workbook.active.title = TIME_SHEET
        self._ensure_result_tables(workbook.active)
        return workbook

    @staticmethod
    def _save_workbook(workbook, path: Path, instruction: str) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        temporary_path = path.with_name(f".{path.stem}.tmp{path.suffix}")
        try:
            workbook.save(temporary_path)
            temporary_path.replace(path)
        except OSError as exc:
            raise ResultsWorkbookError(
                f"Could not save {path.name}. {instruction}"
            ) from exc
        finally:
            if temporary_path.exists():
                temporary_path.unlink()

    @staticmethod
    def _time_sheet(workbook):
        if TIME_SHEET in workbook.sheetnames:
            return workbook[TIME_SHEET]
        sheet = workbook.create_sheet(TIME_SHEET, 0)
        ResultsWorkbook._ensure_result_tables(sheet)
        return sheet

    @staticmethod
    def _angle_sheet(workbook):
        if ANGLE_SHEET in workbook.sheetnames:
            return workbook[ANGLE_SHEET]
        return workbook.create_sheet(ANGLE_SHEET)

    def _upsert_tail_angle_rows(
        self,
        sheet,
        records: list[TailAngleFrameRecord],
        videos: list[TrialVideo],
        conditions: dict[tuple[str, str], str],
    ) -> None:
        """Rewrite the owned angle sheet as one wide row per source trial."""
        trials = ResultsWorkbook._read_tail_angle_trials(sheet)
        for video in videos:
            trials[video.relative_path] = {
                "source_video": video.relative_path,
                "day": video.day,
                "cage": video.cage_number,
                "animal": video.rat_id,
                "group": conditions.get((video.cage_number, video.rat_id), "Unassigned"),
                "trial": video.trial,
                "angles": {},
                "back_paw_positions": {},
            }
        for record in records:
            trial = trials.setdefault(
                record.relative_video,
                {
                    "source_video": record.relative_video,
                    "day": record.day,
                    "cage": record.cage,
                    "animal": record.animal,
                    "group": conditions.get((record.cage, record.animal), "Unassigned"),
                    "trial": record.trial,
                    "angles": {},
                    "back_paw_positions": {},
                },
            )
            trial["angles"][record.frame] = record.signed_tail_angle_degrees
            trial["back_paw_positions"][record.frame] = record.back_paw_position_cm

        # The label map is the authority for this column. Reapply it to every
        # retained row as well, so changing a SHAM/STROKE assignment updates
        # old trials when the workbook is refreshed.
        for trial in trials.values():
            cage = self._normalized(trial["cage"])
            animal = self._normalized(trial["animal"])
            trial["group"] = conditions.get((cage, animal), "Unassigned")

        frame_count = max(
            (
                max(
                    max(trial["angles"], default=-1),
                    max(trial["back_paw_positions"], default=-1),
                )
                for trial in trials.values()
            ),
            default=-1,
        ) + 1
        headers = [*ANGLE_METADATA_HEADERS, *(f"Frame {frame}" for frame in range(frame_count))]
        if sheet.max_row:
            sheet.delete_rows(1, sheet.max_row)
        sheet.append(headers)

        for trial in sorted(
            trials.values(),
            key=lambda item: (
                item["day"],
                ResultsWorkbook._animal_sort_key(str(item["cage"]), str(item["animal"])),
                int(item["trial"]),
                str(item["source_video"]),
            ),
        ):
            sheet.append(
                (
                    trial["source_video"],
                    trial["day"],
                    trial["cage"],
                    trial["animal"],
                    trial["group"],
                    trial["trial"],
                    *(
                        ResultsWorkbook._frame_measurement_cell(
                            trial["angles"].get(frame),
                            trial["back_paw_positions"].get(frame),
                        )
                        for frame in range(frame_count)
                    ),
                )
            )

        from openpyxl.utils import get_column_letter

        # Keep the source-video identity and headers visible while comparing
        # far-right frame columns across trials.
        sheet.freeze_panes = "B2"
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(sheet.max_row, 1)}"
        sheet.column_dimensions["A"].width = 60
        for column in range(2, len(ANGLE_METADATA_HEADERS) + 1):
            sheet.column_dimensions[get_column_letter(column)].width = 12
        for column in range(len(ANGLE_METADATA_HEADERS) + 1, len(headers) + 1):
            sheet.column_dimensions[get_column_letter(column)].width = 25

    @staticmethod
    def _three_decimal(value: object) -> float | None:
        if value is None:
            return None
        return round(float(value), 3)

    @staticmethod
    def _frame_measurement_cell(angle: object, back_paw_position: object) -> str | None:
        rounded_angle = ResultsWorkbook._three_decimal(angle)
        rounded_back_paw = ResultsWorkbook._three_decimal(back_paw_position)
        if rounded_angle is None and rounded_back_paw is None:
            return None
        if rounded_angle is None:
            return f"angle unavailable; {rounded_back_paw:.3f} cm"
        if rounded_back_paw is None:
            return f"{rounded_angle:+.3f} deg; back paw unavailable"
        return f"{rounded_angle:+.3f} deg; {rounded_back_paw:.3f} cm"

    @staticmethod
    def _read_tail_angle_trials(sheet) -> dict[str, dict[str, object]]:
        """Read either the current wide sheet or the prior long sheet format."""
        headers = [sheet.cell(row=1, column=column).value for column in range(1, sheet.max_column + 1)]
        header_names = [str(header or "") for header in headers]
        if not any(header_names):
            return {}
        header_index = {header: index for index, header in enumerate(header_names)}

        if all(header in header_index for header in LEGACY_ANGLE_HEADERS):
            return ResultsWorkbook._read_legacy_tail_angle_trials(sheet, header_index)
        if all(header in header_index for header in ANGLE_METADATA_HEADERS):
            return ResultsWorkbook._read_wide_tail_angle_trials(sheet, header_index)
        raise ResultsWorkbookError(
            f"The {ANGLE_SHEET!r} sheet has unexpected columns. Rename it, then run analysis again."
        )

    @staticmethod
    def _read_wide_tail_angle_trials(sheet, header_index: dict[str, int]) -> dict[str, dict[str, object]]:
        compact_columns: dict[int, int] = {}
        angle_columns: dict[int, int] = {}
        back_paw_columns: dict[int, int] = {}
        for header, column in header_index.items():
            if not header.startswith("Frame "):
                continue
            try:
                if header.endswith(" angle (deg)"):
                    angle_columns[int(header.removeprefix("Frame ").removesuffix(" angle (deg)"))] = column
                elif header.endswith(" back paw (cm)"):
                    back_paw_columns[int(header.removeprefix("Frame ").removesuffix(" back paw (cm)"))] = column
                else:
                    compact_columns[int(header.removeprefix("Frame "))] = column
            except ValueError:
                continue

        trials: dict[str, dict[str, object]] = {}
        for values in sheet.iter_rows(min_row=2, values_only=True):
            source_video = str(values[header_index["Source video"]] or "")
            if not source_video:
                continue
            angles = {
                frame: values[column]
                for frame, column in angle_columns.items()
                if column < len(values) and values[column] is not None
            }
            back_paw_positions = {
                frame: values[column]
                for frame, column in back_paw_columns.items()
                if column < len(values) and values[column] is not None
            }
            for frame, column in compact_columns.items():
                if column >= len(values):
                    continue
                angle, back_paw_position = ResultsWorkbook._parse_frame_measurement_cell(values[column])
                if angle is not None:
                    angles[frame] = angle
                if back_paw_position is not None:
                    back_paw_positions[frame] = back_paw_position
            trials[source_video] = {
                "source_video": source_video,
                "day": values[header_index["Day"]],
                "cage": values[header_index["Cage"]],
                "animal": values[header_index["Animal"]],
                "group": values[header_index["Group"]],
                "trial": values[header_index["Trial"]],
                "angles": angles,
                "back_paw_positions": back_paw_positions,
            }
        return trials

    @staticmethod
    def _parse_frame_measurement_cell(value: object) -> tuple[float | None, float | None]:
        if value is None:
            return None, None
        if isinstance(value, (float, int)):
            # Accept an early, angle-only wide export during conversion.
            return float(value), None
        match = re.fullmatch(
            r"([+-]?\d+(?:\.\d+)?) deg; ([+-]?\d+(?:\.\d+)?) cm",
            str(value).strip(),
        )
        if match is None:
            return None, None
        return float(match.group(1)), float(match.group(2))

    def _write_tail_angle_consistency_sheet(self, workbook, angle_sheet, dataset: str) -> None:
        """Summarize trial-to-trial tail-angle consistency for every mouse/day."""
        trials = self._read_tail_angle_trials(angle_sheet)
        condition_map = ConditionMapStore(self.results_root).load(dataset)
        trial_means: dict[tuple[str, str, str], dict[int, list[float]]] = defaultdict(
            lambda: defaultdict(list)
        )
        overall_trial_means: dict[tuple[str, str, str], list[float]] = defaultdict(list)

        for trial in trials.values():
            day = str(trial["day"])
            cage = str(trial["cage"])
            animal = str(trial["animal"])
            key = (day, cage, animal)
            by_bin: dict[int, list[float]] = defaultdict(list)
            all_angles: list[float] = []
            for frame, angle in trial["angles"].items():
                back_paw_position = trial["back_paw_positions"].get(frame)
                if back_paw_position is None:
                    continue
                bin_start = self._angle_bin_start(float(back_paw_position))
                if bin_start is None:
                    continue
                numeric_angle = float(angle)
                by_bin[bin_start].append(numeric_angle)
                all_angles.append(numeric_angle)

            for bin_start, angles in by_bin.items():
                trial_means[key][bin_start].append(statistics.fmean(angles))
            if all_angles:
                overall_trial_means[key].append(statistics.fmean(all_angles))

        if ANGLE_CONSISTENCY_SHEET in workbook.sheetnames:
            sheet = workbook[ANGLE_CONSISTENCY_SHEET]
            if sheet.max_row:
                sheet.delete_rows(1, sheet.max_row)
        else:
            sheet = workbook.create_sheet(ANGLE_CONSISTENCY_SHEET)
        sheet.append(ANGLE_CONSISTENCY_HEADERS)

        subject_keys = set(trial_means) | set(overall_trial_means)
        for day, cage, animal in sorted(
            subject_keys,
            key=lambda item: (item[0], self._animal_sort_key(item[1], item[2])),
        ):
            values: list[object] = [
                day,
                cage,
                animal,
                condition_map.get((cage, animal), "Unassigned"),
            ]
            for bin_start in ANGLE_BIN_STARTS_CM:
                means = trial_means[(day, cage, animal)][bin_start]
                values.extend((self._sample_standard_deviation(means), len(means)))
            overall_means = overall_trial_means[(day, cage, animal)]
            values.extend((self._sample_standard_deviation(overall_means), len(overall_means)))
            sheet.append(values)

        from openpyxl.utils import get_column_letter

        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = (
            f"A1:{get_column_letter(len(ANGLE_CONSISTENCY_HEADERS))}{max(sheet.max_row, 1)}"
        )
        for column in range(5, len(ANGLE_CONSISTENCY_HEADERS) + 1, 2):
            for row in range(2, sheet.max_row + 1):
                sheet.cell(row=row, column=column).number_format = "0.000"

    @staticmethod
    def _angle_bin_start(back_paw_position_cm: float) -> int | None:
        if not 0.0 <= back_paw_position_cm <= 90.0:
            return None
        return min(int(back_paw_position_cm // 10) * 10, 80)

    @staticmethod
    def _sample_standard_deviation(values: list[float]) -> float | None:
        return round(statistics.stdev(values), 3) if len(values) >= 2 else None

    @staticmethod
    def _read_legacy_tail_angle_trials(sheet, header_index: dict[str, int]) -> dict[str, dict[str, object]]:
        trials: dict[str, dict[str, object]] = {}
        for values in sheet.iter_rows(min_row=2, values_only=True):
            source_video = str(values[header_index["Source video"]] or "")
            frame = values[header_index["Frame"]]
            angle = values[header_index["Signed tail angle (deg)"]]
            if not source_video or frame is None or angle is None:
                continue
            trial = trials.setdefault(
                source_video,
                {
                    "source_video": source_video,
                    "day": values[header_index["Day"]],
                    "cage": values[header_index["Cage"]],
                    "animal": values[header_index["Animal"]],
                    "group": values[header_index["Group"]],
                    "trial": values[header_index["Trial"]],
                    "angles": {},
                    "back_paw_positions": {},
                },
            )
            trial["angles"][int(frame)] = angle
            back_paw_position = values[header_index["Back paw position (cm)"]]
            if back_paw_position is not None:
                trial["back_paw_positions"][int(frame)] = back_paw_position
        return trials

    @staticmethod
    def _ensure_result_tables(sheet) -> None:
        ResultsWorkbook._ensure_table_headers(sheet, "TIME (seconds)", 1, 1)
        ResultsWorkbook._ensure_table_headers(sheet, "DISTANCE (cm)", 1, DISTANCE_TABLE_COLUMN)

    @staticmethod
    def _ensure_table_headers(
        sheet,
        title: str,
        title_row: int,
        start_column: int,
    ) -> None:
        if sheet.cell(row=title_row, column=start_column).value is None:
            sheet.cell(row=title_row, column=start_column, value=title)

        header_row = title_row + 2
        if sheet.cell(row=header_row, column=start_column).value is None:
            sheet.cell(row=header_row, column=start_column, value="S.No.")
        if sheet.cell(row=header_row, column=start_column + 1).value is None:
            sheet.cell(row=header_row, column=start_column + 1, value="Animal I.D.")

    @staticmethod
    def _normalized(value: object) -> str:
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        text = str(value).strip()
        if text.upper().startswith("C") and text[1:].isdigit():
            return str(int(text[1:]))
        return text

    def _animal_row(
        self,
        sheet,
        cage: str,
        subject: str,
        title_row: int,
        start_column: int,
    ) -> int:
        """Find or create an ordered animal row in one result table."""
        first_row = title_row + 3
        last_column = self._last_table_column(sheet, title_row, start_column)
        records = self._table_records(sheet, first_row, last_column, start_column)

        target = (self._normalized(cage), self._normalized(subject))
        if not any(self._record_key(record) == target for record in records):
            records.append(
                [
                    int(target[0]) if target[0].isdigit() else target[0],
                    int(target[1]) if target[1].isdigit() else target[1],
                    *([None] * (last_column - start_column - 1)),
                ]
            )

        records.sort(key=lambda record: self._animal_sort_key(record[0], record[1]))
        for row, record in enumerate(records, start=first_row):
            for offset, value in enumerate(record):
                sheet.cell(row=row, column=start_column + offset, value=value)

        return next(
            row
            for row, record in enumerate(records, start=first_row)
            if self._record_key(record) == target
        )

    @staticmethod
    def _table_records(sheet, first_row: int, last_column: int, start_column: int) -> list[list[object]]:
        records = []
        row = first_row
        while (
            sheet.cell(row=row, column=start_column).value is not None
            or sheet.cell(row=row, column=start_column + 1).value is not None
        ):
            records.append(
                [
                    sheet.cell(row=row, column=column).value
                    for column in range(start_column, last_column + 1)
                ]
            )
            row += 1
        return records

    @staticmethod
    def _record_key(record: list[object]) -> tuple[str, str]:
        return (
            ResultsWorkbook._normalized(record[0]),
            ResultsWorkbook._normalized(record[1]),
        )

    @staticmethod
    def _animal_sort_key(cage: object, subject: object) -> tuple[int, int | str, int, int | str]:
        def component(value: object) -> tuple[int, int | str]:
            text = ResultsWorkbook._normalized(value)
            return (0, int(text)) if text.isdigit() else (1, text.casefold())

        return (*component(cage), *component(subject))

    @staticmethod
    def _last_table_column(sheet, title_row: int, start_column: int) -> int:
        header_row = title_row + 2
        last_column = start_column + 1
        for column in range(start_column + 2, sheet.max_column + 1):
            header = str(sheet.cell(row=header_row, column=column).value or "").strip()
            if not header.upper().startswith("T"):
                break
            last_column = column
        return last_column

    @staticmethod
    def _table_last_row(sheet, title_row: int, start_column: int) -> int:
        row = title_row + 3
        while (
            sheet.cell(row=row, column=start_column).value is not None
            or sheet.cell(row=row, column=start_column + 1).value is not None
        ):
            row += 1
        return max(title_row + 2, row - 1)

    @staticmethod
    def _trial_columns(
        sheet,
        title_row: int,
        start_column: int,
    ) -> dict[tuple[str, int], int]:
        columns: dict[tuple[str, int], int] = {}
        active_day = ""
        header_row = title_row + 2
        last_column = ResultsWorkbook._last_table_column(sheet, title_row, start_column)
        for column in range(start_column + 2, last_column + 1):
            day_value = sheet.cell(row=title_row + 1, column=column).value
            if day_value is not None and str(day_value).strip():
                active_day = str(day_value).strip()
            trial = str(sheet.cell(row=header_row, column=column).value or "").strip().upper()
            if active_day and trial.startswith("T") and trial[1:].isdigit():
                columns[(active_day, int(trial[1:]))] = column
        return columns

    def _order_forelimb_day_columns(self, sheet) -> None:
        """Keep TIME and DISTANCE days aligned in chronological column order."""
        time_columns = self._trial_columns(sheet, 1, 1)
        distance_columns = self._trial_columns(sheet, 1, DISTANCE_TABLE_COLUMN)
        days = sorted(
            {day for day, _trial in time_columns} | {day for day, _trial in distance_columns},
            key=natural_day_key,
        )
        self._order_trial_table_columns(sheet, 1, 1, days)
        self._order_trial_table_columns(sheet, 1, DISTANCE_TABLE_COLUMN, days)

    def _order_trial_table_columns(
        self,
        sheet,
        title_row: int,
        start_column: int,
        days: list[str],
    ) -> None:
        """Rewrite one Forelimb table's trial columns without changing values."""
        old_columns = self._trial_columns(sheet, title_row, start_column)
        old_last_column = self._last_table_column(sheet, title_row, start_column)
        first_data_row = title_row + 3
        last_data_row = self._table_last_row(sheet, title_row, start_column)

        records: list[tuple[object, object, dict[tuple[str, int], object]]] = []
        for row in range(first_data_row, last_data_row + 1):
            cage = sheet.cell(row=row, column=start_column).value
            animal = sheet.cell(row=row, column=start_column + 1).value
            if cage is None and animal is None:
                continue
            records.append(
                (
                    cage,
                    animal,
                    {
                        key: sheet.cell(row=row, column=column).value
                        for key, column in old_columns.items()
                    },
                )
            )

        new_last_column = start_column + 1 + (len(days) * 3)
        clear_last_column = max(old_last_column, new_last_column)
        for row in range(title_row + 1, last_data_row + 1):
            for column in range(start_column + 2, clear_last_column + 1):
                sheet.cell(row=row, column=column).value = None

        for day_index, day in enumerate(days):
            first_column = start_column + 2 + (day_index * 3)
            sheet.cell(row=title_row + 1, column=first_column, value=day)
            for trial in (1, 2, 3):
                sheet.cell(row=title_row + 2, column=first_column + trial - 1, value=f"T{trial}")

        for row, (cage, animal, values) in enumerate(
            sorted(records, key=lambda item: self._animal_sort_key(item[0], item[1])),
            start=first_data_row,
        ):
            sheet.cell(row=row, column=start_column, value=cage)
            sheet.cell(row=row, column=start_column + 1, value=animal)
            for day_index, day in enumerate(days):
                first_column = start_column + 2 + (day_index * 3)
                for trial in (1, 2, 3):
                    cell = sheet.cell(row=row, column=first_column + trial - 1)
                    cell.value = values.get((day, trial))
                    cell.number_format = "0.00"

    @staticmethod
    def _table_rows_by_subject(
        sheet,
        title_row: int,
        start_column: int,
    ) -> dict[tuple[str, str], int]:
        first_row = title_row + 3
        last_row = ResultsWorkbook._table_last_row(sheet, title_row, start_column)
        return {
            (
                ResultsWorkbook._normalized(sheet.cell(row=row, column=start_column).value),
                ResultsWorkbook._normalized(sheet.cell(row=row, column=start_column + 1).value),
            ): row
            for row in range(first_row, last_row + 1)
        }

    @staticmethod
    def _numeric(value: object, *, allow_zero: bool = False) -> float | None:
        try:
            number = float(value)
        except (TypeError, ValueError):
            return None
        if number > 0 or (allow_zero and number == 0):
            return number
        return None

    @staticmethod
    def _clear_speed_tables(sheet) -> None:
        title_rows = [
            row
            for row in range(1, sheet.max_row + 1)
            if str(sheet.cell(row=row, column=1).value or "").strip().upper()
            == "SPEED (CM/S)"
        ]
        for title_row in title_rows:
            last_row = ResultsWorkbook._table_last_row(sheet, title_row, 1)
            last_column = ResultsWorkbook._last_table_column(sheet, title_row, 1)
            for row in range(title_row, last_row + 1):
                for column in range(1, last_column + 1):
                    sheet.cell(row=row, column=column).value = None

    @staticmethod
    def _speed_title_row(sheet) -> int:
        return max(
            SPEED_TABLE_ROW,
            ResultsWorkbook._table_last_row(sheet, 1, 1) + 3,
        )

    def _rebuild_speed_table(self, sheet) -> None:
        speed_title_row = self._speed_title_row(sheet)
        self._ensure_table_headers(sheet, "SPEED (cm/s)", speed_title_row, 1)

        time_columns = self._trial_columns(sheet, 1, 1)
        distance_columns = self._trial_columns(sheet, 1, DISTANCE_TABLE_COLUMN)
        time_rows = self._table_rows_by_subject(sheet, 1, 1)
        distance_rows = self._table_rows_by_subject(sheet, 1, DISTANCE_TABLE_COLUMN)

        for column in range(3, self._last_table_column(sheet, 1, 1) + 1):
            sheet.cell(
                row=speed_title_row + 1,
                column=column,
                value=sheet.cell(row=2, column=column).value,
            )
            sheet.cell(
                row=speed_title_row + 2,
                column=column,
                value=sheet.cell(row=3, column=column).value,
            )

        for offset, (subject_key, time_row) in enumerate(time_rows.items()):
            speed_row = speed_title_row + 3 + offset
            cage, subject = subject_key
            sheet.cell(row=speed_row, column=1, value=sheet.cell(row=time_row, column=1).value)
            sheet.cell(row=speed_row, column=2, value=sheet.cell(row=time_row, column=2).value)

            distance_row = distance_rows.get((cage, subject))
            for trial_key, time_column in time_columns.items():
                time_value = self._numeric(
                    sheet.cell(row=time_row, column=time_column).value
                )
                distance_column = distance_columns.get(trial_key)
                distance_value = (
                    self._numeric(
                        sheet.cell(row=distance_row, column=distance_column).value,
                        allow_zero=True,
                    )
                    if distance_row is not None and distance_column is not None
                    else None
                )
                cell = sheet.cell(row=speed_row, column=time_column)
                cell.value = (
                    round(distance_value / time_value, 2)
                    if time_value is not None and distance_value is not None
                    else None
                )
                cell.number_format = "0.00"

    @staticmethod
    def _trial_column(
        sheet,
        day: str,
        trial: int,
        title_row: int,
        start_column: int,
    ) -> int:
        header_row = title_row + 2
        last_column = ResultsWorkbook._last_table_column(sheet, title_row, start_column)

        active_day = ""
        for column in range(start_column + 2, last_column + 1):
            day_value = sheet.cell(row=title_row + 1, column=column).value
            if day_value is not None and str(day_value).strip():
                active_day = str(day_value).strip()
            trial_value = str(sheet.cell(row=header_row, column=column).value or "").strip().upper()
            if active_day.upper() == day.upper() and trial_value == f"T{trial}":
                return column

        first_column = max(start_column + 2, last_column + 1)
        for offset, trial_number in enumerate((1, 2, 3)):
            column = first_column + offset
            if trial_number == 1:
                sheet.cell(row=title_row + 1, column=column, value=day)
            sheet.cell(row=header_row, column=column, value=f"T{trial_number}")
        return first_column + trial - 1

    def _write_measurement(
        self,
        sheet,
        title_row: int,
        start_column: int,
        annotation: TrialAnnotation,
        value: float | int | None,
    ) -> None:
        row = self._animal_row(
            sheet,
            annotation.group,
            annotation.subject,
            title_row,
            start_column,
        )
        column = self._trial_column(
            sheet,
            annotation.day,
            annotation.trial,
            title_row,
            start_column,
        )
        cell = sheet.cell(
            row=row,
            column=column,
            value=round(value, 2) if value is not None else None,
        )
        cell.number_format = "0.00"

    @staticmethod
    def _audit_sheet(workbook):
        if AUDIT_SHEET in workbook.sheetnames:
            sheet = workbook[AUDIT_SHEET]
            existing_headers = {
                sheet.cell(row=1, column=column).value
                for column in range(1, sheet.max_column + 1)
            }
            from .annotations import TrialAnnotation

            next_column = sheet.max_column + 1
            for name in TrialAnnotation.__dataclass_fields__:
                if name not in existing_headers:
                    sheet.cell(row=1, column=next_column, value=name)
                    next_column += 1
            return sheet

        sheet = workbook.create_sheet(AUDIT_SHEET)
        from .annotations import TrialAnnotation

        for column, name in enumerate(TrialAnnotation.__dataclass_fields__, start=1):
            sheet.cell(row=1, column=column, value=name)
        sheet.freeze_panes = "A2"
        return sheet

    @staticmethod
    def _upsert_audit_row(sheet, annotation: TrialAnnotation) -> None:
        from dataclasses import asdict

        values = asdict(annotation)
        headers = [
            sheet.cell(row=1, column=column).value
            for column in range(1, sheet.max_column + 1)
        ]
        relative_video_column = headers.index("relative_video") + 1
        target_row = next(
            (
                row
                for row in range(2, sheet.max_row + 1)
                if sheet.cell(row=row, column=relative_video_column).value
                == annotation.relative_video
            ),
            sheet.max_row + 1,
        )
        for column, name in enumerate(headers, start=1):
            sheet.cell(row=target_row, column=column, value=values.get(name, ""))
